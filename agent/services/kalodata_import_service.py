"""Kalodata/External catalog staged import service.

Parses the Owner's Kalodata/Fastmoss merged workbook ("MERGED PRODUCTS" +
"COPYWRITING HUB" sheets) into two staged JSON files inside OPERATOR_PACK_DIR:

  * KALODATA_REFERENCE_CATALOG.json — reference-seed records (the SAME shape
    and `fastmoss-ref:` id scheme as the singleton workbook rows, built through
    the existing `_reference_seed` helper) that
    `list_fastmoss_reference_products()` unions into the catalog, so every
    downstream consumer (bulk queue sync, drafts, approval, readiness) works
    unchanged.
  * KALODATA_HUB_ENRICHMENT.json — per-reference enrichment payloads in the
    exact item shape `fastmoss_bulk_promotion_service.import_enrichment`
    already accepts.

Guarantees: zero AI calls, zero direct `product` table writes, idempotent
re-import (atomic file rewrite; queue sync stays INSERT OR IGNORE), never
mutates the singleton workbook.
"""
from __future__ import annotations

import asyncio
import json
import logging
import os
import re
import hashlib
from io import BytesIO
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from types import SimpleNamespace
from typing import Any

from agent.config import BASE_DIR, OPERATOR_PACK_DIR
from agent.models.kalodata_import import (
    CopyIntelligenceDryRunReport,
    CopyIntelligenceSourceRow,
    KalodataHubRow,
    KalodataImportReport,
    KalodataMergedRow,
)
from agent.services.fastmoss_product_reference_service import _reference_seed

logger = logging.getLogger(__name__)

STAGED_CATALOG_FILENAME = "KALODATA_REFERENCE_CATALOG.json"
STAGED_HUB_FILENAME = "KALODATA_HUB_ENRICHMENT.json"
COPY_INTELLIGENCE_WORKBOOKS_DIR = BASE_DIR / "data" / "fastmoss" / "imports" / "copy_intelligence"

MERGED_SHEET = "MERGED PRODUCTS"
HUB_SHEET = "COPYWRITING HUB"


def store_copy_intelligence_workbook_upload(
    *, original_filename: str, payload: bytes
) -> dict[str, str | list[str]]:
    """Store an unmodified full workbook after required-sheet validation."""
    if Path(original_filename).suffix.lower() != ".xlsx":
        raise ValueError("XLSX_REQUIRED")
    if not payload:
        raise ValueError("WORKBOOK_EMPTY")

    import openpyxl

    try:
        workbook = openpyxl.load_workbook(BytesIO(payload), read_only=True, data_only=False)
    except Exception as exc:  # noqa: BLE001 — malformed workbook is never stored
        raise ValueError("INVALID_XLSX") from exc
    try:
        sheet_names = list(workbook.sheetnames)
    finally:
        workbook.close()

    required_sheets = [HUB_SHEET, MERGED_SHEET]
    missing = [sheet for sheet in required_sheets if sheet not in sheet_names]
    if missing:
        raise ValueError(f"MISSING_REQUIRED_SHEETS:{','.join(missing)}")

    fingerprint = hashlib.sha256(payload).hexdigest()
    destination = COPY_INTELLIGENCE_WORKBOOKS_DIR / f"{fingerprint}.xlsx"
    destination.parent.mkdir(parents=True, exist_ok=True)
    if not destination.exists():
        destination.write_bytes(payload)
    return {
        "source_id": fingerprint,
        "original_filename": Path(original_filename).name,
        "fingerprint": fingerprint,
        "sheet_names": sheet_names,
        "required_sheets": required_sheets,
    }


def resolve_copy_intelligence_workbook_source(source_id: str) -> Path:
    """Resolve an upload fingerprint without accepting a caller filesystem path."""
    if not re.fullmatch(r"[0-9a-f]{64}", source_id or ""):
        raise ValueError("INVALID_SOURCE_ID")
    path = COPY_INTELLIGENCE_WORKBOOKS_DIR / f"{source_id}.xlsx"
    if not path.is_file():
        raise FileNotFoundError(source_id)
    return path

# 19-digit TikTok product ids exceed float64's 53-bit mantissa — an Excel
# numeric cell is unrecoverable; the URL is the only lossless source.
_TIKTOK_ID_URL_PATTERNS = (
    re.compile(r"/pdp/(\d{6,})"),
    re.compile(r"/product/(\d{6,})"),
    re.compile(r"[?&](?:pid|product_id)=(\d{6,})"),
)
_FLOAT53_MAX = float(2**53)

_PRICE_NUMBER_RE = re.compile(r"(\d+(?:[.,]\d+)?)")

# Excel serial day 0 (1900 date system, with the standard leap-year quirk
# already absorbed by the -2 offset convention used below).
_EXCEL_EPOCH = date(1899, 12, 30)


def staged_catalog_path() -> Path:
    return Path(OPERATOR_PACK_DIR) / STAGED_CATALOG_FILENAME


def staged_hub_path() -> Path:
    return Path(OPERATOR_PACK_DIR) / STAGED_HUB_FILENAME


def _clean(value: Any) -> str:
    return str(value or "").strip()


def extract_tiktok_product_id(url: Any, cell: Any) -> tuple[str | None, str]:
    """Recover the TikTok product id. URL wins (lossless); a numeric cell is
    trusted only when it is an exact integer below float53 precision."""
    url_text = _clean(url)
    for pattern in _TIKTOK_ID_URL_PATTERNS:
        match = pattern.search(url_text)
        if match:
            return match.group(1), "URL"
    if cell is None or _clean(cell) == "":
        return None, "NONE"
    if isinstance(cell, str):
        digits = _clean(cell)
        if digits.isdigit():
            return digits, "CELL"
        return None, "LOW"
    try:
        value = float(cell)
    except (TypeError, ValueError):
        return None, "LOW"
    if value.is_integer() and 0 < value < _FLOAT53_MAX:
        return str(int(value)), "CELL"
    return None, "LOW"


def parse_price(raw: Any) -> tuple[float | None, float | None, float | None, str | None, bool]:
    """Parse "RM26.50" / "RM6.12 - 19.01" / 26.5 → (mid, min, max, raw, is_range)."""
    if raw is None or _clean(raw) == "":
        return None, None, None, None, False
    if isinstance(raw, (int, float)):
        value = float(raw)
        return value, value, value, str(raw), False
    text = _clean(raw)
    numbers = [float(n.replace(",", ".")) for n in _PRICE_NUMBER_RE.findall(text)]
    if not numbers:
        return None, None, None, text, False
    if len(numbers) == 1:
        return numbers[0], numbers[0], numbers[0], text, False
    low, high = min(numbers[:2]), max(numbers[:2])
    midpoint = round((low + high) / 2, 2)
    return midpoint, low, high, text, True


def excel_date_to_iso(value: Any) -> str | None:
    if value is None or _clean(value) == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        serial = float(value)
        if 1 <= serial <= 200_000:
            return (_EXCEL_EPOCH + timedelta(days=int(serial))).isoformat()
        return None
    text = _clean(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _to_int(value: Any) -> int | None:
    try:
        if value is None or _clean(value) == "":
            return None
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _to_float(value: Any) -> float | None:
    try:
        if value is None or _clean(value) == "":
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def _header_index(header_row: tuple, *candidates: str) -> int | None:
    normalized = [(_clean(h).lower()) for h in header_row]
    for candidate in candidates:
        want = candidate.lower()
        for idx, cell in enumerate(normalized):
            if cell.startswith(want):
                return idx
    return None


def parse_workbook(path: Path) -> tuple[list[KalodataMergedRow], list[KalodataHubRow]]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    merged_rows: list[KalodataMergedRow] = []
    hub_rows: list[KalodataHubRow] = []

    if MERGED_SHEET in workbook.sheetnames:
        sheet = workbook[MERGED_SHEET]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header:
            col = {
                "no": _header_index(header, "No"),
                "sumber": _header_index(header, "Sumber"),
                "name": _header_index(header, "Product Name"),
                "image": _header_index(header, "Image URL", "img_url"),
                "category": _header_index(header, "Category"),
                "price": _header_index(header, "Price"),
                "launch": _header_index(header, "Launch Date"),
                "rating": _header_index(header, "Product Rating"),
                "sold": _header_index(header, "Item Sold"),
                "avg_price": _header_index(header, "Avg Unit Price", "Avg. Unit Price"),
                "commission": _header_index(header, "Commission"),
                "creators": _header_index(header, "Creator Number"),
                "conversion": _header_index(header, "Creator Conversion", "Creator Sales"),
                "tiktok_url": _header_index(header, "TikTok URL", "TikTokUrl"),
                "product_id": _header_index(header, "Product ID"),
                "source_url": _header_index(header, "Source URL", "KalodataUrl"),
            }

            def cell(row: tuple, key: str) -> Any:
                idx = col.get(key)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            for row in rows:
                name = _clean(cell(row, "name"))
                if not name:
                    continue
                price_mid, price_min, price_max, price_raw, _ = parse_price(cell(row, "price"))
                tiktok_url = _clean(cell(row, "tiktok_url")) or None
                product_id, confidence = extract_tiktok_product_id(
                    tiktok_url, cell(row, "product_id")
                )
                merged_rows.append(
                    KalodataMergedRow(
                        row_no=_to_int(cell(row, "no")) or (len(merged_rows) + 1),
                        sumber=_clean(cell(row, "sumber")) or "KALODATA",
                        product_name=name,
                        image_url=_clean(cell(row, "image")) or None,
                        category_path=_clean(cell(row, "category")) or None,
                        price=price_mid,
                        price_min=price_min,
                        price_max=price_max,
                        price_raw=price_raw,
                        launch_date=excel_date_to_iso(cell(row, "launch")),
                        rating=_to_float(cell(row, "rating")),
                        sold_count=_to_int(cell(row, "sold")),
                        avg_unit_price=_to_float(cell(row, "avg_price")),
                        commission_rate=_clean(cell(row, "commission")) or None,
                        creator_number=_to_int(cell(row, "creators")),
                        conversion=_clean(cell(row, "conversion")) or None,
                        tiktok_product_url=tiktok_url,
                        tiktok_product_id=product_id,
                        tiktok_product_id_confidence=confidence,
                        source_url=_clean(cell(row, "source_url")) or None,
                    )
                )

    if HUB_SHEET in workbook.sheetnames:
        sheet = workbook[HUB_SHEET]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header:
            col = {
                "no": _header_index(header, "No"),
                "product_id": _header_index(header, "Product ID"),
                "name": _header_index(header, "Product Name"),
                "product_type": _header_index(header, "Product Type"),
                "category": _header_index(header, "Category"),
                "target_avatar": _header_index(header, "Target Avatar"),
                "pain_point": _header_index(header, "Pain Point"),
                "emotion": _header_index(header, "Emotion"),
                "dream": _header_index(header, "Dream Outcome"),
                "ingredient": _header_index(header, "Key Ingredient"),
                "main_benefit": _header_index(header, "Main Benefit"),
                "secondary_benefit": _header_index(header, "Secondary Benefit"),
                "usp": _header_index(header, "USP"),
                "hook_type": _header_index(header, "Hook Type"),
            }

            def hub_cell(row: tuple, key: str) -> str | None:
                idx = col.get(key)
                if idx is None or idx >= len(row):
                    return None
                return _clean(row[idx]) or None

            for row in rows:
                name = hub_cell(row, "name") or ""
                if not name:
                    continue
                hub_rows.append(
                    KalodataHubRow(
                        row_no=_to_int(row[col["no"]] if col.get("no") is not None else None)
                        or (len(hub_rows) + 1),
                        product_id=hub_cell(row, "product_id"),
                        product_name=name,
                        product_type=hub_cell(row, "product_type"),
                        category=hub_cell(row, "category"),
                        target_avatar=hub_cell(row, "target_avatar"),
                        pain_point=hub_cell(row, "pain_point"),
                        emotion_trigger=hub_cell(row, "emotion"),
                        dream_outcome=hub_cell(row, "dream"),
                        key_ingredient_feature=hub_cell(row, "ingredient"),
                        main_benefit=hub_cell(row, "main_benefit"),
                        secondary_benefit=hub_cell(row, "secondary_benefit"),
                        usp=hub_cell(row, "usp"),
                        hook_type=hub_cell(row, "hook_type"),
                    )
                )

    workbook.close()
    return merged_rows, hub_rows


def _split_category(category_path: str | None) -> tuple[str | None, str | None]:
    if not category_path:
        return None, None
    parts = [p.strip() for p in category_path.split(">") if p.strip()]
    category = parts[0] if parts else None
    subcategory = parts[1] if len(parts) > 1 else None
    return category, subcategory


def build_staged_record(row: KalodataMergedRow, *, source_file: str) -> dict[str, Any]:
    """Build a reference-seed record through the EXISTING `_reference_seed`
    helper so the id scheme and record shape can never drift from the
    singleton-workbook rows."""
    category, subcategory = _split_category(row.category_path)
    seed_input = SimpleNamespace(
        raw_product_title=row.product_name,
        product_display_name=row.product_name,
        product_short_name=None,
        source_url=row.source_url or "",
        tiktok_product_url=row.tiktok_product_url or "",
        category=category,
        sub_category=subcategory,
        type_angle=None,
        product_type=None,
        silo_id=None,
        trigger_id=None,
        submode_formula=None,
        mode_recommendations=None,
        copywriting_angle=None,
        claim_risk_level=None,
        mapping_source=None,
        mapping_confidence=None,
        shop_name=None,
        avg_price_rm=row.price,
        commission_amount=None,
        commission_rate=row.commission_rate,
        image_url=row.image_url,
    )
    record = _reference_seed(seed_input)
    record["source"] = row.sumber or "KALODATA"
    record["fastmoss_source_file"] = source_file
    record["kalodata_meta"] = {
        "row_no": row.row_no,
        "launch_date": row.launch_date,
        "rating": row.rating,
        "sold_count": row.sold_count,
        "avg_unit_price": row.avg_unit_price,
        "creator_number": row.creator_number,
        "conversion": row.conversion,
        "tiktok_product_id": row.tiktok_product_id,
        "tiktok_product_id_confidence": row.tiktok_product_id_confidence,
        "price_min": row.price_min,
        "price_max": row.price_max,
        "price_raw": row.price_raw,
    }
    return record


def _normalize_name(name: str) -> str:
    return re.sub(r"\s+", " ", _clean(name)).lower()


_COPY_HUB_HEADERS = {
    "product_name": ("Product Name",),
    "target_avatar": ("Target Avatar",),
    "pain_point": ("Pain Point",),
    "emotion_trigger": ("Emotion/Trigger",),
    "dream_outcome": ("Dream Outcome",),
    "key_ingredients_features": ("Key Feature", "Key Ingredient/Feature"),
    "hook_type": ("Hook Type",),
    "hook_script": ("Hook Script",),
    "body_script": ("Body Script",),
    "cta_type": ("CTA Type",),
    "cta_script": ("CTA Script",),
    "tone": ("Tone",),
    "pronoun": ("Pronoun",),
    "copy_angle": ("Copy Angle",),
}


def _copy_hub_header_index(header: tuple[Any, ...], candidates: tuple[str, ...]) -> int | None:
    normalized = [_clean(value).casefold() for value in header]
    for candidate in candidates:
        candidate_normalized = candidate.casefold()
        for index, value in enumerate(normalized):
            if value.startswith(candidate_normalized):
                return index
    return None


def parse_copy_intelligence_hub(path: Path) -> list[dict[str, str | int | None]]:
    """Read COPYWRITING HUB without altering the legacy staging parser."""
    import openpyxl

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if HUB_SHEET not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"SHEET_NOT_FOUND:{HUB_SHEET}")
    sheet = workbook[HUB_SHEET]
    iterator = sheet.iter_rows(values_only=True)
    header = next(iterator, None)
    if not header:
        workbook.close()
        return []
    indexes = {key: _copy_hub_header_index(header, candidates) for key, candidates in _COPY_HUB_HEADERS.items()}
    rows: list[dict[str, str | int | None]] = []
    for source_row, values in enumerate(iterator, start=2):
        def value(key: str) -> str | None:
            index = indexes[key]
            return _clean(values[index]) or None if index is not None and index < len(values) else None
        product_name = value("product_name")
        if not product_name:
            continue
        rows.append({"source_row": source_row, "source_product_name": product_name, **{key: value(key) for key in indexes if key != "product_name"}})
    workbook.close()
    return rows


def build_copy_intelligence_dry_run(
    source_path: str | Path,
    product_id_by_tiktok_id: dict[str, str | list[str]] | None = None,
) -> CopyIntelligenceDryRunReport:
    """Build an auditable, no-write COPYWRITING HUB seed plan.

    Workbook numeric IDs are intentionally excluded: they are float-lossy. A
    unique normalized name may identify the workbook reference at MEDIUM only;
    duplicate names are quarantined and no row is auto-approved.
    """
    path = Path(source_path)
    if not path.exists():
        raise FileNotFoundError(str(path))
    merged_rows, _ = parse_workbook(path)
    source_rows = parse_copy_intelligence_hub(path)
    report = CopyIntelligenceDryRunReport(source_workbook=path.name)
    known_product_ids = product_id_by_tiktok_id or {}
    report.total_source_rows = len(source_rows)
    report.examples = {key: [] for key in ("medium", "quarantined", "blank", "suspicious")}
    merged_by_name: dict[str, list[KalodataMergedRow]] = {}
    hub_name_count: dict[str, int] = {}
    for merged in merged_rows:
        merged_by_name.setdefault(_normalize_name(merged.product_name), []).append(merged)
    for row in source_rows:
        name = _normalize_name(str(row["source_product_name"]))
        hub_name_count[name] = hub_name_count.get(name, 0) + 1
    source_fingerprints: set[str] = set()
    target_counts: dict[str, int] = {}
    corruption_probe = [KalodataHubRow(
        row_no=int(row["source_row"]), product_name=str(row["source_product_name"]),
        target_avatar=row.get("target_avatar"), pain_point=row.get("pain_point"),
        emotion_trigger=row.get("emotion_trigger"), dream_outcome=row.get("dream_outcome"),
        key_ingredient_feature=row.get("key_ingredients_features"),
        main_benefit=row.get("body_script"), usp=row.get("hook_script"),
    ) for row in source_rows]
    suspicious_rows = find_hub_internal_corruption(corruption_probe)
    for row in source_rows:
        copy_values = [value for key, value in row.items() if key not in {"source_row", "source_product_name"} and value]
        if not copy_values:
            report.blank_no_copy_rows += 1
            report.examples["blank"].append({"source_row": int(row["source_row"]), "product": str(row["source_product_name"])})
            continue
        report.usable_rows += 1
        normalized_name = _normalize_name(str(row["source_product_name"]))
        candidates = merged_by_name.get(normalized_name, [])
        duplicate_name = hub_name_count[normalized_name] > 1 or len(candidates) != 1
        match_method = "UNMATCHED"
        confidence = "LOW"
        reference_id: str | None = None
        target_product_id: str | None = None
        quarantine_reason: str | None = None
        duplicate_product_ids: list[str] = []
        if len(candidates) == 1 and not duplicate_name:
            reference_id = build_staged_record(candidates[0], source_file=path.name)["id"]
            source_tid = candidates[0].tiktok_product_id
            raw_product_ids = known_product_ids.get(source_tid or "")
            product_ids = [raw_product_ids] if isinstance(raw_product_ids, str) else (raw_product_ids or [])
            if len(product_ids) == 1:
                target_product_id = product_ids[0]
                match_method = "TIKTOK_PRODUCT_ID_MATCH"
                confidence = "HIGH"
                report.matched_high_confidence += 1
            elif len(product_ids) > 1:
                match_method = "DUPLICATE_PRODUCT_TRUTH_TIKTOK_ID"
                duplicate_product_ids = sorted(product_ids)
                quarantine_reason = match_method
                report.low_confidence_quarantined += 1
                report.examples["quarantined"].append({"source_row": int(row["source_row"]), "product": str(row["source_product_name"]), "reason": quarantine_reason, "candidate_product_ids": ",".join(duplicate_product_ids)})
            else:
                match_method = "UNIQUE_NORMALIZED_NAME_TO_SOURCE_REFERENCE"
                confidence = "MEDIUM"
                report.matched_medium_confidence += 1
            target_counts[reference_id] = target_counts.get(reference_id, 0) + 1
        elif not candidates:
            report.unmatched += 1
        else:
            match_method = "AMBIGUOUS_NORMALIZED_NAME"
            report.low_confidence_quarantined += 1
            report.duplicates += max(0, hub_name_count[normalized_name] - 1)
            report.examples["quarantined"].append({"source_row": int(row["source_row"]), "product": str(row["source_product_name"]), "reason": "AMBIGUOUS_NORMALIZED_NAME"})
        if int(row["source_row"]) in suspicious_rows:
            report.suspicious_cross_product_copy += 1
            report.examples["suspicious"].append({"source_row": int(row["source_row"]), "product": str(row["source_product_name"])})
        fingerprint_input = f"{path.name}|{HUB_SHEET}|{row['source_row']}|{reference_id or normalized_name}"
        fingerprint = hashlib.sha256(fingerprint_input.encode("utf-8")).hexdigest()
        if fingerprint in source_fingerprints:
            report.duplicates += 1
            continue
        source_fingerprints.add(fingerprint)
        record = CopyIntelligenceSourceRow(
            source_workbook=path.name, source_sheet=HUB_SHEET, source_row=int(row["source_row"]),
            source_product_name=str(row["source_product_name"]), reference_id=reference_id,
            target_product_id=target_product_id,
            target_avatar=row.get("target_avatar"), pain_point=row.get("pain_point"),
            emotion_trigger=row.get("emotion_trigger"), dream_outcome=row.get("dream_outcome"),
            key_ingredients_features=row.get("key_ingredients_features"), hook_type=row.get("hook_type"),
            hook_script=row.get("hook_script"), body_script=row.get("body_script"),
            cta_type=row.get("cta_type"), cta_script=row.get("cta_script"), tone=row.get("tone"),
            pronoun=row.get("pronoun"), copy_angle=row.get("copy_angle"),
            match_method=match_method, confidence=confidence, source_fingerprint=fingerprint,
            provenance={"workbook": path.name, "sheet": HUB_SHEET, "source_row": str(row["source_row"]), "match_method": match_method, **({"quarantine_reason": quarantine_reason, "duplicate_product_ids": ",".join(duplicate_product_ids)} if quarantine_reason else {})},
        )
        report.records.append(record)
        if confidence == "MEDIUM" and len(report.examples["medium"]) < 3:
            report.examples["medium"].append({"source_row": record.source_row, "product": record.source_product_name, "reference_id": reference_id or ""})
    report.conflicts = sum(count - 1 for count in target_counts.values() if count > 1)
    return report


async def persist_copy_intelligence_seed_records(records: list[CopyIntelligenceSourceRow]) -> dict[str, int]:
    """Explicit write primitive for an owner-authorized seed only.

    No route calls this function. It writes review-only ledger records and
    deliberately cannot mutate Product Truth or approved Copy Sets.
    """
    from agent.db import crud

    created_or_existing = skipped_quarantined = skipped_low_confidence = skipped_unmatched = 0
    for record in records:
        if record.provenance.get("quarantine_reason"):
            skipped_quarantined += 1
            continue
        if record.match_method == "UNMATCHED":
            skipped_unmatched += 1
            continue
        if record.confidence == "LOW":
            skipped_low_confidence += 1
            continue
        row = await crud.create_copy_intelligence_seed(
            source_fingerprint=record.source_fingerprint,
            source_workbook=record.source_workbook, source_sheet=record.source_sheet,
            source_row=record.source_row, source_product_name=record.source_product_name,
            reference_id=record.reference_id, target_product_id=record.target_product_id,
            match_method=record.match_method,
            confidence=record.confidence, status="NEEDS_REVIEW",
            target_avatar=record.target_avatar, pain_point=record.pain_point,
            emotion_trigger=record.emotion_trigger, dream_outcome=record.dream_outcome,
            key_ingredients_features=record.key_ingredients_features, hook_type=record.hook_type,
            hook_script=record.hook_script, body_script=record.body_script,
            cta_type=record.cta_type, cta_script=record.cta_script, tone=record.tone,
            pronoun=record.pronoun, copy_angle=record.copy_angle,
            provenance_json=json.dumps(record.provenance, ensure_ascii=False, sort_keys=True),
        )
        if row.get("created_at"):
            created_or_existing += 1
    return {"records_processed": len(records), "created_or_existing": created_or_existing, "skipped_quarantined": skipped_quarantined, "skipped_low_confidence": skipped_low_confidence, "skipped_unmatched": skipped_unmatched, "status": "NEEDS_REVIEW_ONLY"}


async def list_copy_intelligence_seed_records(
    *, confidence: str | None = None, status: str | None = None,
    search: str | None = None, limit: int = 100,
) -> dict[str, object]:
    """Read the immutable review ledger without invoking the seed primitive."""
    from agent.db import crud

    total, rows = await crud.list_copy_intelligence_seeds(
        confidence=confidence, status=status, search=search, limit=limit,
    )
    items = []
    for row in rows:
        provenance = row.pop("provenance_json", "{}")
        try:
            row["provenance"] = json.loads(provenance or "{}")
        except json.JSONDecodeError:
            row["provenance"] = {}
        items.append(row)
    return {"total": total, "items": items}


# ── Copy Intelligence seed review/approval layer ──────────────────────────────
# One owner reviews ONE persisted ledger row at a time. Approval/rejection is
# fail-closed and audit-stamped; it NEVER exposes a row to generation. Nothing
# downstream (compiler, DeepSeek, Product Truth, Copy Set) reads these rows —
# this layer only transitions status + records who/when/why.
APPROVE_COPY_INTELLIGENCE_PHRASE = "APPROVE COPY INTELLIGENCE"
APPROVE_MEDIUM_COPY_INTELLIGENCE_PHRASE = "APPROVE MEDIUM CONFIDENCE COPY INTELLIGENCE"
REJECT_COPY_INTELLIGENCE_PHRASE = "REJECT COPY INTELLIGENCE"
_REVIEWABLE_SOURCE_STATUS = "NEEDS_REVIEW"


class CopyIntelligenceReviewError(Exception):
    """Fail-closed review error. Callers (the API) surface it verbatim; a review
    that cannot be validated NEVER writes a status transition."""

    def __init__(self, code: str, status_code: int = 422, detail: object = None):
        super().__init__(code)
        self.code = code
        self.status_code = status_code
        self.detail = detail


async def review_copy_intelligence_seed(
    seed_id: str,
    *,
    action: str,
    reviewed_by: str,
    review_note: str,
    confirmation_phrase: str,
) -> dict[str, object]:
    """Approve or reject ONE ledger row. Fail-closed on every guard.

    Allowed transitions are ONLY ``NEEDS_REVIEW -> APPROVED`` and
    ``NEEDS_REVIEW -> REJECTED``. MEDIUM-confidence approval requires the
    stronger confirmation phrase. This function writes only the seed row's
    review audit columns; it never creates a snapshot or Copy Set, never calls
    an AI provider, and never routes the row to generation.
    """
    from agent.db import crud

    decision = (action or "").strip().upper()
    if decision not in ("APPROVE", "REJECT"):
        raise CopyIntelligenceReviewError("INVALID_REVIEW_ACTION", 422, {"action": action})

    reviewer = (reviewed_by or "").strip()
    if not reviewer:
        raise CopyIntelligenceReviewError("REVIEWER_REQUIRED", 422)

    note = (review_note or "").strip()
    if not note:
        raise CopyIntelligenceReviewError("REVIEW_NOTE_REQUIRED", 422)

    row = await crud.get_copy_intelligence_seed(seed_id)
    if not row:
        raise CopyIntelligenceReviewError(
            "COPY_INTELLIGENCE_SEED_NOT_FOUND", 404, {"seed_id": seed_id}
        )

    current_status = row.get("status")
    if current_status != _REVIEWABLE_SOURCE_STATUS:
        raise CopyIntelligenceReviewError(
            "INVALID_SOURCE_STATUS",
            409,
            {"seed_id": seed_id, "status": current_status},
        )

    confidence = row.get("confidence")
    phrase = (confirmation_phrase or "").strip()

    if decision == "APPROVE":
        if confidence == "MEDIUM":
            required_phrase = APPROVE_MEDIUM_COPY_INTELLIGENCE_PHRASE
            if phrase == APPROVE_COPY_INTELLIGENCE_PHRASE:
                # Normal phrase on a MEDIUM row is explicitly not enough.
                raise CopyIntelligenceReviewError(
                    "MEDIUM_CONFIDENCE_PHRASE_REQUIRED",
                    422,
                    {"required_phrase": required_phrase},
                )
        else:
            required_phrase = APPROVE_COPY_INTELLIGENCE_PHRASE
        if phrase != required_phrase:
            raise CopyIntelligenceReviewError(
                "CONFIRMATION_PHRASE_MISMATCH", 422, {"required_phrase": required_phrase}
            )
        new_status = "APPROVED"
    else:
        if phrase != REJECT_COPY_INTELLIGENCE_PHRASE:
            raise CopyIntelligenceReviewError(
                "CONFIRMATION_PHRASE_MISMATCH",
                422,
                {"required_phrase": REJECT_COPY_INTELLIGENCE_PHRASE},
            )
        new_status = "REJECTED"

    reviewed_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    await crud.update_copy_intelligence_seed(
        seed_id,
        status=new_status,
        previous_status=current_status,
        review_action=decision,
        review_note=note,
        reviewed_by=reviewer,
        reviewed_at=reviewed_at,
    )
    return {
        "seed_id": seed_id,
        "previous_status": current_status,
        "new_status": new_status,
        "confidence": confidence,
        "reviewed_by": reviewer,
        "reviewed_at": reviewed_at,
        "review_note": note,
    }


async def build_copy_intelligence_dry_run_for_system(
    source_path: str | Path,
) -> CopyIntelligenceDryRunReport:
    """Resolve high-confidence Product Truth links from URL IDs, read-only."""
    from agent.db import crud

    product_id_by_tiktok_id: dict[str, list[str]] = {}
    for product in await crud.list_products(include_archived=True):
        product_id = _clean(product.get("id"))
        tiktok_id, confidence = extract_tiktok_product_id(product.get("tiktok_product_url"), None)
        if product_id and tiktok_id and confidence == "URL":
            product_id_by_tiktok_id.setdefault(tiktok_id, []).append(product_id)
    return await asyncio.to_thread(
        build_copy_intelligence_dry_run, source_path, product_id_by_tiktok_id
    )


def _name_tokens(name: str) -> set[str]:
    return {w.lower() for w in re.findall(r"[A-Za-z]{4,}", str(name or ""))}


def _hub_copy_blob(hub: KalodataHubRow) -> str:
    return " ".join(
        str(v or "")
        for v in (
            hub.target_avatar, hub.pain_point, hub.emotion_trigger,
            hub.dream_outcome, hub.key_ingredient_feature, hub.main_benefit,
            hub.secondary_benefit, hub.usp, hub.hook_type,
        )
    ).lower()


def find_hub_internal_corruption(hub_rows: list[KalodataHubRow]) -> set[int]:
    """Source-corruption guard (live workbook audit 2026-07-14: dozens of HUB
    rows carry ANOTHER product's copy — scattered offsets, unrecoverable
    automatically; e.g. the SZINDORE perfume row held FOCALLURE lip-clay copy).
    A row is corrupt when its copy names a DISTINCTIVE brand token of a
    different product (token appearing in ≤2 product names) while naming none
    of its own. Copy that merely omits the product name stays accepted —
    wrong-brand copy must never apply, but a missing mention is not proof."""
    token_owners: dict[str, set[int]] = {}
    row_tokens: dict[int, set[str]] = {}
    for hub in hub_rows:
        tokens = _name_tokens(hub.product_name)
        row_tokens[hub.row_no] = tokens
        for t in tokens:
            token_owners.setdefault(t, set()).add(hub.row_no)
    distinctive = {t for t, owners in token_owners.items() if len(owners) <= 2}
    corrupt: set[int] = set()
    for hub in hub_rows:
        if not hub.has_any_enrichment():
            continue
        blob = _hub_copy_blob(hub)
        own = row_tokens.get(hub.row_no, set())
        if any(t in blob for t in own):
            continue
        foreign = distinctive - own
        if any(
            not token_owners[t].issubset({hub.row_no})
            and re.search(rf"\b{re.escape(t)}\b", blob)
            for t in foreign
        ):
            corrupt.add(hub.row_no)
    return corrupt


def build_hub_enrichment(
    hub_rows: list[KalodataHubRow],
    merged_rows: list[KalodataMergedRow],
    record_by_row_no: dict[int, dict[str, Any]],
) -> tuple[dict[str, dict[str, str]], int, list[int], list[int]]:
    """Join HUB rows to staged reference ids by NORMALIZED PRODUCT NAME and map
    columns into the exact `import_enrichment` item field names.

    NEVER by row № — the COPYWRITING HUB sheet is ordered differently from
    MERGED PRODUCTS (live workbook audit 2026-07-14: 590/591 № misaligned),
    and a №-first join attached jeans copy to a perfume. The HUB Product ID
    cells are float-lossy, so the name is the only reliable join key."""
    by_name = {
        _normalize_name(row.product_name): record_by_row_no.get(row.row_no)
        for row in merged_rows
        if record_by_row_no.get(row.row_no)
    }
    corrupt_rows = find_hub_internal_corruption(hub_rows)
    enrichment: dict[str, dict[str, str]] = {}
    matched = 0
    unmatched: list[int] = []
    internally_corrupt: list[int] = []
    for hub in hub_rows:
        if not hub.has_any_enrichment():
            continue
        if hub.row_no in corrupt_rows:
            internally_corrupt.append(hub.row_no)
            continue
        record = by_name.get(_normalize_name(hub.product_name))
        if not record:
            unmatched.append(hub.row_no)
            continue
        reference_id = record["id"]
        benefits_parts = [
            p for p in (hub.main_benefit, hub.secondary_benefit, hub.usp) if p
        ]
        knowledge_parts = []
        if hub.pain_point:
            knowledge_parts.append(f"Pain point: {hub.pain_point}")
        if hub.emotion_trigger:
            knowledge_parts.append(f"Emotion trigger: {hub.emotion_trigger}")
        if hub.dream_outcome:
            knowledge_parts.append(f"Dream outcome: {hub.dream_outcome}")
        if hub.hook_type:
            knowledge_parts.append(f"Hook type: {hub.hook_type}")
        item: dict[str, str] = {}
        if hub.target_avatar:
            item["target_customer_text"] = hub.target_avatar
        if benefits_parts:
            item["benefits_text"] = "\n".join(benefits_parts)
        if hub.key_ingredient_feature:
            item["ingredients_text"] = hub.key_ingredient_feature
        if knowledge_parts:
            item["product_knowledge_text"] = "\n".join(knowledge_parts)
        if not item:
            continue
        price = record.get("price")
        if price is not None:
            item["price"] = price
        enrichment[reference_id] = item
        matched += 1
    return enrichment, matched, unmatched, internally_corrupt


def _atomic_write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=1), encoding="utf-8")
    os.replace(tmp, path)


def import_workbook(
    source_path: str | Path,
    existing_tids: set[str] | None = None,
) -> KalodataImportReport:
    """Parse + stage the workbook. Pure parsing — no AI, no DB writes.

    Duplicate law (Owner, 2026-07-14): the TikTok Product ID is the product
    identity. A row whose tid already exists in the SYSTEM (committed product
    rows or the fastmoss reference workbook — supplied via `existing_tids`)
    or earlier in the SAME file is never staged."""
    path = Path(source_path)
    if not path.exists():
        raise FileNotFoundError(str(path))

    merged_rows, hub_rows = parse_workbook(path)
    report = KalodataImportReport(
        source_path=str(path),
        parsed_merged=len(merged_rows),
        parsed_hub=len(hub_rows),
    )

    staged_records: list[dict[str, Any]] = []
    # HUB lookup covers EVERY valid row — including tid-skipped ones. Their
    # reference ids already live in the queue (first import) or as committed
    # products; dropping them from the lookup silently discarded their HUB
    # copy, so Recompute rebuilt those drafts with empty knowledge fields
    # (Owner-reported 2026-07-14).
    hub_lookup_by_row_no: dict[int, dict[str, Any]] = {}
    seen_ids: set[str] = set()
    seen_tids: set[str] = set()
    system_tids = {t for t in (existing_tids or set()) if t}
    for row in merged_rows:
        if not _clean(row.product_name):
            report.skipped_invalid += 1
            continue
        record = build_staged_record(row, source_file=path.name)
        hub_lookup_by_row_no[row.row_no] = record
        tid = row.tiktok_product_id
        if tid and tid in system_tids:
            report.skipped_existing_tid += 1
            continue
        if record["id"] in seen_ids or (tid and tid in seen_tids):
            report.skipped_duplicate_in_file += 1
            continue
        seen_ids.add(record["id"])
        if tid:
            seen_tids.add(tid)
        staged_records.append(record)
        if row.tiktok_product_id_confidence == "URL":
            report.product_id_from_url += 1
        elif row.tiktok_product_id_confidence == "LOW":
            report.product_id_low_confidence += 1
        if row.price_raw and row.price_min is not None and row.price_min != row.price_max:
            report.price_ranges_parsed += 1

    enrichment, matched, unmatched, internally_corrupt = build_hub_enrichment(
        hub_rows, merged_rows, hub_lookup_by_row_no
    )
    report.hub_matched = matched
    report.hub_unmatched_rows = unmatched[:50]
    report.hub_internally_corrupt_rows = internally_corrupt[:100]

    catalog_path = staged_catalog_path()
    hub_path = staged_hub_path()
    _atomic_write_json(catalog_path, staged_records)
    _atomic_write_json(hub_path, enrichment)
    report.staged = len(staged_records)
    report.staged_catalog_path = str(catalog_path)
    report.staged_hub_path = str(hub_path)
    logger.info(
        "kalodata import: staged=%d hub_matched=%d low_confidence_ids=%d",
        report.staged, report.hub_matched, report.product_id_low_confidence,
    )
    return report


def load_staged_catalog() -> list[dict[str, Any]]:
    path = staged_catalog_path()
    if not path.exists():
        return []
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        return data if isinstance(data, list) else []
    except (ValueError, OSError) as exc:  # fail-closed: bad file = no staged rows
        logger.error("kalodata staged catalog unreadable: %s", exc)
        return []


def load_hub_enrichment() -> dict[str, dict[str, Any]]:
    path = staged_hub_path()
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except (ValueError, OSError) as exc:
        logger.error("kalodata staged hub enrichment unreadable: %s", exc)
        return {}


# Queue rows in these states are HISTORY (committed / resolved) — re-applying
# enrichment would demote or re-open them.
_ENRICHMENT_TERMINAL_STATUSES = {"APPROVED", "DUPLICATE_LINKED", "REJECTED"}


async def apply_hub_enrichment(reference_ids: list[str] | None = None) -> dict[str, Any]:
    """Push staged HUB payloads through the EXISTING import_enrichment door
    (recompute-safe; never PATCHes derived product columns). Terminal queue
    rows (approved / duplicate-linked / rejected) are never touched."""
    from agent.db import crud
    from agent.services.fastmoss_bulk_promotion_service import import_enrichment

    staged = load_hub_enrichment()
    if reference_ids:
        wanted = {r for r in reference_ids if r}
        staged = {k: v for k, v in staged.items() if k in wanted}
    items = []
    skipped_terminal = 0
    for ref_id, fields in staged.items():
        row = await crud.get_bulk_queue_row(ref_id)
        if row and row.get("promotion_status") in _ENRICHMENT_TERMINAL_STATUSES:
            skipped_terminal += 1
            continue
        items.append({"reference_id": ref_id, **fields})
    if not items:
        return {"total": 0, "recomputed": 0, "skipped": 0, "failed": 0,
                "skipped_terminal": skipped_terminal, "results": []}
    result = await import_enrichment(items)
    result["skipped_terminal"] = skipped_terminal
    return result


async def collect_system_tids() -> set[str]:
    """TikTok product ids already known to the system: committed product rows
    plus the fastmoss reference WORKBOOK rows (never the staged Kalodata file —
    a re-import must not exclude itself)."""
    from agent.db import crud

    tids: set[str] = set()
    for product in await crud.list_products(include_archived=True):
        tid, _ = extract_tiktok_product_id(product.get("tiktok_product_url"), None)
        if tid:
            tids.add(tid)
    try:
        from agent.api.operator import _load_products

        for ref in _load_products(limit=2000):
            tid, _ = extract_tiktok_product_id(
                getattr(ref, "tiktok_product_url", None), None
            )
            if tid:
                tids.add(tid)
    except Exception as exc:  # noqa: BLE001 — workbook optional
        logger.warning("collect_system_tids: workbook refs unavailable: %s", exc)
    return tids


async def purge_redundant_queue_rows(dry_run: bool = False) -> dict[str, Any]:
    """Purge NEVER-DRAFTED queue rows whose TikTok product id duplicates either
    (a) another queue row for the same tid (the earliest/drafted row survives)
    or (b) an already-committed product. Drafted/approved rows are history and
    are never deleted."""
    from agent.db import crud

    product_tids: set[str] = set()
    for product in await crud.list_products(include_archived=True):
        tid, _ = extract_tiktok_product_id(product.get("tiktok_product_url"), None)
        if tid:
            product_tids.add(tid)

    rows = await crud.list_all_bulk_queue_rows()
    by_tid: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        tid, _ = extract_tiktok_product_id(row.get("tiktok_product_url"), None)
        if tid:
            by_tid.setdefault(tid, []).append(row)

    def _deletable(row: dict[str, Any]) -> bool:
        return not row.get("draft_id") and row.get("promotion_status") == "PENDING_DRAFT"

    to_delete: list[dict[str, Any]] = []
    for tid, group in by_tid.items():
        if tid in product_tids:
            # product already committed — every never-drafted queue twin goes
            to_delete.extend(r for r in group if _deletable(r))
            continue
        if len(group) > 1:
            keepers = [r for r in group if not _deletable(r)]
            survivors = keepers or group[:1]  # earliest row survives
            survivor_ids = {r["reference_id"] for r in survivors}
            to_delete.extend(
                r for r in group
                if r["reference_id"] not in survivor_ids and _deletable(r)
            )

    reference_ids = sorted({r["reference_id"] for r in to_delete})
    deleted = 0
    if reference_ids and not dry_run:
        deleted = await crud.delete_bulk_queue_rows(reference_ids)
    return {
        "dry_run": dry_run,
        "candidates": len(reference_ids),
        "deleted": deleted,
        "reference_ids": reference_ids[:100],
    }


async def hub_gaps() -> dict[str, Any]:
    """Read-only report: which staged products still lack HUB enrichment."""
    from agent.db import crud

    catalog = load_staged_catalog()
    enrichment = load_hub_enrichment()
    items: list[dict[str, Any]] = []
    with_hub = 0
    for record in catalog:
        reference_id = record.get("id") or ""
        fields = enrichment.get(reference_id) or {}
        missing = [
            key
            for key in (
                "target_customer_text", "benefits_text",
                "ingredients_text", "product_knowledge_text",
            )
            if not _clean(fields.get(key))
        ]
        if not missing:
            with_hub += 1
        queue_row = await crud.get_bulk_queue_row(reference_id) or {}
        items.append(
            {
                "reference_id": reference_id,
                "raw_product_title": record.get("raw_product_title"),
                "missing_hub_fields": missing,
                "promotion_status": queue_row.get("promotion_status"),
                "claim_risk_level": queue_row.get("claim_risk_level"),
                "image_readiness": queue_row.get("image_readiness"),
            }
        )
    return {
        "items": items,
        "totals": {
            "staged": len(catalog),
            "fully_enriched": with_hub,
            "with_any_gap": len(catalog) - with_hub,
        },
    }
