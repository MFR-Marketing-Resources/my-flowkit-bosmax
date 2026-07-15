"""Kalodata staged import — parser/normalizer/staging contract tests."""
import json

import pytest

from agent.services import kalodata_import_service as svc
from agent.services.fastmoss_product_reference_service import (
    FASTMOSS_REFERENCE_ID_PREFIX,
)


# ── normalizers ──────────────────────────────────────────────────────────────
def test_tiktok_id_recovered_from_url_patterns():
    assert svc.extract_tiktok_product_id(
        "https://shop-my.tiktok.com/pdp/1731147231842430988?src=x", None
    ) == ("1731147231842430988", "URL")
    assert svc.extract_tiktok_product_id(
        "https://shop.tiktok.com/view/product/1729674255001688097", 1.72e18
    ) == ("1729674255001688097", "URL")


def test_tiktok_id_float_cell_beyond_precision_is_low_confidence():
    # 19-digit ids exceed float64's 53-bit mantissa — unrecoverable from a cell.
    assert svc.extract_tiktok_product_id(None, 1.73114723184243e18) == (None, "LOW")


def test_tiktok_id_small_integral_cell_accepted():
    assert svc.extract_tiktok_product_id(None, 123456.0) == ("123456", "CELL")
    assert svc.extract_tiktok_product_id(None, "1731147231842430988") == (
        "1731147231842430988",
        "CELL",
    )
    assert svc.extract_tiktok_product_id(None, None) == (None, "NONE")


def test_price_range_parses_midpoint_min_max():
    mid, low, high, raw, is_range = svc.parse_price("RM6.12  - 19.01")
    assert (mid, low, high, is_range) == (12.57, 6.12, 19.01, True)
    assert raw.startswith("RM6.12")


def test_price_single_and_numeric():
    assert svc.parse_price("RM45.90")[:3] == (45.9, 45.9, 45.9)
    assert svc.parse_price(26.5)[:3] == (26.5, 26.5, 26.5)
    assert svc.parse_price(None) == (None, None, None, None, False)


def test_excel_serial_and_datetime_dates():
    assert svc.excel_date_to_iso(45765) == "2025-04-18"
    from datetime import datetime

    assert svc.excel_date_to_iso(datetime(2025, 7, 18)) == "2025-07-18"
    assert svc.excel_date_to_iso("2025-04-18") == "2025-04-18"
    assert svc.excel_date_to_iso(None) is None


# ── workbook fixture ─────────────────────────────────────────────────────────
def _build_workbook(path):
    import openpyxl

    wb = openpyxl.Workbook()
    merged = wb.active
    merged.title = svc.MERGED_SHEET
    merged.append([
        "No", "Sumber", "Product Name", "Image URL", "Category", "Price (RM)",
        "Launch Date", "Product Rating", "Item Sold / Total Sales",
        "Avg Unit Price (RM)", "Commission Rate", "Creator Number",
        "Creator Conversion/Sales Rate", "TikTok URL", "Product ID", "Source URL 1",
    ])
    merged.append([
        1, "KALODATA", "Pengedap Vakum Mudah Alih", "https://img/1.jpg",
        "Home Supplies > Home Organizers > Storage", "RM26.50", 45765, 4.5, 3448,
        123.99, 0.05, 57, 0.7719,
        "https://shop-my.tiktok.com/pdp/1731147231842430988",
        1.73114723184243e18, "https://kalodata/1",
    ])
    merged.append([
        2, "FASTMOSS", "Jalur Pemutih Gigi", "https://img/2.jpg",
        "Beauty & Personal Care > Oral", "RM6.12 - 19.01", 45856, 4.5, 6335,
        47.54, 0.1, 120, 0.8333,
        "https://shop.tiktok.com/view/product/1731684607938880001",
        None, "https://kalodata/2",
    ])
    # duplicate of row 1 (same name+urls → same reference id)
    merged.append([
        3, "KALODATA", "Pengedap Vakum Mudah Alih", "https://img/1.jpg",
        "Home Supplies", "RM26.50", 45765, 4.5, 3448, 123.99, 0.05, 57, 0.7719,
        "https://shop-my.tiktok.com/pdp/1731147231842430988",
        None, "https://kalodata/1",
    ])
    # no image, no urls
    merged.append([
        4, "KALODATA", "Produk Tanpa Gambar", None, "Food", "RM9.90",
        None, None, None, None, None, None, None, None, None, None,
    ])

    hub = wb.create_sheet(svc.HUB_SHEET)
    hub.append([
        "No", "Product ID", "Product Name", "Product Type", "Category",
        "Price (RM)", "Image URL", "Target Avatar", "Pain Point",
        "Emotion/Trigger", "Dream Outcome", "Key Ingredient/Feature",
        "Main Benefit", "Secondary Benefit", "USP", "Hook Type",
    ])
    # DELIBERATELY MISALIGNED row №s (mirrors the real workbook: HUB ordering
    # differs from MERGED PRODUCTS — the join must be by NAME, never by №).
    hub.append([  # № says 2, but the copy belongs to Pengedap Vakum (merged №1)
        2, None, "Pengedap Vakum Mudah Alih", "Home", "Home Supplies", "RM26.50",
        "https://img/1.jpg", "Ibu rumah yang kemas", "Makanan cepat basi",
        "Geram bila bazir", "Dapur tersusun", "Vacuum seal technology",
        "Makanan tahan lebih lama", "Jimat ruang", "Seal kedap udara", "Problem-Solution",
    ])
    hub.append([  # empty enrichment row → skipped
        1, None, "Jalur Pemutih Gigi", None, None, None, None,
        None, None, None, None, None, None, None, None, None,
    ])
    hub.append([  # unmatched product name
        99, None, "Produk Tidak Wujud", None, None, None, None,
        "Avatar X", None, None, None, None, None, None, None, None,
    ])
    wb.save(path)


@pytest.fixture()
def staged_env(tmp_path, monkeypatch):
    monkeypatch.setattr(svc, "OPERATOR_PACK_DIR", tmp_path)
    workbook_path = tmp_path / "kalodata_sample.xlsx"
    _build_workbook(workbook_path)
    return workbook_path


# ── import_workbook ──────────────────────────────────────────────────────────
def test_import_workbook_stages_records_and_hub(staged_env, tmp_path):
    report = svc.import_workbook(staged_env)
    assert report.parsed_merged == 4
    assert report.staged == 3  # duplicate collapsed
    assert report.skipped_duplicate_in_file == 1
    assert report.product_id_from_url == 2
    assert report.price_ranges_parsed == 1
    assert report.hub_matched == 1
    assert 99 in report.hub_unmatched_rows

    catalog = json.loads((tmp_path / svc.STAGED_CATALOG_FILENAME).read_text(encoding="utf-8"))
    assert len(catalog) == 3
    first = catalog[0]
    assert first["id"].startswith(FASTMOSS_REFERENCE_ID_PREFIX)
    assert first["source"] == "KALODATA"
    assert first["reference_only"] is True
    assert first["price"] == 26.5
    assert first["currency"] == "MYR"
    assert first["kalodata_meta"]["tiktok_product_id"] == "1731147231842430988"
    assert first["kalodata_meta"]["launch_date"] == "2025-04-18"

    hub = json.loads((tmp_path / svc.STAGED_HUB_FILENAME).read_text(encoding="utf-8"))
    assert len(hub) == 1
    (ref_id, item), = hub.items()
    # NAME join: HUB row carries № 2, yet its copy lands on the product whose
    # NAME matches (Pengedap Vakum) — never on merged № 2 (Jalur Pemutih Gigi).
    assert ref_id == first["id"]
    assert "Pengedap" in first["raw_product_title"]
    assert item["target_customer_text"] == "Ibu rumah yang kemas"
    assert "Vacuum seal technology" == item["ingredients_text"]
    assert "Makanan tahan lebih lama" in item["benefits_text"]
    assert "Seal kedap udara" in item["benefits_text"]
    assert "Pain point: Makanan cepat basi" in item["product_knowledge_text"]
    assert "Hook type: Problem-Solution" in item["product_knowledge_text"]
    assert item["price"] == 26.5


def test_tid_skipped_rows_still_get_hub_enrichment(staged_env, tmp_path):
    """Rows refused at the tid gate already exist in the QUEUE from earlier
    imports — their HUB copy must still land in the staged hub map, or
    Recompute rebuilds those drafts with empty knowledge fields (Owner-
    reported live incident 2026-07-14)."""
    report = svc.import_workbook(
        staged_env, existing_tids={"1731147231842430988"}
    )
    # Pengedap Vakum (the hub-matched product) was tid-skipped, NOT staged...
    assert report.staged == 2
    # ...yet its HUB copy is still mapped under its reference id
    assert report.hub_matched == 1
    hub = json.loads(
        (tmp_path / svc.STAGED_HUB_FILENAME).read_text(encoding="utf-8")
    )
    (ref_id, item), = hub.items()
    assert item["target_customer_text"] == "Ibu rumah yang kemas"
    catalog_ids = {
        r["id"] for r in json.loads(
            (tmp_path / svc.STAGED_CATALOG_FILENAME).read_text(encoding="utf-8")
        )
    }
    assert ref_id not in catalog_ids  # hub entry exists WITHOUT being staged


def test_import_skips_tids_already_in_system(staged_env, tmp_path):
    # Duplicate law: TikTok Product ID = product identity. Row 1's tid already
    # exists in the system → never staged; the others stage normally.
    report = svc.import_workbook(
        staged_env, existing_tids={"1731147231842430988"}
    )
    # both fixture rows carrying that tid (the original AND its in-file twin)
    # are refused at the system gate
    assert report.skipped_existing_tid == 2
    assert report.staged == 2
    catalog = json.loads((tmp_path / svc.STAGED_CATALOG_FILENAME).read_text(encoding="utf-8"))
    tids = {(r.get("kalodata_meta") or {}).get("tiktok_product_id") for r in catalog}
    assert "1731147231842430988" not in tids


def test_import_dedupes_same_tid_within_file(tmp_path, monkeypatch):
    import openpyxl

    monkeypatch.setattr(svc, "OPERATOR_PACK_DIR", tmp_path)
    wb = openpyxl.Workbook()
    merged = wb.active
    merged.title = svc.MERGED_SHEET
    merged.append(["No", "Sumber", "Product Name", "Image URL", "Category",
                   "Price (RM)", "Launch Date", "Product Rating", "Item Sold",
                   "Avg Unit Price", "Commission Rate", "Creator Number",
                   "Conversion", "TikTok URL", "Product ID", "Source URL"])
    # same tid, DIFFERENT truncated titles + different source urls → different
    # sha1 reference ids, but ONE product
    merged.append([1, "KALODATA", "Kepingan ganti kaca penutup belakang iPhone A",
                   None, "Tech", "RM10", None, None, None, None, None, None, None,
                   "https://shop-my.tiktok.com/pdp/1733902025147385549", None, "https://k/1"])
    merged.append([2, "FASTMOSS", "Kepingan ganti kaca penutup belakang iPh",
                   None, "Tech", "RM10", None, None, None, None, None, None, None,
                   "https://shop.tiktok.com/view/product/1733902025147385549", None, "https://f/2"])
    wb.create_sheet(svc.HUB_SHEET)
    path = tmp_path / "dupe.xlsx"
    wb.save(path)

    report = svc.import_workbook(path)
    assert report.staged == 1
    assert report.skipped_duplicate_in_file == 1


# ── HUB source-corruption guard ──────────────────────────────────────────────
def _hub_row(row_no, name, **copy):
    from agent.models.kalodata_import import KalodataHubRow

    return KalodataHubRow(row_no=row_no, product_name=name, **copy)


def test_hub_guard_flags_wrong_brand_copy_and_keeps_own_brand():
    rows = [
        # corrupt: SZINDORE perfume row carrying FOCALLURE copy (real workbook case)
        _hub_row(1, "SZINDORE BBW AROMA PERFUME",
                 main_benefit="FOCALLURE lip clay tahan lama, warna pekat"),
        # clean: names its own brand
        _hub_row(2, "FOCALLURE Lip Clay Matte",
                 main_benefit="Focallure lip clay tekstur mousse"),
        # clean: copy omits the name entirely — not proof of corruption
        _hub_row(3, "Pengedap Vakum Mudah Alih",
                 main_benefit="Makanan tahan lebih lama, seal kedap udara"),
    ]
    assert svc.find_hub_internal_corruption(rows) == {1}


def test_hub_guard_ignores_common_tokens():
    # "perfume" appears in 3 names → not distinctive; mentioning it is fine
    rows = [
        _hub_row(1, "BELLA PERFUME PINK", main_benefit="Perfume tahan 8 jam"),
        _hub_row(2, "SZINDORE PERFUME MEN", main_benefit="Perfume lelaki premium"),
        _hub_row(3, "AROMA PERFUME CLASSIC", main_benefit="Perfume bau segar"),
    ]
    assert svc.find_hub_internal_corruption(rows) == set()


def test_import_quarantines_internally_corrupt_hub_rows(tmp_path, monkeypatch):
    import openpyxl

    monkeypatch.setattr(svc, "OPERATOR_PACK_DIR", tmp_path)
    wb = openpyxl.Workbook()
    merged = wb.active
    merged.title = svc.MERGED_SHEET
    merged.append(["No", "Sumber", "Product Name", "Image URL", "Category",
                   "Price (RM)", "Launch Date", "Product Rating", "Item Sold",
                   "Avg Unit Price", "Commission Rate", "Creator Number",
                   "Conversion", "TikTok URL", "Product ID", "Source URL"])
    merged.append([1, "KALODATA", "SZINDORE BBW AROMA PERFUME", None, "Beauty",
                   "RM20", None, None, None, None, None, None, None,
                   "https://shop-my.tiktok.com/pdp/1000000000000000001", None, "https://k/1"])
    merged.append([2, "KALODATA", "FOCALLURE Lip Clay Matte", None, "Beauty",
                   "RM15", None, None, None, None, None, None, None,
                   "https://shop-my.tiktok.com/pdp/1000000000000000002", None, "https://k/2"])
    hub = wb.create_sheet(svc.HUB_SHEET)
    hub.append(["No", "Product ID", "Product Name", "Product Type", "Category",
                "Price (RM)", "Image URL", "Target Avatar", "Pain Point",
                "Emotion/Trigger", "Dream Outcome", "Key Ingredient/Feature",
                "Main Benefit", "Secondary Benefit", "USP", "Hook Type"])
    # corrupt at source: perfume row carries the lip-clay product's copy
    hub.append([1, None, "SZINDORE BBW AROMA PERFUME", None, "Beauty", "RM20",
                None, "Makeup lovers", None, None, None, None,
                "FOCALLURE lip clay warna pekat", None, None, None])
    # clean row
    hub.append([2, None, "FOCALLURE Lip Clay Matte", None, "Beauty", "RM15",
                None, "Makeup lovers", None, None, None, None,
                "Focallure lip clay tekstur mousse", None, None, None])
    path = tmp_path / "corrupt.xlsx"
    wb.save(path)

    report = svc.import_workbook(path)
    assert report.hub_internally_corrupt_rows == [1]
    assert report.hub_matched == 1
    staged_hub = json.loads(
        (tmp_path / svc.STAGED_HUB_FILENAME).read_text(encoding="utf-8")
    )
    # only the clean row's enrichment landed; the perfume got NO foreign copy
    assert len(staged_hub) == 1
    (ref_id,) = staged_hub
    catalog = json.loads(
        (tmp_path / svc.STAGED_CATALOG_FILENAME).read_text(encoding="utf-8")
    )
    focallure = next(r for r in catalog if "FOCALLURE" in r["raw_product_title"])
    assert ref_id == focallure["id"]


def test_import_workbook_is_idempotent(staged_env):
    first = svc.import_workbook(staged_env)
    second = svc.import_workbook(staged_env)
    assert first.staged == second.staged == 3
    assert svc.load_staged_catalog() == svc.load_staged_catalog()


def test_copy_intelligence_dry_run_preserves_scripts_and_quarantines_ambiguous_names(tmp_path):
    """The review-only lane reads current HUB headers without row-number joins."""
    import openpyxl

    workbook = openpyxl.Workbook()
    merged = workbook.active
    merged.title = svc.MERGED_SHEET
    merged.append(["No", "Sumber", "Product Name", "Image URL", "Category", "Price (RM)",
                   "Launch Date", "Product Rating", "Item Sold", "Avg Unit Price",
                   "Commission Rate", "Creator Number", "Conversion", "TikTok URL",
                   "Product ID", "Source URL"])
    merged.append([1, "KALODATA", "Produk Unik", None, "Home", "RM10", None, None,
                   None, None, None, None, None,
                   "https://shop.tiktok.com/view/product/1731147231842430988", None, None])
    merged.append([2, "KALODATA", "Produk Sama", None, "Home", "RM11", None, None,
                   None, None, None, None, None,
                   "https://shop.tiktok.com/view/product/1731147231842430999", None, None])
    hub = workbook.create_sheet(svc.HUB_SHEET)
    hub.append(["No", "Product ID", "Product Name", "Target Avatar 1", "Pain Point 1",
                "Emotion/Trigger 1", "Dream Outcome 1", "Key Feature 1", "Benefit 1",
                "Hook Type", "Hook Script 1", "CTA Type", "CTA Script 1"])
    hub.append([99, 1.73114723184243e18, "Produk Unik", "Ibu sibuk", "Ruang sempit",
                "Tenang", "Rumah kemas", "Saiz kompak", "Mudah simpan", "Problem-Solution",
                "Rumah sempit? Cuba susun begini.", "Direct", "Tambah ke cart."])
    hub.append([1, None, "Produk Sama", "A", "P", "E", "D", "K", "B", "H", "H", "C", "C"])
    hub.append([2, None, "Produk Sama", "A", "P", "E", "D", "K", "B", "H", "H", "C", "C"])
    path = tmp_path / "copy_intelligence.xlsx"
    workbook.save(path)

    report = svc.build_copy_intelligence_dry_run(path)

    assert report.total_source_rows == 3
    assert report.usable_rows == 3
    assert report.matched_high_confidence == 0
    assert report.matched_medium_confidence == 1
    assert report.low_confidence_quarantined == 2
    assert report.unmatched == 0
    assert report.duplicates == 2
    unique = next(record for record in report.records if record.source_product_name == "Produk Unik")
    assert unique.source_row == 2  # provenance, never identity
    assert unique.key_ingredients_features == "Saiz kompak"
    assert unique.hook_type == "Problem-Solution"
    assert unique.hook_script == "Rumah sempit? Cuba susun begini."
    assert unique.cta_type == "Direct"
    assert unique.cta_script == "Tambah ke cart."
    assert unique.status == "NEEDS_REVIEW"
    assert unique.match_method == "UNIQUE_NORMALIZED_NAME_TO_SOURCE_REFERENCE"
    ambiguous = [record for record in report.records if record.source_product_name == "Produk Sama"]
    assert {record.match_method for record in ambiguous} == {"AMBIGUOUS_NORMALIZED_NAME"}

    high = svc.build_copy_intelligence_dry_run(
        path, {"1731147231842430988": "product-truth-001"}
    )
    linked = next(record for record in high.records if record.source_product_name == "Produk Unik")
    assert high.matched_high_confidence == 1
    assert linked.target_product_id == "product-truth-001"
    assert linked.match_method == "TIKTOK_PRODUCT_ID_MATCH"


def test_copy_intelligence_duplicate_product_truth_tid_is_quarantined(tmp_path):
    """A duplicated Product Truth TikTok ID can never select an arbitrary target."""
    workbook_path = tmp_path / "duplicate-tid.xlsx"
    _build_workbook(workbook_path)
    import openpyxl
    workbook = openpyxl.load_workbook(workbook_path)
    workbook[svc.MERGED_SHEET].delete_rows(4, 1)  # remove fixture's duplicate source row
    workbook.save(workbook_path)

    report = svc.build_copy_intelligence_dry_run(
        workbook_path,
        {"1731147231842430988": ["product-a", "product-b"]},
    )

    row = next(record for record in report.records if record.source_product_name == "Pengedap Vakum Mudah Alih")
    assert row.confidence == "LOW"
    assert row.target_product_id is None
    assert row.match_method == "DUPLICATE_PRODUCT_TRUTH_TIKTOK_ID"
    assert report.low_confidence_quarantined >= 1


@pytest.mark.asyncio
async def test_system_copy_intelligence_dry_run_executes_without_seed_write(tmp_path, monkeypatch):
    """The API-facing read-only wrapper must resolve Product Truth candidates."""
    from agent.db import crud

    workbook_path = tmp_path / "system-dry-run.xlsx"
    _build_workbook(workbook_path)

    async def list_products(*, include_archived: bool):
        assert include_archived is True
        return []

    monkeypatch.setattr(crud, "list_products", list_products)

    report = await svc.build_copy_intelligence_dry_run_for_system(workbook_path)

    assert report.total_source_rows == 3
    assert report.matched_high_confidence == 0


@pytest.mark.asyncio
async def test_copy_intelligence_seed_is_idempotent_and_never_touches_product_truth():
    from agent.db import crud

    payload = {
        "source_fingerprint": "copy-hub-test-fingerprint",
        "source_workbook": "fixture.xlsx",
        "source_sheet": "COPYWRITING HUB",
        "source_row": 2,
        "source_product_name": "Produk Ujian",
        "match_method": "UNMATCHED",
        "confidence": "LOW",
        "status": "NEEDS_REVIEW",
        "provenance_json": '{"sheet":"COPYWRITING HUB"}',
    }
    first = await crud.create_copy_intelligence_seed(**payload)
    second = await crud.create_copy_intelligence_seed(**payload)

    assert first["seed_id"] == second["seed_id"]
    assert first["status"] == "NEEDS_REVIEW"
    assert first["target_product_id"] is None


@pytest.mark.asyncio
async def test_seed_reimport_preserves_product_truth_and_approved_copy_set():
    """The seed ledger is additive: it cannot mutate product or approved copy."""
    from agent.db import crud
    from agent.models.kalodata_import import CopyIntelligenceSourceRow

    product = await crud.create_product(
        source="MANUAL", raw_product_title="Produk Truth", product_display_name="Produk Truth",
        product_short_name="Produk Truth", copywriting_angle="Manual authority",
    )
    approved = await crud.create_copy_set(
        product["id"], hook="Approved hook", cta="Approved CTA", usp_set_json='["Approved USP"]',
        status="COPY_APPROVED",
    )
    source = CopyIntelligenceSourceRow(
        source_workbook="fixture.xlsx", source_sheet="COPYWRITING HUB", source_row=7,
        source_product_name="Produk Truth", target_product_id=product["id"],
        match_method="TIKTOK_PRODUCT_ID_MATCH", confidence="HIGH", source_fingerprint="seed-reimport-proof",
        hook_script="Imported hook remains review-only", provenance={"workbook": "fixture.xlsx"},
    )
    before_product = await crud.get_product(product["id"])
    before_copy = await crud.get_copy_set(approved["copy_set_id"])

    await svc.persist_copy_intelligence_seed_records([source])
    await svc.persist_copy_intelligence_seed_records([source])

    db = await crud.get_db()
    cursor = await db.execute(
        "SELECT COUNT(*) FROM copy_intelligence_seed WHERE source_fingerprint=?",
        (source.source_fingerprint,),
    )
    assert (await cursor.fetchone())[0] == 1
    assert await crud.get_product(product["id"]) == before_product
    assert await crud.get_copy_set(approved["copy_set_id"]) == before_copy


@pytest.mark.asyncio
async def test_normal_seed_persist_skips_unmatched_and_low_confidence_records():
    from agent.models.kalodata_import import CopyIntelligenceSourceRow

    records = [
        CopyIntelligenceSourceRow(
            source_workbook="fixture.xlsx", source_sheet="COPYWRITING HUB", source_row=81,
            source_product_name="Unmatched", match_method="UNMATCHED", confidence="LOW",
            source_fingerprint="skip-unmatched", provenance={},
        ),
        CopyIntelligenceSourceRow(
            source_workbook="fixture.xlsx", source_sheet="COPYWRITING HUB", source_row=82,
            source_product_name="Duplicate target", match_method="DUPLICATE_PRODUCT_TRUTH_TIKTOK_ID",
            confidence="LOW", source_fingerprint="skip-low", provenance={"quarantine_reason": "DUPLICATE_PRODUCT_TRUTH_TIKTOK_ID"},
        ),
    ]

    result = await svc.persist_copy_intelligence_seed_records(records)

    assert result == {
        "records_processed": 2,
        "created_or_existing": 0,
        "skipped_quarantined": 1,
        "skipped_low_confidence": 0,
        "skipped_unmatched": 1,
        "status": "NEEDS_REVIEW_ONLY",
    }


def test_import_workbook_missing_file_raises(tmp_path, monkeypatch):
    monkeypatch.setattr(svc, "OPERATOR_PACK_DIR", tmp_path)
    with pytest.raises(FileNotFoundError):
        svc.import_workbook(tmp_path / "nope.xlsx")


def test_loaders_fail_closed_on_corrupt_files(tmp_path, monkeypatch):
    monkeypatch.setattr(svc, "OPERATOR_PACK_DIR", tmp_path)
    (tmp_path / svc.STAGED_CATALOG_FILENAME).write_text("{not json", encoding="utf-8")
    (tmp_path / svc.STAGED_HUB_FILENAME).write_text("[]", encoding="utf-8")
    assert svc.load_staged_catalog() == []
    assert svc.load_hub_enrichment() == {}


# ── reference-catalog union ──────────────────────────────────────────────────
@pytest.mark.asyncio
async def test_reference_union_includes_staged_records(staged_env, monkeypatch):
    from agent.services import fastmoss_product_reference_service as ref_svc

    svc.import_workbook(staged_env)

    async def fake_enrich(record, persist=False):
        return dict(record)

    monkeypatch.setattr(ref_svc, "enrich_product", fake_enrich)
    monkeypatch.setattr(ref_svc, "_REFERENCE_CACHE_SIGNATURE", None)
    monkeypatch.setattr(ref_svc, "_REFERENCE_CACHE_ITEMS", [])
    monkeypatch.setattr(ref_svc, "_REFERENCE_CACHE_LOADED_LIMIT", 0)

    items = await ref_svc.list_fastmoss_reference_products(limit=2000)
    kalodata_items = [i for i in items if i.get("source_label") == "Kalodata Reference"]
    assert len(kalodata_items) == 3
    assert all(i["id"].startswith(FASTMOSS_REFERENCE_ID_PREFIX) for i in kalodata_items)
    assert all(i["reference_only"] for i in kalodata_items)

    # cache invalidates when the staged file changes (re-import)
    svc.import_workbook(staged_env)
    items_again = await ref_svc.list_fastmoss_reference_products(limit=2000)
    assert len([i for i in items_again if i.get("source_label") == "Kalodata Reference"]) == 3


@pytest.mark.asyncio
async def test_purge_redundant_queue_rows_keeps_drafted_survivor(monkeypatch):
    """Never-drafted queue twins of the same tid — or of a committed product —
    are purged; drafted/approved rows always survive."""
    from agent.db import crud

    products = [
        {"id": "p-committed", "tiktok_product_url": "https://shop.tiktok.com/view/product/9990001112223334"},
    ]
    queue = [
        # tid duplicated across two queue rows: drafted survives, pending goes
        {"reference_id": "fastmoss-ref:aaa", "draft_id": "draft-1",
         "promotion_status": "READY_FOR_APPROVAL",
         "tiktok_product_url": "https://shop-my.tiktok.com/pdp/1110002223334445"},
        {"reference_id": "fastmoss-ref:bbb", "draft_id": None,
         "promotion_status": "PENDING_DRAFT",
         "tiktok_product_url": "https://shop.tiktok.com/view/product/1110002223334445"},
        # tid already a committed product: pending row goes
        {"reference_id": "fastmoss-ref:ccc", "draft_id": None,
         "promotion_status": "PENDING_DRAFT",
         "tiktok_product_url": "https://shop-my.tiktok.com/pdp/9990001112223334"},
        # unique tid: stays
        {"reference_id": "fastmoss-ref:ddd", "draft_id": None,
         "promotion_status": "PENDING_DRAFT",
         "tiktok_product_url": "https://shop-my.tiktok.com/pdp/5556667778889990"},
    ]
    deleted_ids: list[str] = []

    async def fake_list_products(**kwargs):
        return products

    async def fake_list_rows():
        return queue

    async def fake_delete(reference_ids):
        deleted_ids.extend(reference_ids)
        return len(reference_ids)

    monkeypatch.setattr(crud, "list_products", fake_list_products)
    monkeypatch.setattr(crud, "list_all_bulk_queue_rows", fake_list_rows)
    monkeypatch.setattr(crud, "delete_bulk_queue_rows", fake_delete)

    dry = await svc.purge_redundant_queue_rows(dry_run=True)
    assert dry["candidates"] == 2 and dry["deleted"] == 0 and not deleted_ids

    result = await svc.purge_redundant_queue_rows()
    assert sorted(deleted_ids) == ["fastmoss-ref:bbb", "fastmoss-ref:ccc"]
    assert result["deleted"] == 2


@pytest.mark.asyncio
async def test_duplicate_detector_matches_by_tiktok_product_id(monkeypatch):
    """Draft-time guard: a variant/truncated title with the SAME TikTok id is
    caught even though the title query finds nothing."""
    from agent.db import crud as _crud
    from agent.services.fastmoss_bulk_promotion_service import (
        _detect_queue_duplicate_candidate,
    )

    async def fake_list_products(**kwargs):
        return []  # title query finds nothing (variant title)

    async def fake_find_by_tid(tid):
        assert tid == "1731147231842430988"
        return [{
            "id": "p-1", "product_display_name": "Pengedap Vakum",
            "source": "MANUAL", "mapping_source": "FASTMOSS_PROMOTED",
            "tiktok_product_url": f"https://shop.tiktok.com/view/product/{tid}",
        }]

    monkeypatch.setattr(_crud, "list_products", fake_list_products)
    monkeypatch.setattr(_crud, "find_products_by_tiktok_product_id", fake_find_by_tid)

    candidate = await _detect_queue_duplicate_candidate(
        "fastmoss-ref:x", "Pengedap Vakum Mudah Alih VARIANT TITLE",
        "https://shop-my.tiktok.com/pdp/1731147231842430988",
    )
    assert candidate is not None
    assert candidate["match_reason"] == "TIKTOK_PRODUCT_ID_MATCH_EXISTING_PRODUCT"
    assert candidate["id"] == "p-1"

    # raw FASTMOSS reference rows still never block by tid either
    async def fake_find_raw(tid):
        return [{"id": "p-raw", "source": "FASTMOSS", "mapping_source": None,
                 "tiktok_product_url": "x"}]

    monkeypatch.setattr(_crud, "find_products_by_tiktok_product_id", fake_find_raw)
    assert await _detect_queue_duplicate_candidate(
        "fastmoss-ref:x", "Sesuatu Produk",
        "https://shop-my.tiktok.com/pdp/1731147231842430988",
    ) is None


@pytest.mark.asyncio
async def test_apply_hub_enrichment_delegates_to_import_enrichment(staged_env, monkeypatch):
    svc.import_workbook(staged_env)
    captured: dict = {}

    async def fake_import_enrichment(items):
        captured["items"] = items
        return {"total": len(items), "recomputed": len(items), "skipped": 0,
                "failed": 0, "results": []}

    import agent.services.fastmoss_bulk_promotion_service as bulk
    monkeypatch.setattr(bulk, "import_enrichment", fake_import_enrichment)

    result = await svc.apply_hub_enrichment()
    assert result["total"] == 1
    item = captured["items"][0]
    assert item["reference_id"].startswith(FASTMOSS_REFERENCE_ID_PREFIX)
    # exact import_enrichment field names only
    assert set(item.keys()) <= {
        "reference_id", "price", "benefits_text", "usage_text",
        "target_customer_text", "ingredients_text", "warnings_text",
        "product_knowledge_text",
    }


# ── Copy Intelligence seed review/approval layer ─────────────────────────────
async def _make_needs_review_seed(fingerprint: str, confidence: str = "HIGH") -> dict:
    from agent.db import crud

    return await crud.create_copy_intelligence_seed(
        source_fingerprint=fingerprint,
        source_workbook="fixture.xlsx", source_sheet="COPYWRITING HUB",
        source_row=2, source_product_name="Produk Review",
        match_method="TIKTOK_PRODUCT_ID_MATCH", confidence=confidence,
        status="NEEDS_REVIEW", hook_script="Hook", cta_script="CTA",
        provenance_json='{"sheet":"COPYWRITING HUB"}',
    )


@pytest.mark.asyncio
async def test_review_high_needs_review_to_approved_persists_audit():
    from agent.db import crud

    seed = await _make_needs_review_seed("review-approve-high")
    result = await svc.review_copy_intelligence_seed(
        seed["seed_id"], action="APPROVE", reviewed_by="owner",
        review_note="Verified product identity",
        confirmation_phrase="APPROVE COPY INTELLIGENCE",
    )
    assert result["previous_status"] == "NEEDS_REVIEW"
    assert result["new_status"] == "APPROVED"
    assert result["reviewed_by"] == "owner"
    assert result["review_note"] == "Verified product identity"
    assert result["reviewed_at"]

    row = await crud.get_copy_intelligence_seed(seed["seed_id"])
    assert row["status"] == "APPROVED"
    assert row["previous_status"] == "NEEDS_REVIEW"
    assert row["review_action"] == "APPROVE"
    assert row["reviewed_by"] == "owner"
    assert row["reviewed_at"] == result["reviewed_at"]
    assert row["review_note"] == "Verified product identity"


@pytest.mark.asyncio
async def test_review_high_needs_review_to_rejected():
    from agent.db import crud

    seed = await _make_needs_review_seed("review-reject-high")
    result = await svc.review_copy_intelligence_seed(
        seed["seed_id"], action="REJECT", reviewed_by="owner",
        review_note="Wrong mapping", confirmation_phrase="REJECT COPY INTELLIGENCE",
    )
    assert result["new_status"] == "REJECTED"
    row = await crud.get_copy_intelligence_seed(seed["seed_id"])
    assert row["status"] == "REJECTED"
    assert row["review_action"] == "REJECT"


@pytest.mark.asyncio
async def test_review_medium_approval_with_normal_phrase_is_rejected():
    from agent.db import crud

    seed = await _make_needs_review_seed("review-medium-normal", confidence="MEDIUM")
    with pytest.raises(svc.CopyIntelligenceReviewError) as exc:
        await svc.review_copy_intelligence_seed(
            seed["seed_id"], action="APPROVE", reviewed_by="owner",
            review_note="note", confirmation_phrase="APPROVE COPY INTELLIGENCE",
        )
    assert exc.value.code == "MEDIUM_CONFIDENCE_PHRASE_REQUIRED"
    assert exc.value.status_code == 422
    row = await crud.get_copy_intelligence_seed(seed["seed_id"])
    assert row["status"] == "NEEDS_REVIEW"


@pytest.mark.asyncio
async def test_review_medium_approval_with_stronger_phrase_succeeds():
    from agent.db import crud

    seed = await _make_needs_review_seed("review-medium-strong", confidence="MEDIUM")
    result = await svc.review_copy_intelligence_seed(
        seed["seed_id"], action="APPROVE", reviewed_by="owner",
        review_note="Confirmed identity carefully",
        confirmation_phrase="APPROVE MEDIUM CONFIDENCE COPY INTELLIGENCE",
    )
    assert result["new_status"] == "APPROVED"
    assert (await crud.get_copy_intelligence_seed(seed["seed_id"]))["status"] == "APPROVED"


@pytest.mark.asyncio
async def test_review_wrong_phrase_rejected_fail_closed():
    from agent.db import crud

    seed = await _make_needs_review_seed("review-wrong-phrase")
    with pytest.raises(svc.CopyIntelligenceReviewError) as exc:
        await svc.review_copy_intelligence_seed(
            seed["seed_id"], action="APPROVE", reviewed_by="owner",
            review_note="note", confirmation_phrase="approve",
        )
    assert exc.value.code == "CONFIRMATION_PHRASE_MISMATCH"
    assert (await crud.get_copy_intelligence_seed(seed["seed_id"]))["status"] == "NEEDS_REVIEW"


@pytest.mark.asyncio
async def test_review_missing_note_rejected():
    seed = await _make_needs_review_seed("review-missing-note")
    with pytest.raises(svc.CopyIntelligenceReviewError) as exc:
        await svc.review_copy_intelligence_seed(
            seed["seed_id"], action="APPROVE", reviewed_by="owner",
            review_note="   ", confirmation_phrase="APPROVE COPY INTELLIGENCE",
        )
    assert exc.value.code == "REVIEW_NOTE_REQUIRED"


@pytest.mark.asyncio
async def test_review_missing_reviewer_rejected():
    seed = await _make_needs_review_seed("review-missing-reviewer")
    with pytest.raises(svc.CopyIntelligenceReviewError) as exc:
        await svc.review_copy_intelligence_seed(
            seed["seed_id"], action="APPROVE", reviewed_by="",
            review_note="note", confirmation_phrase="APPROVE COPY INTELLIGENCE",
        )
    assert exc.value.code == "REVIEWER_REQUIRED"


@pytest.mark.asyncio
async def test_review_unknown_seed_is_404():
    with pytest.raises(svc.CopyIntelligenceReviewError) as exc:
        await svc.review_copy_intelligence_seed(
            "does-not-exist", action="APPROVE", reviewed_by="owner",
            review_note="note", confirmation_phrase="APPROVE COPY INTELLIGENCE",
        )
    assert exc.value.code == "COPY_INTELLIGENCE_SEED_NOT_FOUND"
    assert exc.value.status_code == 404


@pytest.mark.asyncio
async def test_review_invalid_source_status_is_409():
    from agent.db import crud

    seed = await _make_needs_review_seed("review-invalid-status")
    await svc.review_copy_intelligence_seed(
        seed["seed_id"], action="APPROVE", reviewed_by="owner",
        review_note="note", confirmation_phrase="APPROVE COPY INTELLIGENCE",
    )
    with pytest.raises(svc.CopyIntelligenceReviewError) as exc:
        await svc.review_copy_intelligence_seed(
            seed["seed_id"], action="REJECT", reviewed_by="owner",
            review_note="note", confirmation_phrase="REJECT COPY INTELLIGENCE",
        )
    assert exc.value.code == "INVALID_SOURCE_STATUS"
    assert exc.value.status_code == 409
    assert (await crud.get_copy_intelligence_seed(seed["seed_id"]))["status"] == "APPROVED"


@pytest.mark.asyncio
async def test_review_does_not_create_snapshot_copy_set_or_touch_product_truth():
    """Approval transitions the seed ONLY — no materialization into generation."""
    from agent.db import crud

    product = await crud.create_product(
        source="MANUAL", raw_product_title="Truth", product_display_name="Truth",
        product_short_name="Truth", copywriting_angle="Manual",
    )
    before_product = await crud.get_product(product["id"])
    db = await crud.get_db()

    async def _count(table: str) -> int:
        cur = await db.execute(f"SELECT COUNT(*) FROM {table}")
        return (await cur.fetchone())[0]

    seed = await _make_needs_review_seed("review-isolation")
    products_before = await _count("product")
    copy_sets_before = await _count("copy_set")
    snapshots_before = await _count("product_intelligence_snapshot")
    seeds_before = await _count("copy_intelligence_seed")

    await svc.review_copy_intelligence_seed(
        seed["seed_id"], action="APPROVE", reviewed_by="owner",
        review_note="note", confirmation_phrase="APPROVE COPY INTELLIGENCE",
    )

    assert await _count("product") == products_before
    assert await _count("copy_set") == copy_sets_before
    assert await _count("product_intelligence_snapshot") == snapshots_before
    assert await _count("copy_intelligence_seed") == seeds_before  # transition, not insert
    assert await crud.get_product(product["id"]) == before_product


def test_invariant_generation_services_never_reference_seed_table():
    """The generation/compiler/DeepSeek services must not read the seed table."""
    from pathlib import Path

    repo_root = Path(__file__).resolve().parents[2]
    generation_files = [
        "agent/services/canonical_prompt_compiler.py",
        "agent/services/ai_copy_assist_service.py",
        "agent/services/ai_copy_provider_adapter.py",
        "agent/services/copy_grounding_service.py",
        "agent/services/copy_binding_service.py",
        "agent/services/workspace_execution_package_service.py",
    ]
    for rel in generation_files:
        text = (repo_root / rel).read_text(encoding="utf-8")
        assert "copy_intelligence_seed" not in text, f"{rel} must not reference the seed table"


# ── Approved Copy Intelligence safe-consumption layer ────────────────────────
async def _make_seed(fingerprint, *, status, confidence="HIGH",
                     target_product_id=None, reference_id=None, name="Produk"):
    from agent.db import crud

    return await crud.create_copy_intelligence_seed(
        source_fingerprint=fingerprint,
        source_workbook="fixture.xlsx", source_sheet="COPYWRITING HUB",
        source_row=2, source_product_name=name, match_method="TIKTOK_PRODUCT_ID_MATCH",
        confidence=confidence, status=status, target_product_id=target_product_id,
        reference_id=reference_id, hook_script="Hook", cta_script="CTA",
        provenance_json='{"sheet":"COPYWRITING HUB"}',
    )


@pytest.mark.asyncio
async def test_approved_context_returns_only_approved_rows():
    # Scope by a unique reference_id so the assertion is independent of any other
    # rows already present in the shared test DB.
    ref = "ref-onlyapproved"
    await _make_seed("ctx-approved", status="APPROVED", reference_id=ref, name="Approved product")
    await _make_seed("ctx-needs", status="NEEDS_REVIEW", reference_id=ref, name="Pending product")
    await _make_seed("ctx-rejected", status="REJECTED", reference_id=ref, name="Rejected product")

    result = await svc.get_approved_copy_intelligence_context(reference_id=ref, limit=100)
    names = {item["source_product_name"] for item in result["items"]}
    statuses = {item["status"] for item in result["items"]}
    assert names == {"Approved product"}
    assert statuses == {"APPROVED"}
    assert "Pending product" not in names
    assert "Rejected product" not in names
    assert result["total"] == 1


@pytest.mark.asyncio
async def test_approved_context_lookup_by_target_product_id():
    from agent.db import crud

    product = await crud.create_product(
        source="MANUAL", raw_product_title="Truth", product_display_name="Truth",
        product_short_name="Truth", copywriting_angle="Manual",
    )
    await _make_seed("ctx-tpid-hit", status="APPROVED", target_product_id=product["id"], name="Bound")
    await _make_seed("ctx-tpid-other", status="APPROVED", name="Unbound")

    result = await svc.get_approved_copy_intelligence_context(target_product_id=product["id"])
    assert {i["source_product_name"] for i in result["items"]} == {"Bound"}
    assert result["items"][0]["target_product_id"] == product["id"]


@pytest.mark.asyncio
async def test_approved_context_lookup_by_reference_id():
    await _make_seed("ctx-ref-hit", status="APPROVED", reference_id="fastmoss-ref:123", name="ByRef")
    await _make_seed("ctx-ref-miss", status="NEEDS_REVIEW", reference_id="fastmoss-ref:123", name="ByRefPending")

    result = await svc.get_approved_copy_intelligence_context(reference_id="fastmoss-ref:123")
    assert {i["source_product_name"] for i in result["items"]} == {"ByRef"}
    assert result["total"] == 1


@pytest.mark.asyncio
async def test_approved_context_empty_when_no_approved_rows():
    ref = "ref-emptyapproved"
    await _make_seed("ctx-only-pending", status="NEEDS_REVIEW", reference_id=ref, name="Pending")
    result = await svc.get_approved_copy_intelligence_context(reference_id=ref)
    assert result == {"total": 0, "items": []}


@pytest.mark.asyncio
async def test_approved_context_is_read_only_and_isolated():
    """Reading approved context mutates nothing — no product/copy_set/seed change."""
    from agent.db import crud

    product = await crud.create_product(
        source="MANUAL", raw_product_title="T", product_display_name="T",
        product_short_name="T", copywriting_angle="Manual",
    )
    await _make_seed("ctx-iso", status="APPROVED", name="Iso")
    db = await crud.get_db()

    async def _count(table):
        cur = await db.execute(f"SELECT COUNT(*) FROM {table}")
        return (await cur.fetchone())[0]

    before = {t: await _count(t) for t in ("product", "copy_set", "product_intelligence_snapshot", "copy_intelligence_seed")}
    before_product = await crud.get_product(product["id"])

    await svc.get_approved_copy_intelligence_context()
    await svc.get_approved_copy_intelligence_context(target_product_id=product["id"])

    after = {t: await _count(t) for t in before}
    assert before == after
    assert await crud.get_product(product["id"]) == before_product


def test_invariant_generation_services_never_reference_approved_context():
    """The approved-context reader lives in the import/ledger service, NOT in any
    generation/compiler service — generation still never reads the seed table."""
    from pathlib import Path

    repo_root = Path(__file__).resolve().parents[2]
    generation_files = [
        "agent/services/canonical_prompt_compiler.py",
        "agent/services/ai_copy_assist_service.py",
        "agent/services/ai_copy_provider_adapter.py",
        "agent/services/copy_grounding_service.py",
        "agent/services/copy_binding_service.py",
        "agent/services/workspace_execution_package_service.py",
    ]
    for rel in generation_files:
        text = (repo_root / rel).read_text(encoding="utf-8")
        assert "get_approved_copy_intelligence_context" not in text, f"{rel} must not consume approved context yet"
        assert "list_approved_copy_intelligence_seeds" not in text, f"{rel} must not query the seed table"
