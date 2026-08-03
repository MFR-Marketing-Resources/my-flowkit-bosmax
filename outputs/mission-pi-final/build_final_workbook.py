#!/usr/bin/env python
"""PI-FINAL Phase 10: regenerate the Master Product Recovery Workbook from the FINAL live DB.

Sheets: Recovery Master / Source Evidence / Claim Review / Canonical-Alias Reconciliation /
Exhaustion-Owner Decisions / Copy Eligibility / Summary / Deletion Candidates.
Reconciles every original residual product id (residual_manifest.json) against its final state.
"""
from __future__ import annotations
import hashlib, json, sqlite3, sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook

REPO = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO))
DB = REPO / "flow_agent.db"
OUT = Path(__file__).resolve().parent

HAS_APPROVED = ("EXISTS (SELECT 1 FROM product_intelligence_snapshot s2 "
                "WHERE s2.product_id = p.id AND s2.status = 'APPROVED')")
LATEST_READY = ("(SELECT s2.readiness_status FROM product_intelligence_snapshot s2 "
                "WHERE s2.product_id = p.id AND s2.status = 'APPROVED' ORDER BY s2.version DESC LIMIT 1)")
LATEST_COMPL = ("(SELECT s2.completeness_score FROM product_intelligence_snapshot s2 "
                "WHERE s2.product_id = p.id AND s2.status = 'APPROVED' ORDER BY s2.version DESC LIMIT 1)")
FIXTURE = ("(LOWER(TRIM(COALESCE(p.raw_product_title,''))) IN ('test product','test item','fixture product')"
           " OR LOWER(TRIM(COALESCE(p.product_short_name,''))) IN ('test product','test item','fixture product')"
           " OR LOWER(COALESCE(p.raw_product_title,'')) LIKE 'test product%'"
           " OR LOWER(COALESCE(p.raw_product_title,'')) LIKE 'smoke %'"
           " OR LOWER(COALESCE(p.raw_product_title,'')) LIKE '%smoke approve%'"
           " OR LOWER(COALESCE(p.raw_product_title,'')) LIKE '%smoke reject%'"
           " OR LOWER(COALESCE(p.raw_product_title,'')) LIKE '%smoke claim review%'"
           " OR LOWER(COALESCE(p.raw_product_title,'')) LIKE 'codex pi %verification%'"
           " OR LOWER(COALESCE(p.id,'')) LIKE 'test|_%' ESCAPE '|'"
           " OR LOWER(COALESCE(p.id,'')) LIKE 'fixture|_%' ESCAPE '|')")
ALIAS = "UPPER(COALESCE(p.archived_reason,'')) LIKE 'DUPLICATE_MERGED_TO_CANONICAL%'"
REAL = f"NOT {FIXTURE} AND NOT {ALIAS}"


def classify(row) -> str:
    if not row["has_snap"]:
        return "MISSING_APPROVED_INTELLIGENCE"
    if row["readiness"] == "READY_WITH_GOVERNED_ABSENCE":
        return "APPROVED_WITH_GOVERNED_ABSENCE"
    if (row["completeness"] or 0) >= 1.0:
        return "FULLY_COMPLETE"
    return "LEGACY_APPROVED_INCOMPLETE"


def main() -> int:
    con = sqlite3.connect(f"file:{DB.as_posix()}?mode=ro", uri=True)
    con.row_factory = sqlite3.Row
    manifest = json.load(open(OUT / "residual_manifest.json", encoding="utf-8"))
    evidence = {e["product_id"]: e for e in json.load(open(OUT / "evidence_all.json", encoding="utf-8"))}
    results = {r["product_id"]: r for r in json.load(open(OUT / "recovery_results.json", encoding="utf-8"))}
    adjudication = json.load(open(OUT / "b03_adjudication_results.json", encoding="utf-8"))

    wb = Workbook()

    # ── Recovery Master ──
    ws = wb.active; ws.title = "Recovery Master"
    ws.append(["product_id", "title", "class_at_baseline", "lifecycle", "lane", "final_class",
               "final_version", "final_readiness", "final_completeness", "final_claim_gate",
               "writer_result", "overlaid_fields", "governed_absent"])
    finals = {}
    for r in manifest["rows"]:
        pid = r["product_id"]
        f = con.execute(
            f"SELECT EXISTS(SELECT 1 FROM product_intelligence_snapshot s WHERE s.product_id=p.id "
            f"AND s.status='APPROVED') AS has_snap, {LATEST_READY} AS readiness, "
            f"{LATEST_COMPL} AS completeness, "
            "(SELECT s2.claim_gate FROM product_intelligence_snapshot s2 WHERE s2.product_id=p.id "
            " AND s2.status='APPROVED' ORDER BY s2.version DESC LIMIT 1) AS gate, "
            "(SELECT MAX(s2.version) FROM product_intelligence_snapshot s2 WHERE s2.product_id=p.id "
            " AND s2.status='APPROVED') AS version "
            "FROM product p WHERE p.id=?", (pid,)).fetchone()
        final_class = classify(f)
        finals[pid] = final_class
        res = results.get(pid, {})
        ws.append([pid, r["raw_product_title"][:120], r["class"], r["lifecycle"], r["lane"],
                   final_class, f["version"], f["readiness"], f["completeness"], f["gate"],
                   res.get("result", ""), ",".join(res.get("overlaid", [])),
                   ",".join(evidence.get(pid, {}).get("governed_absent", []))])

    # ── Source Evidence ──
    ws = wb.create_sheet("Source Evidence")
    ws.append(["product_id", "field", "status", "value", "source_url", "rationale/excerpt"])
    for pid, ev in evidence.items():
        for fname, spec in (ev.get("fields") or {}).items():
            ws.append([pid, fname, spec.get("status"),
                       json.dumps(spec.get("value"), ensure_ascii=False)[:300],
                       spec.get("source_url") or ev.get("best_source_url"),
                       str(spec.get("rationale") or spec.get("excerpt") or "")[:300]])
        for src in ((ev.get("research_log") or {}).get("sources") or []):
            ws.append([pid, "(research-log)", src.get("outcome"), "", src.get("url"),
                       str(src.get("note") or "")[:200]])

    # ── Claim Review ──
    ws = wb.create_sheet("Claim Review")
    ws.append(["product_id", "title", "prior_gate", "result", "final_gate", "final_version", "note"])
    for r in adjudication:
        ws.append([r.get("product_id"), "", r.get("prior_gate", ""), r.get("result"),
                   r.get("gate", ""), r.get("version", ""), json.dumps(r, ensure_ascii=False)[:250]])
    remaining = con.execute(
        f"SELECT COUNT(*) FROM product p WHERE {REAL} AND {HAS_APPROVED} AND "
        "(SELECT s2.claim_gate FROM product_intelligence_snapshot s2 WHERE s2.product_id=p.id "
        " AND s2.status='APPROVED' ORDER BY s2.version DESC LIMIT 1) = 'CLAIM_BLOCKED'").fetchone()[0]
    ws.append(["(remaining CLAIM_BLOCKED latest-approved on canonical real products)", "", "", remaining])

    # ── Canonical-Alias Reconciliation ──
    ws = wb.create_sheet("Canonical-Alias Reconciliation")
    ws.append(["alias_product_id", "title", "archived_reason", "lifecycle"])
    for row in con.execute(
            f"SELECT p.id, p.raw_product_title, p.archived_reason, p.lifecycle_status "
            f"FROM product p WHERE {ALIAS} ORDER BY p.id"):
        ws.append([row[0], (row[1] or "")[:120], (row[2] or "")[:120], row[3]])

    # ── Exhaustion / Owner Decisions ──
    ws = wb.create_sheet("Exhaustion-Owner Decisions")
    ws.append(["product_id", "title", "status", "detail"])
    exhausted = [pid for pid, res in results.items()
                 if res.get("result") not in ("APPROVED",)]
    for pid in exhausted:
        r = results[pid]
        ws.append([pid, next((m["raw_product_title"][:120] for m in manifest["rows"]
                              if m["product_id"] == pid), ""),
                   r.get("result"), json.dumps(r, ensure_ascii=False)[:400]])
    if not exhausted:
        ws.append(["(none)", "every original residual product was recovered", "", ""])

    # ── Copy Eligibility ──
    ws = wb.create_sheet("Copy Eligibility")
    ws.append(["product_id", "lifecycle", "eligible", "reasons"])
    import asyncio
    from agent.services.copy_eligibility_service import copy_eligibility
    from agent.db import schema as _schema

    async def _elig_all():
        rows = con.execute(
            f"SELECT p.id, p.lifecycle_status FROM product p WHERE {REAL} ORDER BY p.id").fetchall()
        out = []
        for pr in rows:
            rec = await copy_eligibility(pr[0])
            out.append((pr[0], pr[1], rec["eligible"], ",".join(rec["reasons"])))
        await _schema.close_db()
        return out

    for tup in asyncio.run(_elig_all()):
        ws.append(list(tup))

    # ── Summary ──
    ws = wb.create_sheet("Summary")
    counts = {}
    for name, pred in (
            ("MISSING_APPROVED_INTELLIGENCE", f"NOT {HAS_APPROVED}"),
            ("APPROVED_WITH_GOVERNED_ABSENCE", f"({HAS_APPROVED} AND {LATEST_READY} = 'READY_WITH_GOVERNED_ABSENCE')"),
            ("FULLY_COMPLETE", f"({HAS_APPROVED} AND COALESCE({LATEST_READY},'') <> 'READY_WITH_GOVERNED_ABSENCE' AND COALESCE({LATEST_COMPL},0) >= 1.0)"),
            ("LEGACY_APPROVED_INCOMPLETE", f"({HAS_APPROVED} AND COALESCE({LATEST_READY},'') <> 'READY_WITH_GOVERNED_ABSENCE' AND COALESCE({LATEST_COMPL},0) < 1.0)")):
        counts[name] = con.execute(f"SELECT COUNT(*) FROM product p WHERE {REAL} AND {pred}").fetchone()[0]
    physical = con.execute("SELECT COUNT(*) FROM product p").fetchone()[0]
    fixtures = con.execute(f"SELECT COUNT(*) FROM product p WHERE {FIXTURE}").fetchone()[0]
    aliases = con.execute(f"SELECT COUNT(*) FROM product p WHERE {ALIAS}").fetchone()[0]
    canonical = con.execute(f"SELECT COUNT(*) FROM product p WHERE {REAL}").fetchone()[0]
    for k, v in [("generated_at", datetime.now(timezone.utc).isoformat()),
                 ("physical_rows", physical), ("test_fixtures", fixtures),
                 ("merged_aliases", aliases), ("canonical_real", canonical),
                 *counts.items(),
                 ("classified_total", sum(counts.values()))]:
        ws.append([k, v])

    # ── Deletion Candidates ──
    ws = wb.create_sheet("Deletion Candidates")
    ws.append(["product_id", "title", "exhaustion_record"])
    # remains empty unless exhaustive proof supports candidates

    path = OUT / "Master_Product_Recovery_FINAL.xlsx"
    wb.save(path)
    sha = hashlib.sha256(path.read_bytes()).hexdigest()
    (OUT / "Master_Product_Recovery_FINAL.sha256").write_text(sha + "\n")
    meta = {"path": str(path), "sha256": sha,
            "residual_ids": len(manifest["rows"]),
            "final_classes_of_residuals": {},
            "summary_counts": {"physical": physical, "fixtures": fixtures,
                               "aliases": aliases, "canonical": canonical, **counts}}
    for pid, cls in finals.items():
        meta["final_classes_of_residuals"][cls] = meta["final_classes_of_residuals"].get(cls, 0) + 1
    print(json.dumps(meta, indent=1))
    return 0


if __name__ == "__main__":
    sys.exit(main())
