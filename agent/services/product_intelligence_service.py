from __future__ import annotations

import re
from collections import Counter, defaultdict
from functools import lru_cache
from typing import Any

from openpyxl import load_workbook

from agent.config import OPERATOR_PACK_DIR
from agent.db import crud
from agent.models.product_intelligence import (
    ProductIntelligenceBackfillPreviewResponse,
    ProductIntelligenceProfile,
    ProductIntelligenceResolveRequest,
    ProductIntelligenceSalesMetrics,
    ProductIntelligenceSummaryResponse,
)
from agent.services.bosmax_product_family import derive_bosmax_product_family
from agent.services.fastmoss_import_service import get_latest_fastmoss_reference_index
from agent.services.product_lifecycle_service import lifecycle_status as resolve_lifecycle_status
from agent.services.product_image_analysis_service import analyze_product_image_payload
from agent.services.product_mapping import normalize_mapping_text


REVIEW_CLAIM_TOKENS = {
    "antibakteria",
    "antibaktiria",
    "antibacterial",
    "anti bacterial",
    "antibacteria",
    "whitening",
    "white",
    "mencerahkan",
    "brightening",
    "anti aging",
    "anti-aging",
    "anti jerawat",
    "jerawat",
    "acne",
    "eczema",
    "resdung",
    "supplement",
    "supplements",
    "vitamin",
    "capsule",
    "capsules",
    "detox",
    "slimming",
    "kurus",
    "fat burner",
    "weight loss",
    "pain relief",
    "relief",
    "wellness",
    "immune",
    "hair growth",
    "growth",
    "anti gugur",
    "gugur",
    "medical",
    "health",
    "tenaga batin",
    "batin lelaki",
    "bahagian intim",
    "intim lelaki",
    "ketegangan",
    "kelelakian",
    "stamina lelaki",
    "prestasi fizikal lelaki",
    "keyakinan kelelakian",
    "otot kelelakian",
    "jamu perapat",
    "jamu wanita",
    "kewanitaan",
    "miss v",
    "faraj",
    "vagina",
    "keputihan",
    "bau",
    "gatal",
    "rapat",
    "ketat",
    "anjal",
    "postpartum",
    "selepas bersalin",
    "intimate",
    "feminine hygiene",
    "feminine care",
    "perapat",
}
BLOCKED_CLAIM_TOKENS = {
    "cure",
    "cures",
    "menyembuhkan",
    "merawat",
    "membesarkan",
    "memanjangkan",
    "mati pucuk",
    "fertility claim",
    "hormone claim",
    "infection treatment",
    "fungus treatment",
}

# PI-RECOMPUTE-20260803-01.  These are deliberately versioned at the queue
# layer as part of the persisted recompute ruleset.  Keep the names stable:
# they are operator-facing review reason codes, not provider output.
UNUSUAL_SPEC_CLAIM_RULES = (
    ("unusual_wattage", re.compile(r"(?<!\w)(\d[\d,.]*)\s*(?:w|watt|watts)\b", re.IGNORECASE), 10_000),
    ("unusual_battery_duration", re.compile(r"(?<!\w)(\d[\d,.]*)\s*(?:jam|hours?|hrs?)\b", re.IGNORECASE), 48),
)
UNUSUAL_WARRANTY_PATTERN = re.compile(
    r"(?:(?:waranti|warranty|jaminan)[^0-9]{0,24}(\d[\d,.]*)\s*(?:tahun|years?)\b|"
    r"(\d[\d,.]*)\s*(?:tahun|years?)[^.;|]{0,24}(?:waranti|warranty|jaminan)\b)",
    re.IGNORECASE,
)


FAMILY_PROFILES: dict[str, dict[str, str]] = {
    "LAUNDRY_DETERGENT_LIQUID_REFILL": {
        "group": "LAUNDRY_CARE",
        "sub_group": "LAUNDRY_CARE",
        "type_of_product": "LIQUID_LAUNDRY_DETERGENT",
        "package_form": "bottle_or_refill_pack",
        "physical_state": "liquid",
        "product_scale_class": "liquid_bottle_or_refill_pack",
        "handling_profile": "stable bottle/refill grip, cap/nozzle/label visibility, pour-angle demonstration",
        "scene_profile": "laundry_routine_utility_demo",
        "camera_profile": "label_forward_pour_ready_ugc",
        "copy_route": "DIRECT",
        "copy_formula": "UTILITY_DEMO",
    },
    "FABRIC_SOFTENER_LIQUID": {
        "group": "LAUNDRY_CARE",
        "sub_group": "LAUNDRY_CARE",
        "type_of_product": "LIQUID_FABRIC_SOFTENER",
        "package_form": "bottle_or_refill_pack",
        "physical_state": "liquid",
        "product_scale_class": "liquid_bottle_or_refill_pack",
        "handling_profile": "stable bottle/refill grip, cap/nozzle/label visibility, pour-angle demonstration",
        "scene_profile": "laundry_routine_utility_demo",
        "camera_profile": "label_forward_pour_ready_ugc",
        "copy_route": "DIRECT",
        "copy_formula": "SOFTNESS_ROUTINE",
    },
    "HOUSEHOLD_CLEANER_GENERAL": {
        "group": "HOUSEHOLD_CARE",
        "sub_group": "HOUSEHOLD_CARE",
        "type_of_product": "HOUSEHOLD_CLEANER",
        "package_form": "bottle_or_refill_pack",
        "physical_state": "liquid",
        "product_scale_class": "utility_container",
        "handling_profile": "grip_trigger_or_cap_label_visibility",
        "scene_profile": "household_cleaning_demo",
        "camera_profile": "utility_closeup_function_demo",
        "copy_route": "DIRECT",
        "copy_formula": "UTILITY_DEMO",
    },
    "HOUSEHOLD_STORAGE_ORGANIZER": {
        "group": "HOME_ORGANIZATION",
        "sub_group": "STORAGE_ORGANIZER",
        "type_of_product": "HOME_STORAGE_ORGANIZER",
        "package_form": "rigid_container",
        "physical_state": "solid",
        "product_scale_class": "medium_rigid_object",
        "handling_profile": "two_hand_open_close_shape_visibility",
        "scene_profile": "organization_before_after_demo",
        "camera_profile": "countertop_reveal_stackability_demo",
        "copy_route": "DIRECT",
        "copy_formula": "ORGANIZATION_UTILITY",
    },
    "HOME_TEXTILE": {
        "group": "HOUSEHOLD_CARE",
        "sub_group": "HOME_TEXTILE",
        "type_of_product": "HOME_TEXTILE",
        "package_form": "folded_textile",
        "physical_state": "textile",
        "product_scale_class": "large_soft_good",
        "handling_profile": "spread_fold_drape_texture_visibility",
        "scene_profile": "home_textile_texture_demo",
        "camera_profile": "texture_closeup_with_broad_reveal",
        "copy_route": "DIRECT",
        "copy_formula": "TEXTURE_COMFORT",
    },
    "APPAREL_SLEEPWEAR": {
        "group": "FASHION_AND_APPAREL",
        "sub_group": "SLEEPWEAR_AND_LOUNGEWEAR",
        "type_of_product": "SLEEPWEAR",
        "package_form": "garment",
        "physical_state": "textile",
        "product_scale_class": "wearable_garment",
        "handling_profile": "drape_seam_shoulder_hanger_visibility",
        "scene_profile": "relaxed_homewear_demo",
        "camera_profile": "fabric_fall_and_fit_demo",
        "copy_route": "DIRECT",
        "copy_formula": "COMFORT_STYLE",
    },
    "fashion_modestwear": {
        "group": "FASHION_AND_APPAREL",
        "sub_group": "MODESTWEAR",
        "type_of_product": "MODESTWEAR",
        "package_form": "garment",
        "physical_state": "textile",
        "product_scale_class": "wearable_garment",
        "handling_profile": "drape_coverage_edge_visibility",
        "scene_profile": "modestwear_styling_demo",
        "camera_profile": "coverage_and_texture_reveal",
        "copy_route": "DIRECT",
        "copy_formula": "MODEST_STYLE",
    },
    "fashion_sportswear": {
        "group": "FASHION_AND_APPAREL",
        "sub_group": "SPORTSWEAR",
        "type_of_product": "SPORTSWEAR",
        "package_form": "garment",
        "physical_state": "textile",
        "product_scale_class": "wearable_garment",
        "handling_profile": "fit_seam_texture_visibility",
        "scene_profile": "activewear_styling_demo",
        "camera_profile": "fit_and_texture_reveal",
        "copy_route": "DIRECT",
        "copy_formula": "FIT_STYLE",
    },
    "fashion_apparel": {
        "group": "FASHION_AND_APPAREL",
        "sub_group": "GENERAL_APPAREL",
        "type_of_product": "APPAREL",
        "package_form": "garment",
        "physical_state": "textile",
        "product_scale_class": "wearable_garment",
        "handling_profile": "drape_fold_seam_visibility",
        "scene_profile": "fashion_styling_demo",
        "camera_profile": "fit_and_texture_reveal",
        "copy_route": "DIRECT",
        "copy_formula": "STYLE_DIRECT",
    },
    "BEAUTY_PERSONAL_CARE": {
        "group": "BEAUTY_AND_PERSONAL_CARE",
        "sub_group": "PERSONAL_CARE",
        "type_of_product": "BEAUTY_PERSONAL_CARE_PRODUCT",
        "package_form": "small_bottle_tube_or_compact",
        "physical_state": "liquid_or_semi_liquid",
        "product_scale_class": "small_handheld",
        "handling_profile": "small_bottle_cap_label_closeup_visibility",
        "scene_profile": "beauty_routine_demo",
        "camera_profile": "closeup_handheld_detail_demo",
        "copy_route": "DIRECT",
        "copy_formula": "ROUTINE_BEAUTY",
    },
    "beauty_fragrance": {
        "group": "BEAUTY_AND_PERSONAL_CARE",
        "sub_group": "FRAGRANCE",
        "type_of_product": "FRAGRANCE",
        "package_form": "small_bottle_or_mist",
        "physical_state": "liquid",
        "product_scale_class": "small_handheld",
        "handling_profile": "small_bottle_nozzle_label_visibility",
        "scene_profile": "fragrance_closeup_demo",
        "camera_profile": "reflective_bottle_reveal",
        "copy_route": "DIRECT",
        "copy_formula": "FRESHNESS_DIRECT",
    },
    "ACCESSORY_SMALL_ITEM": {
        "group": "ACCESSORIES_AND_SMALL_ITEMS",
        "sub_group": "SMALL_ACCESSORY",
        "type_of_product": "ACCESSORY",
        "package_form": "small_rigid_item",
        "physical_state": "solid",
        "product_scale_class": "small_fingertip",
        "handling_profile": "pinch_edge_detail_visibility",
        "scene_profile": "styling_closeup_demo",
        "camera_profile": "macro_detail_reveal",
        "copy_route": "DIRECT",
        "copy_formula": "DETAIL_STYLE",
    },
    "BABY_DIAPER": {
        "group": "BABY_AND_MATERNITY",
        "sub_group": "BABY_DIAPERING",
        "type_of_product": "BABY_DIAPER",
        "package_form": "soft_pack",
        "physical_state": "soft_packaged_goods",
        "product_scale_class": "medium_soft_pack",
        "handling_profile": "front_pack_support_label_visibility",
        "scene_profile": "babycare_trust_demo",
        "camera_profile": "front_pack_reveal",
        "copy_route": "DIRECT",
        "copy_formula": "TRUST_BABYCARE",
    },
    "BABY_WIPES": {
        "group": "BABY_AND_MATERNITY",
        "sub_group": "BABY_HYGIENE",
        "type_of_product": "BABY_WIPES",
        "package_form": "soft_pack",
        "physical_state": "soft_packaged_goods",
        "product_scale_class": "small_soft_pack",
        "handling_profile": "front_pack_seal_visibility",
        "scene_profile": "babycare_trust_demo",
        "camera_profile": "front_pack_reveal",
        "copy_route": "DIRECT",
        "copy_formula": "TRUST_BABYCARE",
    },
    "food_packaged": {
        "group": "FOOD_AND_BEVERAGE",
        "sub_group": "PACKAGED_FOOD",
        "type_of_product": "PACKAGED_FOOD_OR_SAUCE",
        "package_form": "jar_sachet_or_food_pack",
        "physical_state": "solid_or_sauce",
        "product_scale_class": "small_food_pack",
        "handling_profile": "sealed_pack_front_label_visibility",
        "scene_profile": "food_serving_demo",
        "camera_profile": "appetite_led_pack_reveal",
        "copy_route": "DIRECT",
        "copy_formula": "TASTE_CONVENIENCE",
    },
    "stationery_paper": {
        "group": "STATIONERY_AND_GIFTING",
        "sub_group": "PAPER_PACKET",
        "type_of_product": "PAPER_PACKET_OR_ENVELOPE",
        "package_form": "flat_packet",
        "physical_state": "paper",
        "product_scale_class": "small_flat_packet",
        "handling_profile": "flat_packet_pinch_edge_visibility",
        "scene_profile": "flatlay_or_fanout_demo",
        "camera_profile": "topdown_or_macro_paper_reveal",
        "copy_route": "DIRECT",
        "copy_formula": "DETAIL_GIFTING",
    },
    "electronics_wearable": {
        "group": "ELECTRONICS_AND_GADGETS",
        "sub_group": "WEARABLE_DEVICE",
        "type_of_product": "WEARABLE_ELECTRONIC",
        "package_form": "small_rigid_device",
        "physical_state": "solid",
        "product_scale_class": "small_handheld",
        "handling_profile": "device_screen_port_visibility",
        "scene_profile": "tech_closeup_demo",
        "camera_profile": "feature_detail_reveal",
        "copy_route": "DIRECT",
        "copy_formula": "FEATURE_UTILITY",
    },
    "HEALTH_SUPPLEMENT": {
        "group": "HEALTH_AND_WELLNESS",
        "sub_group": "SUPPLEMENT",
        "type_of_product": "SUPPLEMENT_BOTTLE",
        "package_form": "small_bottle",
        "physical_state": "solid_capsule_or_powder",
        "product_scale_class": "small_handheld",
        "handling_profile": "bottle_cap_label_visibility",
        "scene_profile": "wellness_routine_demo",
        "camera_profile": "trust_led_closeup",
        "copy_route": "REVIEW_REQUIRED",
        "copy_formula": "WELLNESS_REVIEW",
    },
    "MALE_HEALTH_SENSITIVE": {
        "group": "MALE_HEALTH_SENSITIVE",
        "sub_group": "MALE_HEALTH_SENSITIVE",
        "type_of_product": "SENSITIVE_MALE_HEALTH_PRODUCT",
        "package_form": "small_bottle_or_box",
        "physical_state": "solid_or_liquid_container",
        "product_scale_class": "small_handheld",
        "handling_profile": "bottle_box_label_visibility",
        "scene_profile": "literal_product_demo_review_required",
        "camera_profile": "literal_product_closeup_review_required",
        "copy_route": "STEALTH",
        "copy_formula": "STEALTH_DIALOGUE_SAFE",
    },
    "FEMALE_HEALTH_SENSITIVE": {
        "group": "FEMALE_HEALTH_SENSITIVE",
        "sub_group": "FEMALE_HEALTH_SENSITIVE",
        "type_of_product": "SENSITIVE_FEMALE_HEALTH_PRODUCT",
        "package_form": "small_bottle_tube_or_jar",
        "physical_state": "solid_or_liquid_container",
        "product_scale_class": "small_handheld",
        "handling_profile": "bottle_jar_label_visibility",
        "scene_profile": "literal_product_demo_review_required",
        "camera_profile": "literal_product_closeup_review_required",
        "copy_route": "STEALTH",
        "copy_formula": "STEALTH_DIALOGUE_SAFE",
    },
    "PET_CARE_GENERAL": {
        "group": "PET_CARE",
        "sub_group": "PET_CARE",
        "type_of_product": "PET_CARE_PRODUCT",
        "package_form": "bag_pack_or_can",
        "physical_state": "solid_or_kibble",
        "product_scale_class": "small_to_medium_pack",
        "handling_profile": "front_pack_label_visibility",
        "scene_profile": "petcare_product_demo",
        "camera_profile": "pack_reveal_and_detail",
        "copy_route": "DIRECT",
        "copy_formula": "PETCARE_DIRECT",
    },
    "AUTO_TOOL_GENERAL": {
        "group": "AUTO_AND_TOOLS",
        "sub_group": "AUTO_AND_TOOLS",
        "type_of_product": "AUTO_OR_TOOL_ITEM",
        "package_form": "rigid_tool_or_pack",
        "physical_state": "solid",
        "product_scale_class": "small_to_medium_tool",
        "handling_profile": "function_grip_visibility",
        "scene_profile": "utility_tool_demo",
        "camera_profile": "feature_function_reveal",
        "copy_route": "DIRECT",
        "copy_formula": "UTILITY_DIRECT",
    },
    "toy_play": {
        "group": "TOYS_AND_HOBBIES",
        "sub_group": "HOBBY_AND_CRAFT",
        "type_of_product": "CRAFT_OR_TOY_ITEM",
        "package_form": "box_or_bag",
        "physical_state": "solid",
        "product_scale_class": "medium_rigid_object",
        "handling_profile": "two_hand_hold_detail_visibility",
        "scene_profile": "hobby_craft_routine_demo",
        "camera_profile": "closeup_process_detail_demo",
        "copy_route": "DIRECT",
        "copy_formula": "CREATIVE_PLAY",
    },
    "REAL_ESTATE_OR_SERVICE": {
        "group": "REAL_ESTATE_OR_SERVICE",
        "sub_group": "REAL_ESTATE_OR_SERVICE",
        "type_of_product": "SERVICE_OR_INTANGIBLE",
        "package_form": "not_applicable",
        "physical_state": "not_applicable",
        "product_scale_class": "not_applicable",
        "handling_profile": "review_required",
        "scene_profile": "review_required",
        "camera_profile": "review_required",
        "copy_route": "REVIEW_REQUIRED",
        "copy_formula": "REVIEW_REQUIRED",
    },
    "UNKNOWN_REVIEW_REQUIRED": {
        "group": "UNKNOWN_REVIEW_REQUIRED",
        "sub_group": "UNKNOWN_REVIEW_REQUIRED",
        "type_of_product": "UNKNOWN_REVIEW_REQUIRED",
        "package_form": "unknown",
        "physical_state": "unknown",
        "product_scale_class": "unknown",
        "handling_profile": "review_required",
        "scene_profile": "review_required",
        "camera_profile": "review_required",
        "copy_route": "REVIEW_REQUIRED",
        "copy_formula": "REVIEW_REQUIRED",
    },
}


def _normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _normalize_key(value: Any) -> str:
    return normalize_mapping_text(value)


def _unique(items: list[str]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for item in items:
        normalized = _normalize_text(item)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        ordered.append(normalized)
    return ordered


def _joined_title_text(product: dict[str, Any]) -> str:
    return " ".join(
        _normalize_key(product.get(field))
        for field in (
            "raw_product_title",
            "product_display_name",
            "product_short_name",
            "product_type",
            "brand",
            "product_knowledge_text",
            "benefits",
            "usage",
            "target_customer",
            "warnings",
            "ingredients",
            "package_notes",
        )
        if product.get(field)
    )


def _joined_product_text(product: dict[str, Any]) -> str:
    return " ".join(
        _normalize_key(product.get(field))
        for field in (
            "raw_product_title",
            "product_display_name",
            "product_short_name",
            "category",
            "subcategory",
            "type",
            "product_type",
            "brand",
            "product_knowledge_text",
            "benefits",
            "usage",
            "target_customer",
            "warnings",
            "ingredients",
            "package_notes",
        )
        if product.get(field)
    )


def _contains_any(haystack: str, keywords: list[str]) -> bool:
    return any(
        bool(
            normalized_keyword
            and re.search(
                rf"(?<!\w){re.escape(normalized_keyword)}(?!\w)",
                haystack,
            )
        )
        for keyword in keywords
        if (normalized_keyword := _normalize_key(keyword))
    )


def _claim_body_text(product: dict[str, Any]) -> str:
    """Return claim-bearing text without treating taxonomy labels as claims.

    Category labels such as ``Health`` and ``Home Supplies`` are context for
    claim review.  They are not, by themselves, a claim.  The previous bulk
    path mixed those labels into a substring scan, which made a category label
    indistinguishable from product copy.
    """
    fields = (
        "raw_product_title",
        "product_name",
        "product_display_name",
        "product_short_name",
        "brand",
        "product_knowledge_text",
        "product_knowledge",
        "benefits",
        "benefits_text",
        "usage",
        "usage_text",
        "target_customer",
        "target_customer_text",
        "warnings",
        "warnings_text",
        "ingredients",
        "ingredients_text",
        "package_notes",
        "paste_anything_about_product",
    )
    return " ".join(
        _normalize_key(product.get(field))
        for field in fields
        if product.get(field)
    )


def _claim_taxonomy_text(product: dict[str, Any]) -> str:
    return " ".join(
        _normalize_key(product.get(field))
        for field in ("category", "subcategory", "type", "product_type", "product_type_id")
        if product.get(field)
    )


def _claim_token_matches(text: str, token: str) -> bool:
    normalized_token = _normalize_key(token)
    return bool(
        normalized_token
        and re.search(
            rf"(?<!\w){re.escape(normalized_token)}(?!\w)",
            text,
        )
    )


def _claim_context(product: dict[str, Any], family: str | None) -> dict[str, bool]:
    body = _claim_body_text(product)
    taxonomy = _claim_taxonomy_text(product)
    full = f"{body} {taxonomy}".strip()
    fashion = bool(
        family
        and (
            family.startswith("fashion")
            or family.startswith("APPAREL")
        )
    ) or _contains_any(
        full,
        [
            "fashion",
            "apparel",
            "clothing",
            "garment",
            "menswear",
            "womenswear",
            "underwear",
            "seluar",
            "pants",
            "baju",
            "leggings",
            "socks",
            "stokin",
            "textile",
        ],
    )
    health_or_beauty = bool(
        family
        and family
        in {
            "MALE_HEALTH_SENSITIVE",
            "FEMALE_HEALTH_SENSITIVE",
            "HEALTH_SUPPLEMENT",
            "BEAUTY_PERSONAL_CARE",
            "beauty_fragrance",
        }
    ) or _contains_any(
        taxonomy,
        [
            "health",
            "supplement",
            "medical",
            "beauty",
            "personal care",
            "cosmetics",
            "skincare",
            "oral care",
        ],
    )
    plant_or_pest = _contains_any(
        full,
        [
            "herbicide",
            "pesticide",
            "weed killer",
            "weed",
            "grass killer",
            "grass",
            "rumpai",
            "racun herba",
            "racun rumpai",
            "fertilizer",
            "baja",
            "plant growth",
            "tumbuhan",
            "garden",
        ],
    )
    neutral_odor = _contains_any(
        full,
        [
            "air freshener",
            "car perfume",
            "room perfume",
            "fragrance",
            "perfume",
            "deodorizer",
            "shoe rack",
            "socks",
            "stokin",
        ],
    )
    body_odor_claim = _contains_any(
        body,
        [
            "bau mulut",
            "bau ketiak",
            "bau badan",
            "bau kaki",
            "bad breath",
            "underarm",
            "deodorant",
            "toothpaste",
            "tooth cleaning",
            "mouth spray",
            "oral",
            "gigi",
            "nafas segar",
        ],
    )
    fashion_fit_claim = _contains_any(
        body,
        [
            "kulit",
            "skin",
            "payudara",
            "breast",
            "estrogen",
            "intim",
            "faraj",
            "vagina",
        ],
    )
    # Genuinely intimate/sensitive body cues — deliberately EXCLUDES the generic
    # "kulit"/"skin", which appears in innocuous safety text ("boleh memotong
    # kulit" on a fishing line) and must not gate a product into sensitive review.
    intimate_body_claim = _contains_any(
        body,
        [
            "payudara",
            "breast",
            "estrogen",
            "intim",
            "faraj",
            "vagina",
            "miss v",
            "kewanitaan",
            "keputihan",
            "kegel",
            "postpartum",
        ],
    )
    return {
        "fashion": fashion,
        "health_or_beauty": health_or_beauty,
        "plant_or_pest": plant_or_pest,
        "neutral_odor": neutral_odor,
        "body_odor_claim": body_odor_claim,
        "fashion_fit_claim": fashion_fit_claim,
        "intimate_body_claim": intimate_body_claim,
    }


def _parse_claim_number(value: str) -> float | None:
    try:
        return float(value.replace(",", ""))
    except (TypeError, ValueError):
        return None


def _unusual_spec_claim_tokens(product: dict[str, Any]) -> list[str]:
    raw_text = " ".join(
        str(product.get(field) or "")
        for field in (
            "raw_product_title",
            "product_name",
            "product_display_name",
            "product_short_name",
            "product_knowledge_text",
            "product_knowledge",
            "benefits",
            "benefits_text",
            "usage",
            "usage_text",
            "package_notes",
            "paste_anything_about_product",
        )
    )
    tokens: list[str] = []
    for token, pattern, threshold in UNUSUAL_SPEC_CLAIM_RULES:
        for match in pattern.finditer(raw_text):
            value = _parse_claim_number(match.group(1))
            if value is not None and value >= threshold:
                tokens.append(token)
                break
    for match in UNUSUAL_WARRANTY_PATTERN.finditer(raw_text):
        value = _parse_claim_number(match.group(1) or match.group(2))
        if value is not None and value >= 5:
            tokens.append("unusual_warranty_duration")
            break
    return tokens


def evaluate_product_claims(
    product: dict[str, Any],
    *,
    family: str | None = None,
) -> tuple[list[str], list[str], list[str]]:
    """Evaluate claims with source/category context and explicit boundaries.

    Returns ``(blocked_tokens, review_tokens, warnings)``.  This is the one
    deterministic matcher used by both Product Intelligence and Smart
    Registration completion, so queue recompute cannot drift from the review
    draft classifier.
    """
    body = _claim_body_text(product)
    context = _claim_context(product, family)
    matched_blocked = sorted(
        token for token in BLOCKED_CLAIM_TOKENS if _claim_token_matches(body, token)
    )
    matched_review: list[str] = []
    for token in REVIEW_CLAIM_TOKENS:
        if not _claim_token_matches(body, token):
            continue
        if token == "white":
            # Colour/product-name usage is not a whitening claim.  Explicit
            # complexion, skin or teeth language remains review-gated.
            white_claim = bool(
                re.search(
                    r"(?:white|putih).{0,42}(?:skin|kulit|teeth|gigi|complexion|mencerah|pencerah|pemutih)|"
                    r"(?:skin|kulit|teeth|gigi|complexion).{0,42}(?:white|putih)",
                    body,
                )
            )
            if not white_claim:
                continue
        elif token in {"anjal", "ketat", "rapat", "ketegangan"}:
            # Generic tightness / tension / elasticity words only carry a
            # sensitive-health meaning inside a health, beauty, or genuinely
            # intimate context.  A tight braided fishing line, a snug seal, a
            # taut rope, or a fitted garment (whose warning says it can cut the
            # "kulit") uses these words literally and must not be gated into
            # sensitive claim review.
            if not context["health_or_beauty"] and not context["intimate_body_claim"]:
                continue
        elif token == "growth":
            if context["plant_or_pest"] and not context["health_or_beauty"]:
                continue
        elif token == "bau":
            if context["neutral_odor"] and not context["body_odor_claim"]:
                continue
        matched_review.append(token)

    for token in _unusual_spec_claim_tokens(product):
        if token not in matched_review:
            matched_review.append(token)

    warnings: list[str] = []
    if matched_blocked:
        warnings.append("claim_context:blocked_claim_text")
    if matched_review:
        warnings.append("claim_context:human_review_required")
    if any(token.startswith("unusual_") for token in matched_review):
        warnings.append("claim_context:unusual_specification")
    return matched_blocked, sorted(matched_review), warnings


def _first_non_empty(*values: Any) -> str | None:
    for value in values:
        text = _normalize_text(value)
        if text:
            return text
    return None


def _to_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    text = _normalize_text(value).replace(",", "")
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _coerce_source(value: Any) -> str:
    source = _normalize_text(value).upper()
    if source in {"FASTMOSS", "MANUAL", "TIKTOKSHOP", "TEST"}:
        return source
    if not source:
        return "UNKNOWN"
    return source


def _product_names(product: dict[str, Any]) -> list[str]:
    return _unique(
        [
            product.get("raw_product_title"),
            product.get("product_display_name"),
            product.get("product_short_name"),
        ]
    )


def _find_sheet_header(ws) -> list[str]:
    for row in ws.iter_rows(values_only=True, max_row=10):
        values = [str(value).strip() if value is not None else "" for value in row]
        lowered = {value.lower() for value in values if value}
        if (
            "product name" in lowered
            or "product title" in lowered
            or "rank" in lowered
        ):
            return values
    return []


def _row_mapping(headers: list[str], values: list[Any]) -> dict[str, Any]:
    return {
        headers[index]: values[index]
        for index in range(min(len(headers), len(values)))
        if headers[index]
    }


def _iter_sales_workbook_records() -> list[dict[str, Any]]:
    path = OPERATOR_PACK_DIR / "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx"
    if not path.exists():
        return []

    configs = [
        {
            "sheet": "Product Sales Rank",
            "name_fields": ["Product Name"],
            "shop_fields": ["Shop Name"],
            "source_fields": ["FastMoss Product Detail", "FastMoss Shop Detail"],
            "tiktok_fields": ["TikTok Product Detail"],
            "metric_fields": [
                {
                    "source_column": "Total Units Sold",
                    "metric_name": "product_sold_count",
                    "metric_scope": "PRODUCT",
                    "truth_status": "VERIFIED_PRODUCT_LEVEL",
                    "warning": None,
                },
                {
                    "source_column": "Shop Total Units Sold",
                    "metric_name": "shop_total_sold_count",
                    "metric_scope": "SHOP",
                    "truth_status": "SHOP_LEVEL_AGGREGATE",
                    "warning": "SHOP_LEVEL_METRIC_NOT_PRODUCT_SALES",
                },
                {
                    "source_column": "Orders",
                    "metric_name": "order_count",
                    "metric_scope": "UNKNOWN",
                    "truth_status": "NOT_VERIFIED",
                    "warning": "SALES_METRIC_SCOPE_NOT_VERIFIED",
                },
            ],
        },
        {
            "sheet": "Most Promoted Products",
            "name_fields": ["Product Name"],
            "shop_fields": ["Shop Name"],
            "source_fields": ["FastMoss Product Detail", "FastMoss Shop Detail"],
            "tiktok_fields": ["TikTok Product Detail"],
            "metric_fields": [
                {
                    "source_column": "Total Units Sold",
                    "metric_name": "product_sold_count",
                    "metric_scope": "PRODUCT",
                    "truth_status": "VERIFIED_PRODUCT_LEVEL",
                    "warning": None,
                },
                {
                    "source_column": "Shop Units Sold",
                    "metric_name": "shop_total_sold_count",
                    "metric_scope": "SHOP",
                    "truth_status": "SHOP_LEVEL_AGGREGATE",
                    "warning": "SHOP_LEVEL_METRIC_NOT_PRODUCT_SALES",
                },
            ],
        },
        {
            "sheet": "Video Product List",
            "name_fields": ["Product Title"],
            "shop_fields": [],
            "source_fields": ["FastMoss Product Detail Page Link"],
            "tiktok_fields": ["TikTok Product Link"],
            "metric_fields": [
                {
                    "source_column": "Video Total Units Sold",
                    "metric_name": "product_sold_count",
                    "metric_scope": "PRODUCT",
                    "truth_status": "VERIFIED_PRODUCT_LEVEL",
                    "warning": None,
                },
                {
                    "source_column": "Video Units Sold",
                    "metric_name": "product_sold_count",
                    "metric_scope": "PRODUCT",
                    "truth_status": "VERIFIED_PRODUCT_LEVEL",
                    "warning": None,
                },
            ],
        },
        {
            "sheet": "Product Search Data",
            "name_fields": ["Product Name"],
            "shop_fields": ["Store Name"],
            "source_fields": ["FastMoss", "FastMoss Shop"],
            "tiktok_fields": ["TikTok"],
            "metric_fields": [
                {
                    "source_column": "Total Sales Volume",
                    "metric_name": "total_sales_volume",
                    "metric_scope": "UNKNOWN",
                    "truth_status": "NOT_VERIFIED",
                    "warning": "SALES_METRIC_SCOPE_NOT_VERIFIED",
                },
                {
                    "source_column": "7-Day Sales Volume",
                    "metric_name": "total_sales_volume",
                    "metric_scope": "UNKNOWN",
                    "truth_status": "NOT_VERIFIED",
                    "warning": "SALES_METRIC_SCOPE_NOT_VERIFIED",
                },
            ],
        },
        {
            "sheet": "New Products Ranking",
            "name_fields": ["Product Name"],
            "shop_fields": ["Shop"],
            "source_fields": [],
            "tiktok_fields": [],
            "metric_fields": [
                {
                    "source_column": "Units Sold",
                    "metric_name": "product_sold_count",
                    "metric_scope": "PRODUCT",
                    "truth_status": "VERIFIED_PRODUCT_LEVEL",
                    "warning": None,
                },
                {
                    "source_column": "Shop Units Sold",
                    "metric_name": "shop_total_sold_count",
                    "metric_scope": "SHOP",
                    "truth_status": "SHOP_LEVEL_AGGREGATE",
                    "warning": "SHOP_LEVEL_METRIC_NOT_PRODUCT_SALES",
                },
            ],
        },
        {
            "sheet": "Copywriting_Product_Map",
            "name_fields": ["Product Name"],
            "shop_fields": ["Shop Name"],
            "source_fields": [],
            "tiktok_fields": [],
            "metric_fields": [
                {
                    "source_column": "Total Units Sold",
                    "metric_name": "product_sold_count",
                    "metric_scope": "PRODUCT",
                    "truth_status": "VERIFIED_PRODUCT_LEVEL",
                    "warning": None,
                },
                {
                    "source_column": "Orders",
                    "metric_name": "order_count",
                    "metric_scope": "UNKNOWN",
                    "truth_status": "NOT_VERIFIED",
                    "warning": "SALES_METRIC_SCOPE_NOT_VERIFIED",
                },
            ],
        },
    ]

    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []
    for config in configs:
        if config["sheet"] not in workbook.sheetnames:
            continue
        ws = workbook[config["sheet"]]
        headers = _find_sheet_header(ws)
        if not headers:
            continue
        for row in ws.iter_rows(values_only=True):
            values = list(row)
            if not any(value is not None and str(value).strip() for value in values):
                continue
            if headers and values[: len(headers)] == headers[: len(values[: len(headers)])]:
                continue
            data = _row_mapping(headers, values)
            names = _unique([data.get(field) for field in config["name_fields"]])
            if not names:
                continue
            shop_names = _unique([data.get(field) for field in config["shop_fields"]])
            metric_values = [
                {**metric, "value": value}
                for metric in config["metric_fields"]
                if (value := _to_int(data.get(metric["source_column"]))) is not None
            ]
            source_urls = _unique([data.get(field) for field in config["source_fields"]])
            tiktok_urls = _unique([data.get(field) for field in config["tiktok_fields"]])
            rows.append(
                {
                    "source": "LEGACY_COMBINED_WORKBOOK",
                    "sheet": config["sheet"],
                    "file_type_id": config["sheet"],
                    "matched_by": "legacy_combined_workbook",
                    "batch_id": None,
                    "names": names,
                    "shop_names": shop_names,
                    "metric_values": metric_values,
                    "source_urls": source_urls,
                    "tiktok_urls": tiktok_urls,
                }
            )
    return rows


@lru_cache(maxsize=1)
def _sales_metrics_index() -> dict[str, Any]:
    records = _iter_sales_workbook_records()
    by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    by_source_url: dict[str, list[dict[str, Any]]] = defaultdict(list)
    by_tiktok_url: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        for name in record["names"]:
            by_name[_normalize_key(name)].append(record)
        for url in record["source_urls"]:
            by_source_url[_normalize_text(url)].append(record)
        for url in record["tiktok_urls"]:
            by_tiktok_url[_normalize_text(url)].append(record)
    return {
        "source": "LEGACY_COMBINED_WORKBOOK",
        "batch_id": None,
        "records": records,
        "by_name": by_name,
        "by_source_url": by_source_url,
        "by_tiktok_url": by_tiktok_url,
    }


def _latest_sales_metrics_index() -> dict[str, Any] | None:
    latest = get_latest_fastmoss_reference_index()
    if not latest or not latest.get("records"):
        return None
    return {
        "source": "LATEST_FASTMOSS_IMPORT_BATCH",
        "batch_id": latest.get("batch_id"),
        "records": latest["records"],
        "by_name": latest["by_name"],
        "by_source_url": latest["by_source_url"],
        "by_tiktok_url": latest["by_tiktok_url"],
    }


def _match_sales_records(index: dict[str, Any], product: dict[str, Any]) -> tuple[list[dict[str, Any]], list[str], str | None]:
    matched: list[dict[str, Any]] = []
    provenance: list[str] = []
    matched_by: str | None = None

    source_url = _normalize_text(product.get("source_url"))
    tiktok_url = _normalize_text(product.get("tiktok_product_url"))
    if source_url and source_url in index["by_source_url"]:
        matched.extend(index["by_source_url"][source_url])
        provenance.append("sales_metrics:matched_source_url")
        matched_by = matched_by or "SOURCE_URL"
    if tiktok_url and tiktok_url in index["by_tiktok_url"]:
        matched.extend(index["by_tiktok_url"][tiktok_url])
        provenance.append("sales_metrics:matched_tiktok_product_url")
        matched_by = matched_by or "TIKTOK_PRODUCT_URL"

    if not matched:
        for name in _product_names(product):
            normalized = _normalize_key(name)
            if normalized in index["by_name"]:
                matched.extend(index["by_name"][normalized])
                provenance.append("sales_metrics:matched_exact_name")
                matched_by = "EXACT_NAME"
                break

    if not matched:
        raw_names = [name for name in _product_names(product) if len(_normalize_key(name)) >= 10]
        for candidate in raw_names:
            normalized = _normalize_key(candidate)
            fuzzy_by_identity: dict[int, dict[str, Any]] = {}
            # The by-name index is normalized once when the sales index is built.
            # Reusing those keys prevents every product taxonomy read from
            # regex-normalizing the complete sales-record corpus again.
            for indexed_name, indexed_records in index["by_name"].items():
                if (
                    normalized not in indexed_name
                    and indexed_name not in normalized
                ):
                    continue
                for record in indexed_records:
                    fuzzy_by_identity.setdefault(id(record), record)
            fuzzy = list(fuzzy_by_identity.values())
            if len(fuzzy) == 1:
                matched.extend(fuzzy)
                provenance.append("sales_metrics:matched_unique_fuzzy_name")
                matched_by = "UNIQUE_FUZZY_NAME"
                break

    unique_records: list[dict[str, Any]] = []
    seen_keys: set[tuple[str, tuple[str, ...]]] = set()
    for record in matched:
        record_key = (
            str(record.get("file_type_id") or record.get("sheet") or "UNKNOWN"),
            tuple(sorted(_normalize_key(name) for name in record["names"])),
        )
        if record_key in seen_keys:
            continue
        seen_keys.add(record_key)
        unique_records.append(record)
    return unique_records, provenance, matched_by


def _resolve_sales_metrics(product: dict[str, Any]) -> tuple[ProductIntelligenceSalesMetrics, list[str]]:
    def _resolved_from_index(index: dict[str, Any]) -> tuple[ProductIntelligenceSalesMetrics, list[str]]:
        unique_records, provenance, matched_by = _match_sales_records(index, product)
        if not unique_records:
            return (
                ProductIntelligenceSalesMetrics(
                    sold_count=None,
                    product_sold_count=None,
                    shop_total_sold_count=None,
                    shop_count=None,
                    shop_names=[],
                    source_status="NOT_FOUND",
                    sold_count_metric_scope="UNKNOWN",
                    sold_count_truth_status="NOT_VERIFIED",
                    sales_metric_warnings=[],
                    sales_metric_provenance=[f"sales_metrics:source={index['source']}"],
                    sales_metrics_source=index["source"],
                    sales_metrics_batch_id=index.get("batch_id"),
                ),
                provenance,
            )

        metric_candidates = [
            {
                **metric,
                "file_type_id": record.get("file_type_id") or record.get("sheet"),
                "matched_by": matched_by,
            }
            for record in unique_records
            for metric in record.get("metric_values", [])
            if metric.get("value") is not None
        ]
        product_candidates = [
            candidate
            for candidate in metric_candidates
            if candidate.get("metric_scope") == "PRODUCT"
            and candidate.get("truth_status") == "VERIFIED_PRODUCT_LEVEL"
        ]
        shop_candidates = [
            candidate
            for candidate in metric_candidates
            if candidate.get("metric_scope") == "SHOP"
            and candidate.get("truth_status") == "SHOP_LEVEL_AGGREGATE"
        ]
        unknown_candidates = [
            candidate
            for candidate in metric_candidates
            if candidate.get("truth_status") == "NOT_VERIFIED"
            or candidate.get("metric_scope") == "UNKNOWN"
        ]

        product_metric = max(product_candidates, key=lambda item: int(item["value"])) if product_candidates else None
        shop_metric = max(shop_candidates, key=lambda item: int(item["value"])) if shop_candidates else None
        unknown_metric = max(unknown_candidates, key=lambda item: int(item["value"])) if unknown_candidates else None
        selected_metric = product_metric or shop_metric or unknown_metric
        shop_names = _unique(
            shop_name
            for record in unique_records
            for shop_name in record["shop_names"]
        )
        sheets = sorted(
            {
                str(record.get("file_type_id") or record.get("sheet") or "UNKNOWN")
                for record in unique_records
            }
        )
        if sheets:
            provenance.append("sales_metrics:sources=" + ",".join(sheets))

        warnings: list[str] = []
        scope = "UNKNOWN"
        truth_status = "NOT_VERIFIED"
        product_sold_count = None
        shop_total_sold_count = None
        if product_metric:
            product_sold_count = int(product_metric["value"])
            scope = "PRODUCT"
            truth_status = "VERIFIED_PRODUCT_LEVEL"
        elif shop_metric:
            shop_total_sold_count = int(shop_metric["value"])
            scope = "SHOP"
            truth_status = "SHOP_LEVEL_AGGREGATE"
            warnings.append("SHOP_LEVEL_METRIC_NOT_PRODUCT_SALES")
        elif unknown_metric:
            scope = "UNKNOWN"
            truth_status = "NOT_VERIFIED"
            warnings.append("SALES_METRIC_SCOPE_NOT_VERIFIED")

        metric_provenance = [
            f"sales_metrics:source={index['source']}",
            *([f"sales_metrics:batch_id={index['batch_id']}"] if index.get("batch_id") else []),
            *([f"sales_metrics:matched_by={matched_by}"] if matched_by else []),
            *([f"sales_metrics:file_type={selected_metric['file_type_id']}"] if selected_metric and selected_metric.get("file_type_id") else []),
            *([f"sales_metrics:column={selected_metric['source_column']}"] if selected_metric and selected_metric.get("source_column") else []),
        ]

        return (
            ProductIntelligenceSalesMetrics(
                sold_count=product_sold_count,
                product_sold_count=product_sold_count,
                shop_total_sold_count=shop_total_sold_count,
                shop_count=len(shop_names) if shop_names else None,
                shop_names=shop_names,
                source_status="FOUND",
                sold_count_metric_scope=scope,
                sold_count_truth_status=truth_status,
                sales_metric_warnings=_unique(warnings),
                sales_metric_provenance=_unique(metric_provenance),
                sales_metrics_source=index["source"],
                sales_metrics_batch_id=index.get("batch_id"),
                matched_file_type=str(selected_metric.get("file_type_id")) if selected_metric else None,
                matched_by=matched_by,
                raw_metric_column=str(selected_metric.get("source_column")) if selected_metric else None,
            ),
            provenance,
        )

    latest_index = _latest_sales_metrics_index()
    if latest_index:
        latest_metrics, latest_provenance = _resolved_from_index(latest_index)
        if latest_metrics.source_status == "FOUND":
            return latest_metrics, latest_provenance

    legacy_metrics, legacy_provenance = _resolved_from_index(_sales_metrics_index())
    if latest_index and legacy_metrics.source_status == "NOT_FOUND":
        return (
            ProductIntelligenceSalesMetrics(
                sold_count=None,
                product_sold_count=None,
                shop_total_sold_count=None,
                shop_count=None,
                shop_names=[],
                source_status="NOT_FOUND",
                sold_count_metric_scope="UNKNOWN",
                sold_count_truth_status="NOT_VERIFIED",
                sales_metric_warnings=[],
                sales_metric_provenance=[
                    "sales_metrics:source=NOT_FOUND",
                    f"sales_metrics:latest_batch_id={latest_index.get('batch_id')}",
                ],
                sales_metrics_source="NOT_FOUND",
                sales_metrics_batch_id=latest_index.get("batch_id"),
            ),
            legacy_provenance,
        )
    return legacy_metrics, legacy_provenance


def _resolve_image_analysis(product: dict[str, Any]) -> dict[str, Any]:
    return analyze_product_image_payload(
        product,
        allow_provider_execution=product.get("allow_live_image_analysis") is True,
    )


IMAGE_PACKAGE_FORM_MAP = {
    "bottle": "bottle_or_refill_pack",
    "refill_pouch": "bottle_or_refill_pack",
    "tube": "small_bottle_tube_or_compact",
    "box": "boxed_item",
    "packet": "flat_packet",
    "garment": "garment",
    "jar": "jar_container",
    "roll_on_bottle": "small_roll_on_bottle",
}

IMAGE_PHYSICAL_STATE_MAP = {
    "garment": "textile",
    "packet": "paper",
    "box": "solid",
}


def _image_analysis_is_high_confidence(image_analysis: dict[str, Any]) -> bool:
    return (
        str(image_analysis.get("status")) == "ANALYZED"
        and str(image_analysis.get("visual_confidence")) == "HIGH"
    )


def _apply_image_evidence_to_profile(
    *,
    profile: dict[str, str],
    family: str,
    image_analysis: dict[str, Any],
    warnings: list[str],
    provenance: list[str],
) -> tuple[dict[str, str], bool]:
    updated = dict(profile)
    changed = False
    detected_package = str(image_analysis.get("detected_package") or "").strip()
    if not _image_analysis_is_high_confidence(image_analysis) or not detected_package:
        return updated, changed

    image_package_form = IMAGE_PACKAGE_FORM_MAP.get(detected_package)
    image_physical_state = IMAGE_PHYSICAL_STATE_MAP.get(detected_package)

    if image_package_form and updated.get("package_form") != image_package_form:
        updated["package_form"] = image_package_form
        provenance.append(f"image_analysis:package_form={image_package_form}")
        changed = True
    if image_physical_state and updated.get("physical_state") in {"unknown", "liquid_or_semi_liquid"}:
        updated["physical_state"] = image_physical_state
        provenance.append(f"image_analysis:physical_state={image_physical_state}")
        changed = True

    if family in {"LAUNDRY_DETERGENT_LIQUID_REFILL", "FABRIC_SOFTENER_LIQUID"} and detected_package == "garment":
        warnings.append("IMAGE_TITLE_CONFLICT_REVIEW_REQUIRED")
        provenance.append("image_analysis:title_conflict=garment_vs_liquid_detergent")
    if family.startswith("fashion") and detected_package in {"bottle", "refill_pouch", "tube", "jar", "roll_on_bottle"}:
        warnings.append("IMAGE_TITLE_CONFLICT_REVIEW_REQUIRED")
        provenance.append("image_analysis:title_conflict=container_vs_apparel")

    return updated, changed


def _resolve_claim_gate(
    product: dict[str, Any],
    family: str,
    copy_route: str,
) -> tuple[str, list[str], list[str]]:
    matched_blocked, matched_review, warnings = evaluate_product_claims(
        product,
        family=family,
    )
    if matched_blocked:
        warnings.append("claim_gate:blocked_tokens_present")
        return "CLAIM_BLOCKED", matched_review + matched_blocked, warnings
    if family == "MALE_HEALTH_SENSITIVE":
        if "male_health_sensitive" not in matched_review:
            matched_review.append("male_health_sensitive")
        warnings.append("claim_gate:male_health_sensitive")
        return "CLAIM_REVIEW_REQUIRED", matched_review, warnings
    if family == "FEMALE_HEALTH_SENSITIVE":
        if "female_health_sensitive" not in matched_review:
            matched_review.append("female_health_sensitive")
        warnings.append("claim_gate:female_health_sensitive")
        return "CLAIM_REVIEW_REQUIRED", matched_review, warnings
    if matched_review:
        if any(token in matched_review for token in ["antibakteria", "antibaktiria", "antibacterial", "anti bacterial", "antibacteria"]):
            if "antibacterial_claim" not in matched_review:
                matched_review.append("antibacterial_claim")
        warnings.append("claim_gate:review_tokens_present")
        return "CLAIM_REVIEW_REQUIRED", matched_review, warnings
    if copy_route in {"STEALTH", "REVIEW_REQUIRED"}:
        warnings.append("claim_gate:route_requires_review")
        return "CLAIM_REVIEW_REQUIRED", matched_review, warnings
    return "CLAIM_SAFE", [], warnings


def _resolve_family_from_title(product: dict[str, Any]) -> tuple[str | None, str | None]:
    haystack = _joined_title_text(product)
    if _contains_any(
        haystack,
        [
            "sabun dobi",
            "liquid laundry detergent",
            "laundry detergent",
            "detergen",
            "detergent refill",
            "pencuci baju",
            "isi ulang",
        ],
    ):
        return (
            "LAUNDRY_DETERGENT_LIQUID_REFILL",
            "title_evidence:laundry_detergent_keywords",
        )
    if _contains_any(
        haystack,
        ["softener", "fabric softener", "pelembut", "pewangi pakaian"],
    ):
        return ("FABRIC_SOFTENER_LIQUID", "title_evidence:fabric_softener_keywords")
    if _contains_any(
        haystack,
        ["male health", "suami isteri", "batin", "kuat lelaki"],
    ) or (
        "tahan lama" in haystack and not _contains_any(haystack, ["makeup", "kosmetik", "lipstick", "lipmatte", "setting spray", "baju", "pants", "seluar"])
    ):
        # Strict Isolation: MALE_HEALTH_SENSITIVE requires specific sensitive health tokens.
        # 'tahan lama' alone in makeup/fashion context is not sensitive.
        return ("MALE_HEALTH_SENSITIVE", "title_evidence:male_health_sensitive_keywords")
    if _contains_any(
        haystack,
        [
            "female health",
            "wanita",
            "perempuan",
            "perapat",
            "keputihan",
            "miss v",
            "intim wanita",
            "kewanitaan",
            "jamu wanita",
        ],
    ) or (
        # "ketat"/"rapat" (tight/close) are generic physical descriptors — a
        # tight fishing line, a tight seal, a snug fit.  They only resolve to
        # intimate female health when corroborated by an actual feminine /
        # intimate cue, never on their own.
        _contains_any(haystack, ["ketat", "rapat"])
        and _contains_any(
            haystack,
            [
                "intim",
                "faraj",
                "vagina",
                "miss v",
                "kewanitaan",
                "keputihan",
                "kegel",
                "postpartum",
                "selepas bersalin",
                "bersalin",
                "merapat",
                "kesegaran wanita",
                "feminine",
            ],
        )
    ):
        return ("FEMALE_HEALTH_SENSITIVE", "title_evidence:female_health_sensitive_keywords")
    if _contains_any(
        haystack,
        ["supplement", "capsule", "vitamin", "wellness supplement", "beauty supplement"],
    ):
        return ("HEALTH_SUPPLEMENT", "title_evidence:supplement_keywords")
    if _contains_any(
        haystack,
        ["organizer", "storage", "rak", "bekas simpan", "container set"],
    ):
        return ("HOUSEHOLD_STORAGE_ORGANIZER", "title_evidence:storage_keywords")
    if _contains_any(
        haystack,
        ["cleaner", "all purpose cleaner", "floor cleaner", "toilet cleaner", "sabun pencuci"],
    ):
        return ("HOUSEHOLD_CLEANER_GENERAL", "title_evidence:cleaner_keywords")
    if _contains_any(
        haystack,
        ["sleepwear", "loungewear", "nightdress", "baju tidur", "kelawar", "nightie"],
    ):
        return ("APPAREL_SLEEPWEAR", "title_evidence:sleepwear_keywords")
    if _contains_any(
        haystack,
        ["instant sarung", "sarung syria", "khimar", "telekung", "tudung labuh", "moscrepe"],
    ):
        return ("fashion_modestwear", "title_evidence:modestwear_keywords")
    if _contains_any(
        haystack,
        ["jersey", "jersi", "athleisure", "baju sukan", "quick dry"],
    ):
        return ("fashion_sportswear", "title_evidence:sportswear_keywords")
    if _contains_any(haystack, ["baby wipes", "wet wipes", "wet tissue", "tisu basah"]):
        return ("BABY_WIPES", "title_evidence:baby_wipes_keywords")
    if _contains_any(haystack, ["diaper", "lampin", "pull ups", "pull-ups", "baby diaper"]):
        return ("BABY_DIAPER", "title_evidence:baby_diaper_keywords")
    if _contains_any(
        haystack,
        ["body spray", "perfume", "fragrance", "body mist", "mist"],
    ):
        # Guardrail: avoid matching 'fragrance-free' as fragrance
        if "fragrance-free" in haystack or "fragrance free" in haystack:
             pass # let it fall through
        else:
             return ("beauty_fragrance", "title_evidence:fragrance_keywords")
    if _contains_any(
        haystack,
        [
            "lip balm",
            "lip gloss",
            "lipstick",
            "serum",
            "cleanser",
            "moisturizer",
            "foundation",
            "concealer",
            "body wash",
            "soap",
            "skincare",
            "beauty",
        ],
    ):
        return ("BEAUTY_PERSONAL_CARE", "title_evidence:beauty_keywords")
    if _contains_any(
        haystack,
        ["envelope", "duit raya", "money packet", "angpow", "red packet", "sampul"],
    ):
        return ("stationery_paper", "title_evidence:paper_packet_keywords")
    if _contains_any(
        haystack,
        ["towel", "tuala", "blanket", "comforter", "selimut", "bedsheet", "cadar", "curtain", "pillow", "mat-rug", "rug", "bedding"],
    ):
        # Guardrail: HOME_TEXTILE must not hijack beauty products with 'matte' or 'powder'
        if _contains_any(haystack, ["matte", "powder", "lipmatte"]) and not _contains_any(haystack, ["towel", "blanket", "bedsheet"]):
             pass
        else:
             return ("HOME_TEXTILE", "title_evidence:home_textile_keywords")
    # Isolated 'mat' check to avoid partial matches like 'lipmatte'
    if re.search(r"\bmat\b", haystack) and not _contains_any(haystack, ["matte", "lipmatte"]):
         return ("HOME_TEXTILE", "title_evidence:home_textile_keywords")
    if _contains_any(haystack, ["cat food", "cat treat", "pet", "kucing"]):
        return ("PET_CARE_GENERAL", "title_evidence:petcare_keywords")
    if _contains_any(haystack, ["sauce", "sambal", "popcorn", "chocolate", "biscuits", "cookies", "food"]):
        return ("food_packaged", "title_evidence:food_keywords")
    if _contains_any(haystack, ["smartwatch", "wearable", "charger", "adapter", "cable"]):
        return ("electronics_wearable", "title_evidence:electronics_keywords")
    if _contains_any(
        haystack,
        [
            "car phone holder",
            "phone holder",
            "phone mount",
            "dashboard mount",
            "windshield mount",
            "suction cup",
        ],
    ):
        return ("ACCESSORY_SMALL_ITEM", "title_evidence:phone_mount_accessory_keywords")
    if _contains_any(haystack, ["brooch", "earring", "pin", "charm", "pendant", "clip", "accessory"]):
        return ("ACCESSORY_SMALL_ITEM", "title_evidence:accessory_keywords")
    if _contains_any(haystack, ["tool", "hardware", "automotive", "motorcycle", "car care"]):
        return ("AUTO_TOOL_GENERAL", "title_evidence:auto_tool_keywords")
    if _contains_any(haystack, ["service", "consultation", "homestay", "property", "rumah untuk dijual"]):
        return ("REAL_ESTATE_OR_SERVICE", "title_evidence:service_keywords")
    return None, None


def _resolve_family_from_taxonomy(product: dict[str, Any]) -> tuple[str, str]:
    category = _normalize_key(product.get("category"))
    subcategory = _normalize_key(product.get("subcategory"))
    type_name = _normalize_key(product.get("type"))
    taxonomy = " ".join(part for part in [category, subcategory, type_name] if part)

    if any(token in taxonomy for token in ["laundry detergent", "household cleaners", "home care supplies"]):
        return "LAUNDRY_DETERGENT_LIQUID_REFILL", "taxonomy_fallback:laundry_or_cleaner"
    if any(token in taxonomy for token in ["beauty and personal care", "cosmetics", "fragrance", "bath and body"]):
        if "fragrance" in taxonomy:
            return "beauty_fragrance", "taxonomy_fallback:fragrance"
        return "BEAUTY_PERSONAL_CARE", "taxonomy_fallback:beauty_personal_care"
    if any(token in taxonomy for token in ["womenswear and underwear", "fashion", "muslim fashion", "menswear and underwear"]):
        return "fashion_apparel", "taxonomy_fallback:fashion_apparel"
    if any(token in taxonomy for token in ["textiles and soft furnishings", "bedding", "carpet", "curtains"]):
        return "HOME_TEXTILE", "taxonomy_fallback:home_textile"
    if any(token in taxonomy for token in ["kitchen storage", "food container", "home organization"]):
        return "HOUSEHOLD_STORAGE_ORGANIZER", "taxonomy_fallback:storage"
    if any(token in taxonomy for token in ["stationery", "envelope"]):
        return "stationery_paper", "taxonomy_fallback:stationery"
    if any(token in taxonomy for token in ["baby and maternity", "baby care", "diapers"]):
        if "wipes" in taxonomy:
            return "BABY_WIPES", "taxonomy_fallback:baby_wipes"
        return "BABY_DIAPER", "taxonomy_fallback:baby_diaper"
    if "male health" in taxonomy:
        return "MALE_HEALTH_SENSITIVE", "taxonomy_fallback:male_health_sensitive"
    if "female health" in taxonomy or "feminine care" in taxonomy:
        return "FEMALE_HEALTH_SENSITIVE", "taxonomy_fallback:female_health_sensitive"
    if any(token in taxonomy for token in ["food and beverage", "food and beverages", "kitchenware"]):
        return "food_packaged", "taxonomy_fallback:food"
    if any(token in taxonomy for token in ["health", "supplements"]):
        return "HEALTH_SUPPLEMENT", "taxonomy_fallback:health_supplement"
    if any(token in taxonomy for token in ["pet supplies"]):
        return "PET_CARE_GENERAL", "taxonomy_fallback:petcare"
    if any(token in taxonomy for token in ["tools and hardware", "automotive and motorcycle", "home improvement"]):
        return "AUTO_TOOL_GENERAL", "taxonomy_fallback:auto_tools"
    if any(token in taxonomy for token in ["phones and electronics", "electronics", "computers and office equipment", "household appliances"]):
        return "electronics_wearable", "taxonomy_fallback:electronics"
    # Generic physical-good categories that previously fell through to
    # UNKNOWN_REVIEW_REQUIRED and blocked commit on CLEAR_PRODUCT_FAMILY_INFERENCE.
    # All resolve to safe DIRECT-copy generic families (owner-approved auto-classify).
    if any(token in taxonomy for token in ["sports and outdoor", "sporting goods"]):
        return "AUTO_TOOL_GENERAL", "taxonomy_fallback:sports_outdoor"
    if any(token in taxonomy for token in ["toys and hobbies", "toys"]):
        return "toy_play", "taxonomy_fallback:toys"
    if any(token in taxonomy for token in ["books", "magazines"]):
        return "stationery_paper", "taxonomy_fallback:books_media"
    if any(token in taxonomy for token in ["shoes", "footwear"]):
        return "fashion_apparel", "taxonomy_fallback:shoes"
    if "luggage and bags" in taxonomy:
        return "ACCESSORY_SMALL_ITEM", "taxonomy_fallback:luggage_bags"
    return "UNKNOWN_REVIEW_REQUIRED", "taxonomy_fallback:unknown"


def _resolve_family(product: dict[str, Any]) -> tuple[str, str, bool, str | None]:
    category = _normalize_key(product.get("category"))
    type_name = _normalize_key(product.get("type"))
    if category == "kitchenware" and type_name == "specialty kitchen utensils":
        return (
            "AUTO_TOOL_GENERAL",
            "taxonomy_evidence:specialty_kitchen_utensils",
            False,
            None,
        )

    title_family, title_reason = _resolve_family_from_title(product)
    family_context = derive_bosmax_product_family(product)

    if title_family:
        family = title_family
        reason = title_reason or "title_evidence"
    elif family_context["bosmax_product_family"] != "GENERIC_UNCLASSIFIED":
        family = str(family_context["bosmax_product_family"])
        reason = "family_resolver:" + str(family_context["bosmax_product_family_reason"])
    else:
        family, reason = _resolve_family_from_taxonomy(product)

    taxonomy_conflict = bool(family_context["bosmax_source_taxonomy_conflict"])
    conflict_reason = (
        str(family_context["bosmax_source_taxonomy_conflict_reason"]).strip() or None
    )
    if family == "LAUNDRY_DETERGENT_LIQUID_REFILL" and "baby" in category:
        taxonomy_conflict = True
        conflict_reason = (
            conflict_reason
            or "Title evidence indicates laundry detergent, but source taxonomy is under baby-care lanes."
        )
    if family == "HOUSEHOLD_STORAGE_ORGANIZER" and _contains_any(
        _joined_title_text(product), ["sabun dobi", "detergent", "laundry"]
    ):
        taxonomy_conflict = True
        conflict_reason = conflict_reason or "Storage taxonomy conflicts with laundry detergent title evidence."

    return family, reason, taxonomy_conflict, conflict_reason


def _profile_for_family(family: str) -> dict[str, str]:
    return dict(FAMILY_PROFILES.get(family) or FAMILY_PROFILES["UNKNOWN_REVIEW_REQUIRED"])


def _resolve_confidence(
    reason: str,
    taxonomy_conflict: bool,
    family: str,
    source_taxonomy: dict[str, str | None],
) -> str:
    if family == "UNKNOWN_REVIEW_REQUIRED":
        return "LOW"
    if reason.startswith("title_evidence:"):
        return "MEDIUM" if taxonomy_conflict else "HIGH"
    if reason.startswith("family_resolver:"):
        return "MEDIUM" if taxonomy_conflict else "HIGH"
    if not any(source_taxonomy.values()):
        return "LOW"
    return "LOW" if taxonomy_conflict else "MEDIUM"


def _destination_readiness(
    *,
    copy_route: str,
    claim_gate: str,
    confidence: str,
    image_analysis_status: str,
) -> dict[str, str]:
    review_required = copy_route != "DIRECT" or claim_gate != "CLAIM_SAFE" or confidence == "LOW"
    text_to_video = "READY" if not review_required else "NEEDS_REVIEW"
    semantic_image_ready = image_analysis_status == "ANALYZED"
    frames = (
        "READY"
        if semantic_image_ready and not review_required
        else "NEEDS_REVIEW"
    )
    ingredients = (
        "READY"
        if semantic_image_ready and confidence in {"HIGH", "MEDIUM"}
        else "NEEDS_REVIEW"
    )
    image = (
        "READY"
        if semantic_image_ready and not review_required
        else "NEEDS_REVIEW"
    )
    return {
        "TEXT_TO_VIDEO": text_to_video,
        "FRAMES": frames,
        "INGREDIENTS": ingredients,
        "IMAGE": image,
    }


def resolve_product_intelligence_profile(product: dict[str, Any]) -> dict[str, Any]:
    payload = dict(product)
    family, family_reason, taxonomy_conflict, taxonomy_conflict_reason = _resolve_family(payload)
    profile = _profile_for_family(family)
    source = _coerce_source(payload.get("source"))
    copy_route = profile["copy_route"]
    claim_gate, claim_tokens, claim_warnings = _resolve_claim_gate(payload, family, copy_route)
    image_analysis = _resolve_image_analysis(payload)
    sales_metrics, sales_provenance = _resolve_sales_metrics(payload)
    source_taxonomy = {
        "category": _first_non_empty(payload.get("category")),
        "subcategory": _first_non_empty(payload.get("subcategory")),
        "type": _first_non_empty(payload.get("type")),
    }
    confidence = _resolve_confidence(
        family_reason,
        taxonomy_conflict,
        family,
        source_taxonomy,
    )
    warnings: list[str] = []
    provenance = [
        "resolver:product_intelligence_service",
        "evidence_priority:title_then_taxonomy_then_workbook_then_fallback",
        f"family:{family}",
        family_reason,
    ]
    profile, image_profile_changed = _apply_image_evidence_to_profile(
        profile=profile,
        family=family,
        image_analysis=image_analysis,
        warnings=warnings,
        provenance=provenance,
    )

    # 2. Consume ProductTruthService for read-only reconciliation audit
    from agent.services.product_truth_service import ProductTruthService
    truth_profile = ProductTruthService.build_computed_profile(payload)
    recon = truth_profile.reconciliation
    
    # 3. Apply Reconciliation Overrides
    # If ProductTruth identifies a category boundary lock violation or a severe contradiction, 
    # we MUST NOT allow HIGH confidence.
    if recon.confidence_label == "NEEDS_REVIEW" or "FLAG_CATEGORY_BOUNDARY_LOCK_VIOLATION" in recon.contradiction_flags:
        confidence = "LOW"
        warnings.append("MAPPING_RECONCILIATION_CONTRADICTION")
        provenance.append(f"reconciliation:flagged_{recon.confidence_label}")
    else:
        # Cap intelligence confidence by reconciliation confidence
        if recon.confidence_label == "LOW" and confidence != "LOW":
             confidence = "LOW"
             warnings.append("MAPPING_RECONCILIATION_WEAK_SIGNAL")
        elif recon.confidence_label == "MEDIUM" and confidence == "HIGH":
             confidence = "MEDIUM"
             warnings.append("MAPPING_RECONCILIATION_UNCERTAIN")

    # Special Case: MALE_HEALTH_SENSITIVE guardrail
    # If it was mapped to MALE_HEALTH_SENSITIVE but Truth doesn't corroborate it (e.g. it's actually fashion), 
    # we force it back to fashion or unknown.
    if family == "MALE_HEALTH_SENSITIVE" and "FLAG_CATEGORY_BOUNDARY_LOCK_VIOLATION" in recon.contradiction_flags:
        # If the title says 'seluar' or 'pants', it's likely fashion
        if _contains_any(_joined_title_text(payload), ["seluar", "pants", "jersey", "baju"]):
             family = "fashion_apparel"
             profile = _profile_for_family(family)
             provenance.append("reconciliation:male_health_to_fashion_recovery")

    # 4. Final readiness and status
    readiness = _destination_readiness(
        copy_route=copy_route,
        claim_gate=claim_gate,
        confidence=confidence,
        image_analysis_status=str(image_analysis.get("status") or ""),
    )

    if taxonomy_conflict:
        warnings.append("TAXONOMY_CONFLICT")
        provenance.append("taxonomy_conflict:source_taxonomy_overridden")
    
    warnings.extend(
        warning
        for warning in claim_warnings
        if warning not in warnings
    )
    if sales_metrics.source_status == "NOT_FOUND":
        warnings.append("SALES_METRICS_NOT_FOUND")
    else:
        warnings.extend(
            warning
            for warning in sales_metrics.sales_metric_warnings
            if warning not in warnings
        )
    provenance.extend(sales_provenance)
    provenance.extend(
        note
        for note in sales_metrics.sales_metric_provenance
        if note not in provenance
    )
    
    image_warnings = [str(warning) for warning in image_analysis.get("warnings", []) if str(warning).strip()]
    warnings.extend(image_warnings)
    
    if image_analysis.get("status") == "VISION_PROVIDER_NOT_CONFIGURED":
        provenance.append("image_analysis:provider_not_configured")
    elif image_analysis.get("status") == "ANALYSIS_SKIPPED":
        provenance.append("image_analysis:provider_execution_disabled")
    elif image_analysis.get("status") == "IMAGE_MISSING":
        warnings.append("IMAGE_NOT_AVAILABLE")
        warnings.append("IMAGE_REFERENCE_REQUIRED")
    elif image_analysis.get("status") == "IMAGE_INACCESSIBLE":
        warnings.append("IMAGE_INACCESSIBLE")
    elif image_analysis.get("status") == "UNSUPPORTED_IMAGE_FORMAT":
        warnings.append("UNSUPPORTED_IMAGE_FORMAT")
    elif image_analysis.get("status") == "ANALYSIS_FAILED":
        warnings.append("SEMANTIC_IMAGE_ANALYSIS_FAILED")
        
    if image_profile_changed:
        provenance.append("image_analysis:high_confidence_support_applied")
    if confidence == "LOW":
        warnings.append("INTELLIGENCE_LOW_CONFIDENCE")
    if "IMAGE_TITLE_CONFLICT_REVIEW_REQUIRED" in warnings:
        confidence = "MEDIUM" if confidence == "HIGH" else confidence

    intelligence_status = "READY" if confidence in {"HIGH", "MEDIUM"} and family != "UNKNOWN_REVIEW_REQUIRED" else "NEEDS_REVIEW"
    if family == "UNKNOWN_REVIEW_REQUIRED":
        warnings.append("UNKNOWN_REVIEW_REQUIRED")

    return ProductIntelligenceProfile(
        product_id=_first_non_empty(payload.get("id"), payload.get("product_id")),
        source=source,
        normalized_title=_first_non_empty(
            payload.get("product_short_name"),
            payload.get("product_display_name"),
            payload.get("raw_product_title"),
        )
        or "",
        brand=_first_non_empty(payload.get("brand")),
        group=profile["group"],
        sub_group=profile["sub_group"],
        type_of_product=profile["type_of_product"],
        bosmax_product_family=family,
        package_form=profile["package_form"],
        physical_state=profile["physical_state"],
        product_scale_class=profile["product_scale_class"],
        handling_profile=profile["handling_profile"],
        scene_profile=profile["scene_profile"],
        camera_profile=profile["camera_profile"],
        copy_route=copy_route,
        claim_gate=claim_gate,
        claim_tokens=claim_tokens,
        copy_formula=profile["copy_formula"],
        destination_readiness=readiness,
        sales_metrics=sales_metrics,
        image_analysis=image_analysis,
        confidence=confidence,
        warnings=_unique(warnings),
        provenance=_unique(provenance),
        intelligence_status=intelligence_status,
        taxonomy_conflict=taxonomy_conflict,
        taxonomy_conflict_reason=taxonomy_conflict_reason,
        source_taxonomy=source_taxonomy,
    ).model_dump()


def resolve_product_intelligence_catalog_projection(
    product: dict[str, Any],
    *,
    include_reconciliation: bool = False,
    include_sales_metrics: bool = False,
) -> dict[str, Any]:
    """Build the cheap, read-only fields needed to browse the product registry.

    The full profile intentionally performs Product Truth reconciliation, image
    analysis, and sales-workbook matching.  Those operations are correct for a
    detail row, but they are not necessary for deciding which rows belong on a
    paginated catalog page.  Keep this projection deterministic and provider-free;
    the API still calls ``resolve_product_intelligence_profile`` through the normal
    detail enricher when a non-registry response needs the complete row.

    ``include_reconciliation`` is reserved for registry filters that need the
    final confidence/status label.  It is deliberately opt-in so the default
    catalog request does not perform a full-catalog Product Truth pass.
    ``include_sales_metrics`` is likewise opt-in for the two sales sorts.
    """
    payload = dict(product)
    family, family_reason, taxonomy_conflict, taxonomy_conflict_reason = _resolve_family(payload)
    family_profile = _profile_for_family(family)
    copy_route = family_profile["copy_route"]
    claim_gate, claim_tokens, _claim_warnings = _resolve_claim_gate(
        payload,
        family,
        copy_route,
    )
    source = _coerce_source(payload.get("source"))
    source_taxonomy = {
        "category": _first_non_empty(payload.get("category")),
        "subcategory": _first_non_empty(payload.get("subcategory")),
        "type": _first_non_empty(payload.get("type")),
    }
    confidence = _resolve_confidence(
        family_reason,
        taxonomy_conflict,
        family,
        source_taxonomy,
    )

    if include_reconciliation:
        if source == "MANUAL":
            # ProductTruthService intentionally treats the manual lane as
            # SOURCE_ANCHOR_UNVERIFIED, which caps its reconciliation label at
            # LOW.  Avoid rebuilding the full profile for every manual row.
            confidence = "LOW"
        else:
            from agent.services.product_truth_service import ProductTruthService

            truth_profile = ProductTruthService.build_computed_profile(payload)
            reconciliation = truth_profile.reconciliation
            if (
                reconciliation.confidence_label == "NEEDS_REVIEW"
                or "FLAG_CATEGORY_BOUNDARY_LOCK_VIOLATION"
                in reconciliation.contradiction_flags
            ):
                confidence = "LOW"
            elif reconciliation.confidence_label == "LOW":
                confidence = "LOW"
            elif reconciliation.confidence_label == "MEDIUM" and confidence == "HIGH":
                confidence = "MEDIUM"

    result: dict[str, Any] = {
        "source": source,
        "group": family_profile["group"],
        "sub_group": family_profile["sub_group"],
        "type_of_product": family_profile["type_of_product"],
        "bosmax_product_family": family,
        "package_form": family_profile["package_form"],
        "physical_state": family_profile["physical_state"],
        "product_scale_class": family_profile["product_scale_class"],
        "handling_profile": family_profile["handling_profile"],
        "scene_profile": family_profile["scene_profile"],
        "camera_profile": family_profile["camera_profile"],
        "copy_route": copy_route,
        "claim_gate": claim_gate,
        "claim_tokens": claim_tokens,
        "copy_formula": family_profile["copy_formula"],
        "destination_readiness": {},
        "intelligence_confidence": confidence,
        "intelligence_status": (
            "READY"
            if confidence in {"HIGH", "MEDIUM"}
            and family != "UNKNOWN_REVIEW_REQUIRED"
            else "NEEDS_REVIEW"
        ),
        "taxonomy_conflict": taxonomy_conflict,
        "taxonomy_conflict_reason": taxonomy_conflict_reason,
        "bosmax_source_taxonomy_conflict": taxonomy_conflict,
        "bosmax_source_taxonomy_conflict_reason": taxonomy_conflict_reason,
        "bosmax_product_family_reason": family_reason,
    }

    if include_sales_metrics:
        result.update(resolve_product_sales_metrics_catalog_projection(payload))

    return result


def resolve_product_sales_metrics_catalog_projection(
    product: dict[str, Any],
) -> dict[str, Any]:
    """Return sales fields for a bounded page row without full enrichment."""
    sales_metrics, _sales_provenance = _resolve_sales_metrics(dict(product))
    return {
        "sales_metrics": sales_metrics.model_dump(),
        "product_sold_count": sales_metrics.product_sold_count,
        "shop_total_sold_count": sales_metrics.shop_total_sold_count,
    }


def inject_product_intelligence_fields(product: dict[str, Any], profile: dict[str, Any]) -> dict[str, Any]:
    payload = dict(product)
    payload["product_intelligence"] = profile
    payload["group"] = profile["group"]
    payload["sub_group"] = profile["sub_group"]
    payload["type_of_product"] = profile["type_of_product"]
    payload["bosmax_product_family"] = profile["bosmax_product_family"]
    payload["package_form"] = profile["package_form"]
    payload["physical_state"] = profile["physical_state"]
    payload["product_scale_class"] = profile["product_scale_class"]
    payload["handling_profile"] = profile["handling_profile"]
    payload["scene_profile"] = profile["scene_profile"]
    payload["camera_profile"] = profile["camera_profile"]
    payload["copy_route"] = profile["copy_route"]
    payload["claim_gate"] = profile["claim_gate"]
    payload["claim_tokens"] = profile["claim_tokens"]
    payload["copy_formula"] = profile["copy_formula"]
    payload["destination_readiness"] = profile["destination_readiness"]
    payload["sales_metrics"] = profile["sales_metrics"]
    payload["image_analysis"] = profile["image_analysis"]
    payload["intelligence_confidence"] = profile["confidence"]
    payload["intelligence_status"] = profile["intelligence_status"]
    payload["intelligence_warnings"] = profile["warnings"]
    payload["intelligence_provenance"] = profile["provenance"]
    payload["taxonomy_conflict"] = profile["taxonomy_conflict"]
    payload["taxonomy_conflict_reason"] = profile["taxonomy_conflict_reason"]
    payload["bosmax_source_taxonomy_conflict"] = profile["taxonomy_conflict"]
    payload["bosmax_source_taxonomy_conflict_reason"] = profile["taxonomy_conflict_reason"]
    payload["bosmax_product_family_reason"] = next(
        (
            entry
            for entry in profile["provenance"]
            if entry.startswith("title_evidence:")
            or entry.startswith("family_resolver:")
            or entry.startswith("taxonomy_fallback:")
        ),
        None,
    )
    payload["shop_count"] = profile["sales_metrics"]["shop_count"]
    payload["shop_names"] = profile["sales_metrics"]["shop_names"]
    payload["sold_count"] = profile["sales_metrics"]["sold_count"]
    payload["product_sold_count"] = profile["sales_metrics"]["product_sold_count"]
    payload["shop_total_sold_count"] = profile["sales_metrics"]["shop_total_sold_count"]
    payload["sold_count_metric_scope"] = profile["sales_metrics"]["sold_count_metric_scope"]
    payload["sold_count_truth_status"] = profile["sales_metrics"]["sold_count_truth_status"]
    payload["sales_metric_warnings"] = profile["sales_metrics"]["sales_metric_warnings"]
    payload["sales_metric_provenance"] = profile["sales_metrics"]["sales_metric_provenance"]
    payload["sales_metrics_source"] = profile["sales_metrics"]["sales_metrics_source"]
    payload["sales_metrics_batch_id"] = profile["sales_metrics"]["sales_metrics_batch_id"]
    payload["matched_file_type"] = profile["sales_metrics"]["matched_file_type"]
    payload["matched_by"] = profile["sales_metrics"]["matched_by"]
    payload["raw_metric_column"] = profile["sales_metrics"]["raw_metric_column"]
    payload["image_analysis_status"] = profile["image_analysis"]["status"]
    return payload


async def _load_profile_for_row(product: dict[str, Any]) -> dict[str, Any]:
    return resolve_product_intelligence_profile(product)


async def resolve_product_intelligence_request(
    request_input: ProductIntelligenceResolveRequest | dict[str, Any],
) -> dict[str, Any]:
    request = (
        request_input
        if isinstance(request_input, ProductIntelligenceResolveRequest)
        else ProductIntelligenceResolveRequest.model_validate(request_input)
    )
    if request.product_id:
        product = await crud.get_product(request.product_id)
        if not product:
            return {
                "status": "PRODUCT_NOT_FOUND",
                "product_id": request.product_id,
            }
        merged = dict(product)
        if request.product_payload:
            merged.update(request.product_payload)
        return resolve_product_intelligence_profile(merged)
    if request.product_payload:
        return resolve_product_intelligence_profile(dict(request.product_payload))
    return {
        "status": "PRODUCT_CONTEXT_REQUIRED",
        "warnings": ["PRODUCT_CONTEXT_REQUIRED"],
    }


async def get_product_intelligence_by_id(product_id: str) -> dict[str, Any]:
    product = await crud.get_product(product_id)
    if not product:
        return {
            "status": "PRODUCT_NOT_FOUND",
            "product_id": product_id,
        }
    profile = resolve_product_intelligence_profile(product)
    profile["lifecycle_status"] = resolve_lifecycle_status(product)
    if profile["lifecycle_status"] == "ARCHIVED":
        profile["safe_to_generate_prompt"] = False
        profile["blocker"] = "PRODUCT_ARCHIVED"
    return profile


def _distribution(profiles: list[dict[str, Any]], key: str) -> dict[str, int]:
    counter = Counter(str(profile.get(key) or "UNKNOWN") for profile in profiles)
    return dict(sorted(counter.items(), key=lambda item: (-item[1], item[0])))


async def get_product_intelligence_summary() -> dict[str, Any]:
    products = await crud.list_products(limit=10000, include_archived=False)
    profiles = [resolve_product_intelligence_profile(product) for product in products]
    products_by_source = Counter(_coerce_source(product.get("source")) for product in products)
    products_by_current_category = Counter(
        _first_non_empty(product.get("category")) or "__MISSING__"
        for product in products
    )
    products_by_current_type = Counter(
        _first_non_empty(product.get("type")) or "__MISSING__"
        for product in products
    )
    summary = ProductIntelligenceSummaryResponse(
        total_products=len(products),
        products_by_source=dict(sorted(products_by_source.items(), key=lambda item: (-item[1], item[0]))),
        products_by_current_category=dict(sorted(products_by_current_category.items(), key=lambda item: (-item[1], item[0]))),
        products_by_current_type=dict(sorted(products_by_current_type.items(), key=lambda item: (-item[1], item[0]))),
        products_with_missing_category_or_type=sum(
            1
            for product in products
            if not _first_non_empty(product.get("category")) or not _first_non_empty(product.get("type"))
        ),
        products_with_source_taxonomy_conflict_risk=sum(1 for profile in profiles if profile["taxonomy_conflict"]),
        products_with_image_available=sum(
            1
            for profile in profiles
            if profile["image_analysis"]["status"]
            not in {"IMAGE_MISSING", "IMAGE_INACCESSIBLE", "UNSUPPORTED_IMAGE_FORMAT"}
        ),
        products_with_image_not_available=sum(
            1 for profile in profiles if profile["image_analysis"]["status"] == "IMAGE_MISSING"
        ),
        products_with_image_not_analyzed=sum(
            1
            for profile in profiles
            if profile["image_analysis"]["status"] in {"NOT_ANALYZED", "VISION_PROVIDER_NOT_CONFIGURED"}
        ),
        products_with_sold_count_available=sum(1 for profile in profiles if profile["sales_metrics"]["sold_count"] is not None),
        products_with_shop_count_available=sum(1 for profile in profiles if profile["sales_metrics"]["shop_count"] is not None),
        products_with_shop_names_available=sum(1 for profile in profiles if bool(profile["sales_metrics"]["shop_names"])),
        group_distribution=_distribution(profiles, "group"),
        copy_route_distribution=_distribution(profiles, "copy_route"),
        claim_gate_distribution=_distribution(profiles, "claim_gate"),
        confidence_distribution=_distribution(profiles, "confidence"),
        sample_conflicts=[
            {
                "product_id": profile.get("product_id"),
                "normalized_title": profile.get("normalized_title"),
                "group": profile.get("group"),
                "bosmax_product_family": profile.get("bosmax_product_family"),
                "taxonomy_conflict_reason": profile.get("taxonomy_conflict_reason"),
            }
            for profile in profiles
            if profile["taxonomy_conflict"]
        ][:10],
    )
    return summary.model_dump()


async def get_product_intelligence_backfill_preview() -> dict[str, Any]:
    products = await crud.list_products(limit=10000, include_archived=False)
    profiles = [resolve_product_intelligence_profile(product) for product in products]
    failures = [
        {
            "product_id": profile.get("product_id"),
            "normalized_title": profile.get("normalized_title"),
            "group": profile.get("group"),
            "warnings": profile.get("warnings", []),
            "confidence": profile.get("confidence"),
        }
        for profile in profiles
        if profile["confidence"] == "LOW"
    ]
    conflicts = [
        {
            "product_id": profile.get("product_id"),
            "normalized_title": profile.get("normalized_title"),
            "group": profile.get("group"),
            "bosmax_product_family": profile.get("bosmax_product_family"),
            "taxonomy_conflict_reason": profile.get("taxonomy_conflict_reason"),
        }
        for profile in profiles
        if profile["taxonomy_conflict"]
    ]
    payload = ProductIntelligenceBackfillPreviewResponse(
        total_products=len(products),
        resolved=sum(1 for profile in profiles if profile["group"] != "UNKNOWN_REVIEW_REQUIRED"),
        high_confidence=sum(1 for profile in profiles if profile["confidence"] == "HIGH"),
        medium_confidence=sum(1 for profile in profiles if profile["confidence"] == "MEDIUM"),
        low_confidence=sum(1 for profile in profiles if profile["confidence"] == "LOW"),
        needs_review=sum(1 for profile in profiles if profile["intelligence_status"] == "NEEDS_REVIEW"),
        taxonomy_conflicts=len(conflicts),
        copy_route_distribution=_distribution(profiles, "copy_route"),
        claim_gate_distribution=_distribution(profiles, "claim_gate"),
        group_distribution=_distribution(profiles, "group"),
        sample_failures=failures[:10],
        sample_conflicts=conflicts[:10],
        write_back_status="READ_ONLY_NO_DB_WRITES",
    )
    return payload.model_dump()
