from __future__ import annotations

import os
import re
from datetime import datetime, timezone
from functools import lru_cache
from pathlib import Path
from typing import Any, Optional

import yaml
from fastapi import APIRouter, HTTPException
from openpyxl import load_workbook
from pydantic import BaseModel

from agent import config as agent_config
from agent import runtime_release
from agent.config import OPERATOR_PACK_DIR
from agent.db import crud
from agent.db.schema import get_db
from agent.services import batch_executor
from agent.services.flow_client import get_flow_client
from agent.services.product_intelligence import enrich_product
from agent.services.product_mapping import resolve_product_mapping
from agent.services.product_preflight import build_product_preflight
from agent.services.reporting_service import catalog_population_summary

router = APIRouter(prefix="/api/operator", tags=["operator"])


class OperatorProduct(BaseModel):
    product_id: str | None = None
    product_name: str
    raw_product_title: str | None = None
    product_display_name: str | None = None
    product_short_name: str | None = None
    category: str
    sub_category: str
    type_angle: str
    product_type: str | None = None
    silo_id: str | None = None
    trigger_id: str | None = None
    submode_formula: str | None = None
    mode_recommendations: list[str] = []
    copywriting_angle: str | None = None
    claim_risk_level: str | None = None
    mapping_source: str | None = None
    mapping_confidence: str | None = None
    missing_fields: list[str] = []
    raw_category: str | None = None
    avg_price_rm: float | None = None
    status: str | None = None
    copy_angle: str | None = None
    hook: str | None = None
    usp_1: str | None = None
    usp_2: str | None = None
    usp_3: str | None = None
    body: str | None = None
    cta: str | None = None
    shop_name: str | None = None
    image_url: str | None = None
    source_url: str | None = None
    tiktok_product_url: str | None = None
    commission_rate: str | None = None
    commission_amount: float | None = None


class WorkbookSummary(BaseModel):
    workbook: str
    sheets: list[str]


class ContentPackSummary(BaseModel):
    pack_dir: str
    available: bool
    files: list[str]
    engines: list[str]
    durations_by_engine: dict[str, list[str]]
    avatars: list[str]
    headwear_styles: list[str]
    camera_styles: list[str]
    product_types: list[str]
    triggers: list[str]
    silos: list[str]
    formulas: list[str]
    materials: list[str]
    language_defaults: list[str]
    products: list[OperatorProduct]
    workbooks: list[WorkbookSummary]
    notes: list[str]


class BlueprintInput(BaseModel):
    product_name: str
    category: str = ""
    sub_category: str = ""
    type_angle: str = ""
    product_type: str
    target_language: str
    duration_target: str
    engine_id: str
    avatar_id: str
    headwear_style: str
    camera_style: str
    scene_context: str
    trigger_id: str
    silo_id: str
    submode_formula: str
    hook: str = ""
    usp_1: str = ""
    usp_2: str = ""
    usp_3: str = ""
    body: str = ""
    cta: str = ""
    material: str = "realistic"
    orientation: str = "VERTICAL"


class BlueprintCharacter(BaseModel):
    name: str
    entity_type: str
    description: str


class BlueprintScene(BaseModel):
    display_order: int
    prompt: str
    image_prompt: str
    video_prompt: str
    character_names: list[str]
    chain_type: str = "ROOT"


class BlueprintResponse(BaseModel):
    project: dict[str, Any]
    video: dict[str, Any]
    scenes: list[BlueprintScene]
    notes: list[str]


class FlowReadinessSmokeRequest(BaseModel):
    product_id: str | None = None
    batch_id: str | None = None
    variant_id: str | None = None
    mode: str = "F2V"


class FlowProviderReadinessRequest(BaseModel):
    provider_browser_authority_mode: str = "DEDICATED_CDP_UAT"
    owner_profile_label: str | None = None
    extension_source_path: str | None = None
    provider_execution_route: str = "EXACT_PRODUCT_DETERMINISTIC_COMPOSITE"
    scene_scaffold_route: str = "AGENT_T2V"
    flow_tab_id: int | None = None
    mode: str = "T2V"


_AGENT_T2V_PROVIDER_ROUTES_WITHOUT_UI_COMPOSER = frozenset(
    {
        "EXACT_PRODUCT_DETERMINISTIC_COMPOSITE",
        "API_FIRST_GENERATIVE_REFERENCE",
    }
)


def _provider_ui_composer_required(body: FlowProviderReadinessRequest) -> bool:
    """Return whether provider readiness must prove the legacy DOM composer.

    AGENT_T2V routes use the authenticated Flow-agent transport and the exact
    project/session challenge; they do not drive the legacy Frames/Ingredients
    composer.  Requiring that selector proof makes unrelated project-sidebar
    labels look like an active sub-mode and blocks a valid agent route.
    """
    route = str(body.provider_execution_route or "").strip().upper()
    scaffold = str(body.scene_scaffold_route or "").strip().upper()
    return not (
        scaffold == "AGENT_T2V"
        and route in _AGENT_T2V_PROVIDER_ROUTES_WITHOUT_UI_COMPOSER
    )


class ReloadFlowTabRequest(BaseModel):
    pass


class OpenTargetFlowProjectRequest(BaseModel):
    flow_project_url: str


class OpenFlowNewProjectRequest(BaseModel):
    mode: str = "F2V"


class FlowPageStateDiagnosticRequest(BaseModel):
    mode: str = "F2V"


def _pack_file(name: str) -> Path:
    return OPERATOR_PACK_DIR / name


def _require_pack() -> Path:
    if not OPERATOR_PACK_DIR.exists():
        raise HTTPException(404, f"Operator pack directory not found: {OPERATOR_PACK_DIR}")
    return OPERATOR_PACK_DIR


def _load_yaml_file(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    data = yaml.safe_load(path.read_text(encoding="utf-8"))
    return data if isinstance(data, dict) else {}


def _merged_registry_values(values: list[str], required: list[str]) -> list[str]:
    ordered: list[str] = []
    seen: set[str] = set()
    for item in [*values, *required]:
        if not item or item in seen:
            continue
        ordered.append(item)
        seen.add(item)
    return ordered


def _clean_text(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def _normalized_product_key(value: str | None) -> str:
    cleaned = _clean_text(value or "")
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned.casefold()


def _seconds(duration: str) -> int:
    match = re.search(r"(\d+)", duration or "")
    return int(match.group(1)) if match else 8


def _scene_count(duration_target: str) -> int:
    sec = _seconds(duration_target)
    if sec <= 30:
        return 4
    return 8


def _timed_video_prompt(core_action: str, benefit: str, cta: str) -> str:
    return (
        f'0-3s: {core_action}. '
        f'3-6s: {benefit}. '
        f'6-8s: {cta}.'
    )


def _find_sheet_header(ws) -> list[str]:
    for row in ws.iter_rows(values_only=True, max_row=10):
        values = [str(v).strip() if v is not None else "" for v in row]
        normalized = {value.lower() for value in values if value}
        if "product name" in normalized or "product title" in normalized:
            return values
    return []


def _row_mapping(headers: list[str], values: list[Any]) -> dict[str, Any]:
    return {headers[i]: values[i] for i in range(min(len(headers), len(values))) if headers[i]}


def _parse_commission_amount(price_value: Any, commission_rate: str | None) -> float | None:
    if price_value in {None, ""} or not commission_rate:
        return None
    from agent.services.product_intelligence import normalize_commission_rate
    frac = normalize_commission_rate(commission_rate)
    if frac is None:
        return None
    rate = float(frac)
    price_text = str(price_value).replace('RM', '').strip()
    if '-' in price_text:
        price_text = price_text.split('-', 1)[0].strip()
    try:
        price = float(price_text)
    except ValueError:
        return None
    return round(price * rate, 2)


def _build_product_asset_lookup(wb) -> dict[str, dict[str, Any]]:
    lookup: dict[str, dict[str, Any]] = {}
    sheet_configs = [
        {
            "sheet": "Product Sales Rank",
            "name_field": "Product Name",
            "image_fields": ["Product Image"],
            "source_fields": ["FastMoss Product Detail"],
            "tiktok_fields": ["TikTok Product Detail"],
            "commission_fields": ["Commission Rate"],
            "price_fields": ["Price"],
        },
        {
            "sheet": "Most Promoted Products",
            "name_field": "Product Name",
            "image_fields": ["Product Image Link"],
            "source_fields": ["FastMoss Product Detail"],
            "tiktok_fields": ["TikTok Product Detail"],
            "commission_fields": ["Commission Rate"],
            "price_fields": ["Price"],
        },
        {
            "sheet": "Video Product List",
            "name_field": "Product Title",
            "image_fields": ["Product Cover"],
            "source_fields": ["FastMoss Product Detail Page Link"],
            "tiktok_fields": ["TikTok Product Link"],
            "commission_fields": [],
            "price_fields": ["Product Price"],
        },
        {
            "sheet": "Product Search Data",
            "name_field": "Product Name",
            "image_fields": ["Product Image"],
            "source_fields": ["FastMoss", "FastMoss Shop"],
            "tiktok_fields": ["TikTok"],
            "commission_fields": ["Commission Rate"],
            "price_fields": ["Selling Price / Price"],
        },
        {
            "sheet": "New Products Ranking",
            "name_field": "Product Name",
            "image_fields": ["Cover"],
            "source_fields": [],
            "tiktok_fields": [],
            "commission_fields": ["Commission Rate"],
            "price_fields": ["Selling Price"],
        },
    ]

    for config in sheet_configs:
        if config["sheet"] not in wb.sheetnames:
            continue
        ws = wb[config["sheet"]]
        headers = _find_sheet_header(ws)
        if not headers:
            continue
        for row in ws.iter_rows(values_only=True):
            values = list(row)
            if not any(v is not None and str(v).strip() for v in values):
                continue
            if headers and values[:len(headers)] == headers[:len(values[:len(headers)])]:
                continue
            data = _row_mapping(headers, values)
            name = _clean_text(str(data.get(config["name_field"]) or ""))
            if not name:
                continue
            key = _normalized_product_key(name)
            record = lookup.setdefault(key, {})
            image_url = next((_clean_text(str(data.get(field) or "")) for field in config["image_fields"] if data.get(field)), "")
            source_url = next((_clean_text(str(data.get(field) or "")) for field in config["source_fields"] if data.get(field)), "")
            tiktok_url = next((_clean_text(str(data.get(field) or "")) for field in config["tiktok_fields"] if data.get(field)), "")
            commission_rate = next((_clean_text(str(data.get(field) or "")) for field in config["commission_fields"] if data.get(field)), "")
            price_value = next((data.get(field) for field in config["price_fields"] if data.get(field) is not None), None)
            commission_amount = _parse_commission_amount(price_value, commission_rate)

            if image_url and not record.get("image_url"):
                record["image_url"] = image_url
            if source_url and not record.get("source_url"):
                record["source_url"] = source_url
            if tiktok_url and not record.get("tiktok_product_url"):
                record["tiktok_product_url"] = tiktok_url
            if commission_rate and not record.get("commission_rate"):
                record["commission_rate"] = commission_rate
            if commission_amount is not None and record.get("commission_amount") is None:
                record["commission_amount"] = commission_amount
    return lookup


@lru_cache(maxsize=1)
def _content_pack_summary() -> ContentPackSummary:
    pack_dir = _require_pack()
    master = _load_yaml_file(_pack_file("MASTER_IGNITION_TEMPLATE.yaml"))
    script_registry = _load_yaml_file(_pack_file("SCRIPT_REGISTRY_UNIFIED.yaml"))

    files = sorted(p.name for p in pack_dir.iterdir() if p.is_file())

    durations_by_engine = {
        key: list(values or [])
        for key, values in (master.get("duration_target") or {}).items()
        if isinstance(values, list)
    }

    workbook_summaries = []
    for workbook_name in [
        "Category and product list.xlsx",
        "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx",
    ]:
        path = _pack_file(workbook_name)
        if path.exists():
            wb = load_workbook(path, read_only=True, data_only=True)
            workbook_summaries.append(
                WorkbookSummary(workbook=workbook_name, sheets=wb.sheetnames)
            )

    products = _load_products()

    return ContentPackSummary(
        pack_dir=str(pack_dir),
        available=True,
        files=files,
        engines=list(master.get("engine_id") or []),
        durations_by_engine=durations_by_engine,
        avatars=list(master.get("avatar_id") or []),
        headwear_styles=list(master.get("headwear_style") or []),
        camera_styles=list(master.get("camera_style") or []),
        product_types=_merged_registry_values(list(master.get("product_type") or ["STEALTH", "DIRECT"]), ["UNIVERSAL", "STEALTH"]),
        triggers=_merged_registry_values(list(master.get("trigger_id") or []), ["TRUST_01", "CONFIDENCE_01", "AUTHORITY_01", "COMFORT_01", "EGO_01", "GIFTING_01"]),
        silos=_merged_registry_values(list(master.get("silo_id") or []), ["baby_care_universal_01", "fashion_mass_01", "perfume_mass_01", "fnb_mass_01", "household_or_beauty_mass_01", "electronics_mass_01", "household_mass_01", "health_supp_stealth_01", "stationery_mass_01"]),
        formulas=_merged_registry_values(list(master.get("submode_formula") or []), ["PAS", "AIDA", "TRUST_STACK", "FEATURE_BENEFIT"]),
        materials=["realistic", "3d_pixar", "anime"],
        language_defaults=["Malay", "English"],
        products=products,
        workbooks=workbook_summaries,
        notes=[
            "This operator pack supplies registries, copy hooks, and product intelligence.",
            "Current Flow Kit uses canonical Google Flow labels: Images, Ingredients, Frames, and Text to Video.",
            "Scene context remains the visual authority; hook/USP/CTA are used for wording and scene framing only.",
            f"Malay tone rules are sourced from SCRIPT_REGISTRY_UNIFIED.yaml ({len(script_registry.get('language_tone_rules', {}).get('rules', []))} rules detected).",
        ],
    )


def _load_products(limit: int = 120) -> list[OperatorProduct]:
    path = _pack_file("FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx")
    if not path.exists():
        return []

    wb = load_workbook(path, read_only=True, data_only=True)
    if "Copywriting_Product_Map" not in wb.sheetnames:
        return []

    asset_lookup = _build_product_asset_lookup(wb)

    ws = wb["Copywriting_Product_Map"]
    headers: list[str] = []
    products: list[OperatorProduct] = []

    for row in ws.iter_rows(values_only=True):
        values = list(row)
        if not any(v is not None and str(v).strip() for v in values):
            continue
        if values and values[0] == "Rank":
            headers = [str(v).strip() if v is not None else "" for v in values]
            continue
        if not headers:
            continue
        data = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
        name = _clean_text(str(data.get("Product Name") or ""))
        if not name:
            continue
        asset_data = asset_lookup.get(_normalized_product_key(name), {})
            
        # Basic normalization for Excel source
        short_name = re.sub(r'(\d+PCS|Premium|All size|S/M/L/XL/XXL/XXXL|Ultra-thin|breathable|disposable|tape|pull-ups|disposable diaper tape diaper pants pull-ups)', '', name, flags=re.IGNORECASE)
        short_name = " ".join(short_name.split()[:4]).strip()
        display_name = " ".join(name.split()[:9]).strip()

        mapping = resolve_product_mapping(
            product={
                "raw_product_title": name,
                "product_display_name": display_name,
                "product_short_name": short_name,
                "category": _clean_text(str(data.get("Category") or "")),
                "subcategory": _clean_text(str(data.get("Sub Category") or "")),
                "type": _clean_text(str(data.get("Type / Product Angle") or "")),
                "source": "FASTMOSS",
            },
            source_hint="FASTMOSS",
        )

        products.append(
            OperatorProduct(
                product_name=name,
                raw_product_title=name,
                product_short_name=mapping["product_short_name"],
                product_display_name=display_name,
                category=mapping["category"],
                sub_category=mapping["subcategory"],
                type_angle=mapping["type"],
                product_type=mapping["product_type"] or None,
                silo_id=mapping["silo"] or None,
                trigger_id=mapping["trigger_id"] or None,
                submode_formula=mapping["formula"] or None,
                mode_recommendations=list(mapping.get("mode_recommendations", [])),
                copywriting_angle=mapping.get("copywriting_angle") or None,
                claim_risk_level=mapping.get("claim_risk_level") or None,
                mapping_source=mapping.get("mapping_source") or None,
                mapping_confidence=mapping.get("mapping_confidence") or None,
                missing_fields=list(mapping.get("missing_fields", [])),
                image_url=asset_data.get("image_url") or None,
                source_url=asset_data.get("source_url") or None,
                tiktok_product_url=asset_data.get("tiktok_product_url") or None,
                commission_rate=asset_data.get("commission_rate") or None,
                commission_amount=asset_data.get("commission_amount"),
                raw_category=_clean_text(str(data.get("Raw Category") or "")) or None,
                avg_price_rm=float(data["Avg Price (RM)"]) if data.get("Avg Price (RM)") is not None else None,
                status=_clean_text(str(data.get("Product Status") or "")) or None,
                copy_angle=_clean_text(str(data.get("Copywriting Angle / Hook Direction") or "")) or None,
                hook=_clean_text(str(data.get("Hook") or "")) or None,
                usp_1=_clean_text(str(data.get("USP 1") or "")) or None,
                usp_2=_clean_text(str(data.get("USP 2") or "")) or None,
                usp_3=_clean_text(str(data.get("USP 3") or "")) or None,
                body=_clean_text(str(data.get("Body") or "")) or None,
                cta=_clean_text(str(data.get("CTA") or "")) or None,
                shop_name=_clean_text(str(data.get("Shop Name") or "")) or None,
            )
        )
        if len(products) >= limit:
            break
    return products


async def _find_latest_batch_context(product_id: str | None) -> dict[str, Any] | None:
    if not product_id:
        return None
    db = await get_db()
    cursor = await db.execute(
        """
        SELECT b.id AS batch_id, b.mode AS batch_mode, bv.variant_id, bv.queue_status
        FROM batch_variant bv
        JOIN batch b ON b.id = bv.batch_id
        WHERE bv.product_id = ?
        ORDER BY b.created_at DESC, bv.variation_index ASC
        LIMIT 1
        """,
        (product_id,),
    )
    row = await cursor.fetchone()
    return dict(row) if row else None


def _extract_numeric_credit(value: Any) -> int | float | None:
    """Find a finite numeric credit balance without treating auth metadata as proof."""
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return value if value == value and value not in (float("inf"), float("-inf")) else None
    if isinstance(value, str):
        stripped = value.strip()
        if re.fullmatch(r"-?\d+(?:\.\d+)?", stripped):
            try:
                return float(stripped) if "." in stripped else int(stripped)
            except ValueError:
                return None
        return None
    if isinstance(value, dict):
        credit_keys = {"credits", "creditbalance", "availablecredits", "remainingcredits", "balance"}
        for key, child in value.items():
            if str(key).lower().replace("_", "") in credit_keys:
                found = _extract_numeric_credit(child)
                if found is not None:
                    return found
        for child in value.values():
            found = _extract_numeric_credit(child)
            if found is not None:
                return found
    if isinstance(value, (list, tuple)):
        for child in value:
            found = _extract_numeric_credit(child)
            if found is not None:
                return found
    return None


_ACTIVE_PROVIDER_JOB_STATUSES = frozenset({
    "INITIAL_SUBMITTING",
    "INITIAL_POLLING",
    "INITIAL_RECOVERY_REQUIRED",
    "EXTEND_SUBMITTING",
    "EXTEND_POLLING",
    "CONCAT_SUBMITTING",
    "CONCAT_POLLING",
    "FINAL_SAVING",
    "RUNNING",
    "GENERATING",
    "IN_PROGRESS",
    "POLLING",
})


async def _provider_readiness_runtime() -> dict[str, Any]:
    origin_main = runtime_release.git_output("rev-parse", "origin/main")
    provenance = runtime_release.resolve_provenance(
        runtime_release.source_root(),
        agent_config.DB_PATH,
        origin_main=origin_main,
    )
    provenance["origin_main"] = origin_main
    provenance["runtime_current_main"] = bool(
        provenance.get("runtime_sha")
        and provenance.get("origin_main")
        and provenance.get("runtime_sha") == provenance.get("origin_main")
        and provenance.get("canonical_runtime") is True
        and provenance.get("source_stale") is False
        and provenance.get("release_dirty") is False
        and provenance.get("db_canonical") is True
        and provenance.get("bundle_matches") is True
    )
    return provenance


def _classify_flow_primary_blocker(
    extension_connected: bool,
    composer: dict[str, Any],
    smoke: dict[str, Any] | None,
) -> str | None:
    if not extension_connected:
        return "EXTENSION_RUNTIME_STALE_NEEDS_RELOAD"

    detail = str(composer.get("raw_error") or composer.get("detail") or composer.get("error") or "")
    if any(token in detail for token in ["ERR_UNKNOWN_MESSAGE_TYPE", "ERR_CONTENT_SCRIPT_STALE", "ERR_NO_RECEIVER", "ERR_MESSAGE_RESPONSE_TIMEOUT", "Timed out waiting"]):
        return "CONTENT_SCRIPT_STALE_OR_NOT_INJECTED"
    if not composer.get("flow_tab_found"):
        return "FLOW_PROJECT_LIST_NOT_EDITOR"
    if composer.get("flow_tab_found") and not composer.get("content_script_loaded", False):
        return "CONTENT_SCRIPT_STALE_OR_NOT_INJECTED"
    if "FLOW_MODE_MISMATCH" in detail or "ABORT_FLOW_MODE_MISMATCH" in detail:
        return "FLOW_MODE_MISMATCH"
    if not composer.get("signed_in_likely", True):
        return "FLOW_EDITOR_NOT_AUTHENTICATED"

    flow_url = str(composer.get("flow_url") or "")
    if "/project/" not in flow_url and "/edit/" not in flow_url:
        return "FLOW_PROJECT_LIST_NOT_EDITOR"
    if composer.get("composer_found") and not composer.get("composer_editable"):
        return "COMPOSER_NOT_EDITABLE"
    if smoke and smoke.get("status") == "FAIL_MODE_MISMATCH":
        return "FLOW_MODE_MISMATCH"
    if composer.get("composer_found") and not composer.get("generate_button_found"):
        return "GENERATE_BUTTON_NOT_FOUND"
    if not composer.get("composer_found"):
        return "FLOW_PROJECT_LIST_NOT_EDITOR"
    return None


def _classify_flow_page_state(diagnostic: dict[str, Any]) -> tuple[str, bool]:
    login_markers = [str(item) for item in diagnostic.get("visible_login_markers") or []]
    loading_markers = [str(item) for item in diagnostic.get("visible_loading_markers") or []]
    error_markers = [str(item) for item in diagnostic.get("visible_error_markers") or []]
    editor_markers = [str(item) for item in diagnostic.get("visible_project_editor_markers") or []]
    composer_markers = [str(item) for item in diagnostic.get("visible_composer_placeholder_markers") or []]
    button_texts = [str(item) for item in diagnostic.get("button_texts") or []]
    aria_labels = [str(item) for item in diagnostic.get("aria_labels") or []]
    body_text = str(diagnostic.get("body_text_first_2000_chars") or "")
    ready_state = str(diagnostic.get("document_ready_state") or "")

    lowered_body = body_text.lower()
    wrong_profile_tokens = ("choose an account", "use another account", "switch account")
    if any(token in lowered_body for token in wrong_profile_tokens):
        return "FLOW_WRONG_GOOGLE_PROFILE", False

    if login_markers:
        return "FLOW_LOGIN_WALL_VISIBLE", False

    error_tokens = ("access denied", "request access", "not found", "403", "404", "permission denied", "you need access")
    if error_markers or any(token in lowered_body for token in error_tokens):
        return "FLOW_PROJECT_ACCESS_DENIED_OR_NOT_FOUND", False

    if ready_state != "complete" or loading_markers:
        return "FLOW_PAGE_LOADING_STUCK", False

    generate_marker_present = any("generate" in value.lower() for value in [*button_texts, *aria_labels])
    selector_miss_proven = bool(
        editor_markers and (composer_markers or generate_marker_present)
        and (not diagnostic.get("composer_found") or not diagnostic.get("generate_button_found"))
    )
    if selector_miss_proven:
        return "FLOW_PROJECT_EDITOR_LOADED_SELECTOR_MISS", True

    if editor_markers:
        return "FLOW_COMPOSER_NOT_PRESENT_ON_THIS_MODE", False

    return "FLOW_PAGE_LOADING_STUCK", False


@router.post("/reload-flow-tab")
async def reload_flow_tab(_: ReloadFlowTabRequest):
    result = await get_flow_client().reload_flow_tab()
    detail = str(result.get("error") or result.get("raw_error") or "")
    primary_blocker = None
    if any(token in detail for token in ["ERR_UNKNOWN_MESSAGE_TYPE", "ERR_CONTENT_SCRIPT_STALE", "ERR_NO_RECEIVER", "ERR_MESSAGE_RESPONSE_TIMEOUT"]):
        primary_blocker = "CONTENT_SCRIPT_STALE_OR_NOT_INJECTED"
    return {
        **result,
        "primary_blocker": primary_blocker,
        "last_checked_at": datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
    }


@router.post("/reload-extension")
async def reload_extension(_: ReloadFlowTabRequest):
    # One-way runtime reload: the extension replies ok then calls
    # chrome.runtime.reload() ~100ms later, so the WS connection drops and
    # re-registers. Callers must poll /api/local-agent/status for reconnect.
    result = await get_flow_client().reload_extension()
    return {
        **result,
        "last_checked_at": datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
    }


@router.post("/open-target-flow-project")
async def open_target_flow_project(body: OpenTargetFlowProjectRequest):
    result = await get_flow_client().open_target_flow_project(body.flow_project_url)
    flow_url = str(result.get("flow_url") or result.get("flow_url_after") or "")
    primary_blocker = None
    if not flow_url or ("/project/" not in flow_url and "/edit/" not in flow_url):
        primary_blocker = "FLOW_PROJECT_EDITOR_NOT_OPEN"
    return {
        **result,
        "flow_project_url": body.flow_project_url,
        "primary_blocker": primary_blocker,
        "last_checked_at": datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
    }


@router.post("/open-flow-new-project")
async def open_flow_new_project(body: OpenFlowNewProjectRequest):
    result = await get_flow_client().open_flow_new_project(body.mode)
    detail = str(result.get("detail") or result.get("error") or result.get("raw_error") or "")
    primary_blocker = None
    if "Extension disconnected" in detail:
        primary_blocker = "EXTENSION_RUNTIME_STALE_NEEDS_RELOAD"
    elif "FLOW_MODE_MISMATCH" in detail or "ABORT_FLOW_MODE_MISMATCH" in detail:
        primary_blocker = "FLOW_MODE_MISMATCH"
    elif not result.get("open_flow_root"):
        primary_blocker = "FLOW_ROOT_OPEN_FAILED"
    elif not result.get("project_list_or_landing_detected"):
        primary_blocker = result.get("error") or "FLOW_PROJECT_LIST_OR_LANDING_NOT_DETECTED"
    elif not result.get("new_project_clicked") and result.get("new_project_clicked") != "SKIPPED_ALREADY_IN_EDITOR":
        primary_blocker = result.get("error") or "FLOW_PROJECT_CREATION_PATH_MISSING"
    elif not result.get("editor_ready"):
        primary_blocker = result.get("error") or "FLOW_EDITOR_NOT_READY"
    return {
        **result,
        "primary_blocker": primary_blocker,
        "last_checked_at": datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
    }


@router.get("/preflight")
async def get_operator_preflight(product_id: str):
    product = await crud.get_product(product_id)
    if not product:
        raise HTTPException(404, "Product not found")
    enriched = await enrich_product(product, persist=True)
    batch_context = await _find_latest_batch_context(product_id)
    return {
        "product_id": product_id,
        "preflight": build_product_preflight(enriched),
        "batch_context": batch_context,
        "product": enriched,
    }


@router.post("/flow-readiness-smoke")
async def flow_readiness_smoke(body: FlowReadinessSmokeRequest):
    status = await get_flow_client().get_status()
    extension_connected = bool(status.get("connected"))
    composer = await get_flow_client().check_flow_composer_ready(body.mode)
    batch_context = None
    if body.batch_id and body.variant_id:
        batch_context = {"batch_id": body.batch_id, "variant_id": body.variant_id}
    elif body.product_id:
        batch_context = await _find_latest_batch_context(body.product_id)

    smoke_result = None
    execute_status = "SKIPPED"
    if batch_context and batch_context.get("batch_id") and batch_context.get("variant_id"):
        smoke_result = await batch_executor.smoke_execute_flow_job(batch_context["batch_id"], batch_context["variant_id"])
        execute_status = "PASS" if smoke_result.get("ok") else "STRUCTURED_FAIL"

    primary_blocker = _classify_flow_primary_blocker(extension_connected, composer, smoke_result)
    return {
        "status": "BLOCKED" if primary_blocker else "READY",
        "checked_mode": body.mode,
        "extension_runtime": "PASS" if extension_connected else "FAIL",
        "flow_tab_found": composer.get("flow_tab_found", False),
        "flow_tab_id": composer.get("flow_tab_id"),
        "flow_url": composer.get("flow_url"),
        "extension_protocol_version": composer.get("extension_protocol_version"),
        "content_script_protocol_version": composer.get("content_script_protocol_version"),
        "content_script_loaded": composer.get("content_script_loaded", False),
        "content_script_alive": composer.get("content_script_alive", False),
        "last_content_script_seen_at": composer.get("last_content_script_seen_at"),
        "signed_in_likely": composer.get("signed_in_likely", False),
        "composer_found": composer.get("composer_found", False),
        "composer_editable": composer.get("composer_editable", False),
        "generate_button_found": composer.get("generate_button_found", False),
        "current_mode_visible": composer.get("current_mode_visible") or composer.get("observed", {}).get("topMode") or "UNKNOWN",
        "blocking_modal_detected": composer.get("blocking_modal_detected", False),
        "flow_composer_ready": bool(composer.get("ok")),
        "execute_flow_job_smoke": execute_status,
        "primary_blocker": primary_blocker,
        "last_checked_at": composer.get("last_checked_at") or datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
        "raw_error": composer.get("raw_error") or composer.get("detail") or composer.get("error"),
        "composer": composer,
        "smoke_result": smoke_result,
        "batch_context": batch_context,
    }


@router.post("/flow-provider-readiness")
async def flow_provider_readiness(body: FlowProviderReadinessRequest):
    """Read-only provider-authority receipt for the selected execution route.

    This endpoint deliberately separates the owner-profile extension bridge
    from the dedicated CDP Browser UAT. It never opens a tab, changes account
    state, submits a provider request, or touches product data.
    """
    authority_mode = str(body.provider_browser_authority_mode or "").strip().upper()
    supported_modes = {"DEDICATED_CDP_UAT", "OWNER_PROFILE_EXTENSION_BRIDGE"}
    runtime = await _provider_readiness_runtime()
    base = {
        "provider_browser_authority_mode": authority_mode,
        "owner_profile_label": body.owner_profile_label,
        "extension_source_path": body.extension_source_path,
        "provider_execution_route": body.provider_execution_route,
        "scene_scaffold_route": body.scene_scaffold_route,
        "ui_composer_required": _provider_ui_composer_required(body),
        "runtime_sha": runtime.get("runtime_sha"),
        "origin_main": runtime.get("origin_main"),
        "runtime_current_main": runtime.get("runtime_current_main") is True,
        "release_dir": runtime.get("release_dir"),
        "source_stale": runtime.get("source_stale"),
        "release_dirty": runtime.get("release_dirty"),
        "db_canonical": runtime.get("db_canonical"),
        "bundle_matches": runtime.get("bundle_matches"),
        "dashboard_bundle": runtime.get("dashboard_bundle"),
        "flow_transport_connected": False,
        "flow_transport_bound_to_provider_browser": False,
        "flow_auth_status": "UNKNOWN",
        "numeric_credit_balance": None,
        "flow_tab_found": False,
        "flow_tab_id": None,
        "flow_project_url": None,
        "flow_project_id": None,
        "content_script_loaded": False,
        "content_script_alive": False,
        "session_challenge_verified": False,
        "same_extension_session": False,
        "same_flow_tab": False,
        "extension_session_id": None,
        "extension_id": None,
        "extension_version": None,
        "extension_build": None,
        "extension_build_match": False,
        "video_job_in_flight": False,
    }
    if authority_mode not in supported_modes:
        return {
            **base,
            "FLOW_PROVIDER_UAT_READY": False,
            "primary_blocker": "INVALID_PROVIDER_BROWSER_AUTHORITY_MODE",
        }

    client = get_flow_client()
    status = await client.get_status(timeout=5)
    extension_connected = client.connected and status.get("connected") is True
    base.update({
        "flow_transport_connected": extension_connected,
        "extension_session_id": status.get("extension_session_id"),
        "extension_id": status.get("extension_id"),
        "extension_version": status.get("extension_version"),
        "extension_build": status.get("extension_build") or status.get("background_build_id"),
    })

    challenge: dict[str, Any] = {}
    if extension_connected:
        challenge = await client.verify_provider_session_challenge(body.flow_tab_id)
        base.update({
            "extension_session_id": challenge.get("extension_session_id")
            or status.get("extension_session_id"),
            "extension_id": challenge.get("extension_id") or status.get("extension_id"),
            "extension_version": challenge.get("extension_version")
            or status.get("extension_version"),
            "extension_build": challenge.get("extension_build")
            or status.get("extension_build")
            or status.get("background_build_id"),
            "extension_build_match": challenge.get("extension_build_match") is True,
            "flow_tab_found": challenge.get("flow_tab_found") is True,
            "flow_tab_id": challenge.get("flow_tab_id"),
            "flow_project_url": challenge.get("flow_project_url") or challenge.get("flow_url"),
            "flow_project_id": challenge.get("flow_project_id"),
            "content_script_loaded": challenge.get("content_script_loaded") is True,
            "content_script_alive": challenge.get("content_script_alive") is True,
            "session_challenge_verified": challenge.get("session_challenge_verified") is True,
            "same_extension_session": challenge.get("same_extension_session") is True,
            "same_flow_tab": challenge.get("same_flow_tab") is True,
            "flow_transport_bound_to_provider_browser": challenge.get("session_challenge_verified") is True,
        })

    credit_payload: Any = None
    if extension_connected:
        credit_response = await client.get_credits()
        credit_payload = credit_response.get("data", credit_response)
    credit_balance = _extract_numeric_credit(credit_payload)
    base["numeric_credit_balance"] = credit_balance

    flow_key_present = bool(
        status.get("flowKeyPresent") or status.get("flow_key_present")
    )
    authenticated = bool(
        extension_connected
        and base["session_challenge_verified"]
        and base["same_extension_session"]
        and base["same_flow_tab"]
        and base["flow_tab_found"]
        and base["content_script_alive"]
        and flow_key_present
        and credit_balance is not None
    )
    base["flow_auth_status"] = "AUTHENTICATED" if authenticated else (
        "UNAUTHENTICATED" if extension_connected else "UNKNOWN"
    )

    try:
        jobs = await crud.list_video_production_jobs(limit=100)
    except Exception as exc:  # pragma: no cover - defensive read-only receipt path
        jobs = []
        base["video_job_error"] = str(exc)
    active_jobs = [
        {
            "job_id": job.get("job_id"),
            "status": job.get("status"),
        }
        for job in jobs
        if str(job.get("status") or "").upper() in _ACTIVE_PROVIDER_JOB_STATUSES
    ]
    base["video_job_in_flight"] = bool(active_jobs)

    composer: dict[str, Any] = {}
    if base["ui_composer_required"] and extension_connected:
        composer = await client.check_flow_composer_ready(body.mode)
    base["composer_found"] = composer.get("composer_found") is True
    base["composer_runtime_ready"] = composer.get("runtime_ready") is True
    base["composer"] = composer

    blocker = None
    if not base["runtime_current_main"]:
        blocker = "RUNTIME_NOT_CURRENT_MAIN"
    elif not extension_connected:
        blocker = "EXTENSION_BRIDGE_NOT_CONNECTED"
    elif not base["same_extension_session"]:
        blocker = "EXTENSION_SESSION_MISMATCH"
    elif not base["session_challenge_verified"]:
        blocker = (
            "FLOW_SESSION_CHALLENGE_FAILED"
            if challenge.get("challenge_nonce_match") is not True
            else "FLOW_SESSION_CHALLENGE_FAILED"
        )
    elif not base["flow_tab_found"] or not base["flow_project_url"] or not base["flow_project_id"]:
        blocker = "FLOW_PROJECT_NOT_FOUND"
    elif not base["content_script_alive"]:
        blocker = "FLOW_CONTENT_SCRIPT_NOT_ALIVE"
    elif not base["extension_build_match"]:
        blocker = "EXTENSION_BUILD_MISMATCH"
    elif not authenticated:
        blocker = "FLOW_AUTH_UNAUTHENTICATED"
    elif base["video_job_in_flight"]:
        blocker = "PROVIDER_JOB_IN_FLIGHT"
    elif base["ui_composer_required"] and not (
        base["composer_found"] and base["composer_runtime_ready"]
    ):
        blocker = "FLOW_COMPOSER_NOT_READY"

    ready = blocker is None
    return {
        **base,
        "flow_key_present": flow_key_present,
        "active_video_jobs": active_jobs,
        "FLOW_PROVIDER_UAT_READY": ready,
        "primary_blocker": "FLOW_PROVIDER_UAT_READY" if ready else blocker,
        "challenge": {
            "challenge_verified": base["session_challenge_verified"],
            "same_extension_session": base["same_extension_session"],
            "same_flow_tab": base["same_flow_tab"],
            "challenge_nonce_match": challenge.get("challenge_nonce_match") is True,
            "content_build_id": challenge.get("content_build_id"),
            "content_script_protocol_version": challenge.get("content_script_protocol_version"),
        },
    }


@router.post("/flow-page-state-diagnostic")
async def flow_page_state_diagnostic(body: FlowPageStateDiagnosticRequest):
    diagnostic = await get_flow_client().flow_page_state_diagnostic(body.mode)
    classification, selector_miss_proven = _classify_flow_page_state(diagnostic)
    return {
        **diagnostic,
        "classification": classification,
        "selector_miss_proven": selector_miss_proven,
        "last_checked_at": datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
    }


def _build_story(body: BlueprintInput) -> str:
    lines = [
        f"ENGINE: {body.engine_id}",
        f"DURATION: {body.duration_target}",
        f"PRODUCT: {body.product_name}",
        f"PRODUCT_TYPE: {body.product_type}",
        f"CATEGORY: {body.category}",
        f"SUB_CATEGORY: {body.sub_category}",
        f"TYPE_ANGLE: {body.type_angle}",
        f"TARGET_LANGUAGE: {body.target_language}",
        f"AVATAR: {body.avatar_id}",
        f"HEADWEAR: {body.headwear_style}",
        f"CAMERA_STYLE: {body.camera_style}",
        f"SCENE_CONTEXT: {body.scene_context}",
        f"TRIGGER: {body.trigger_id}",
        f"SILO: {body.silo_id}",
        f"FORMULA: {body.submode_formula}",
        f"HOOK: {body.hook}",
        f"USP_1: {body.usp_1}",
        f"USP_2: {body.usp_2}",
        f"USP_3: {body.usp_3}",
        f"BODY: {body.body}",
        f"CTA: {body.cta}",
    ]
    return "\n".join(line for line in lines if not line.endswith(": "))


def _build_characters(body: BlueprintInput) -> list[BlueprintCharacter]:
    avatar_name = body.avatar_id.replace("_", " ").title()
    product_label = _clean_text(body.product_name)
    context_label = _clean_text(body.scene_context)
    characters = [
        BlueprintCharacter(
            name=avatar_name,
            entity_type="character",
            description=(
                f"Primary presenter for {product_label}. "
                f"Headwear: {body.headwear_style}. Camera style: {body.camera_style}. "
                f"Use tone appropriate for {body.silo_id} in {body.target_language}."
            ),
        ),
        BlueprintCharacter(
            name=product_label,
            entity_type="visual_asset",
            description=(
                f"Hero product asset. Category: {body.category}. Sub-category: {body.sub_category}. "
                f"Type angle: {body.type_angle}. Keep product appearance stable across every scene."
            ),
        ),
    ]
    if context_label:
        characters.append(
            BlueprintCharacter(
                name=f"{product_label} Scene Context",
                entity_type="location",
                description=f"Visual environment authority: {context_label}.",
            )
        )
    return characters


def _build_scenes(body: BlueprintInput) -> list[BlueprintScene]:
    sec = _seconds(body.duration_target)
    scene_total = _scene_count(body.duration_target)
    avatar_name = body.avatar_id.replace("_", " ").title()
    product_label = body.product_name
    context = body.scene_context or f"{body.category} product environment"
    benefits = [part for part in [body.usp_1, body.usp_2, body.usp_3] if part]
    while len(benefits) < 3:
        benefits.append(body.body or "Show practical use and clear product value.")

    base_scenes = [
        (
            f"{avatar_name} opens in {context}. Hook the viewer with {body.hook or 'a sharp opening angle'} while keeping {product_label} visible.",
            f"Cinematic opening frame in {context}. {avatar_name} introduces {product_label}. Focus on visual hook, clean composition, and immediate product readability.",
            _timed_video_prompt(
                f"{avatar_name} enters frame and establishes the problem space around {product_label}",
                f"camera movement reinforces the hook while product remains readable",
                f"end on a compelling visual beat that invites the next scene",
            ),
        ),
        (
            f"Demonstrate {product_label} in use inside {context}. Highlight {benefits[0]}.",
            f"Close product interaction shot inside {context}. Show practical usage of {product_label} with emphasis on {benefits[0]}.",
            _timed_video_prompt(
                f"{avatar_name} demonstrates {product_label} in a believable use case",
                f"show the first clear benefit: {benefits[0]}",
                "finish on a stable proof-oriented frame",
            ),
        ),
        (
            f"Escalate proof for {product_label}. Highlight {benefits[1]} and {benefits[2]}.",
            f"Proof-driven scene in {context}. Product hero framing plus tactile detail shots. Reinforce {benefits[1]} and {benefits[2]}.",
            _timed_video_prompt(
                "show layered proof and tactile close-ups",
                f"translate product value into visual confidence: {benefits[1]}",
                f"close with one more proof beat on {benefits[2]}",
            ),
        ),
        (
            f"Resolve with CTA for {product_label}. Use urgency and decisiveness without changing the scene authority from {context}.",
            f"Final action frame in {context}. Product remains central. End on a direct CTA visual for {product_label}.",
            _timed_video_prompt(
                "build final decision energy around the product",
                f"use body copy rhythm: {body.body or 'clarify value and remove hesitation'}",
                f"deliver CTA visually and verbally: {body.cta or 'act now'}",
            ),
        ),
    ]

    scenes = [
        BlueprintScene(
            display_order=index,
            prompt=prompt,
            image_prompt=image_prompt,
            video_prompt=video_prompt,
            character_names=[avatar_name, product_label],
            chain_type="ROOT" if index == 0 else "CONTINUATION",
        )
        for index, (prompt, image_prompt, video_prompt) in enumerate(base_scenes)
    ]

    if scene_total > 4:
        for extra_index in range(4, scene_total):
            scenes.append(
                BlueprintScene(
                    display_order=extra_index,
                    prompt=(
                        f"Extend the campaign with a variation scene for {product_label} in {context}. "
                        f"Preserve identity and re-emphasize {benefits[(extra_index - 4) % len(benefits)]}."
                    ),
                    image_prompt=(
                        f"Variation scene for {product_label} inside {context}. "
                        f"Maintain product identity, avatar continuity, and clean benefit framing."
                    ),
                    video_prompt=_timed_video_prompt(
                        "carry continuity from the previous scene without changing environment authority",
                        f"restate the strongest visual benefit: {benefits[(extra_index - 4) % len(benefits)]}",
                        "close on a high-clarity transition frame",
                    ),
                    character_names=[avatar_name, product_label],
                    chain_type="CONTINUATION",
                )
            )

    if sec <= 10:
        return scenes[:4]
    return scenes


@router.get("/content-pack", response_model=ContentPackSummary)
async def get_content_pack():
    return _content_pack_summary()


@router.post("/blueprint", response_model=BlueprintResponse)
async def build_blueprint(body: BlueprintInput):
    summary = _content_pack_summary()
    if body.engine_id not in summary.engines:
        raise HTTPException(400, f"Unsupported engine_id: {body.engine_id}")

    allowed_durations = summary.durations_by_engine.get(body.engine_id, [])
    if allowed_durations and body.duration_target not in allowed_durations:
        raise HTTPException(
            400,
            f"Duration {body.duration_target} is not allowed for {body.engine_id}. Allowed: {', '.join(allowed_durations)}",
        )

    characters = _build_characters(body)
    scenes = _build_scenes(body)
    project_name = f"{body.product_name} | {body.engine_id} | {body.duration_target}"

    return BlueprintResponse(
        project={
            "name": project_name,
            "description": f"{body.category} / {body.sub_category} / {body.type_angle}".strip(" /"),
            "story": _build_story(body),
            "language": "ms" if body.target_language.lower() == "malay" else "en",
            "material": body.material,
            "allow_music": False,
            "allow_voice": True,
            "characters": [item.model_dump() for item in characters],
        },
        video={
            "title": body.product_name,
            "description": f"{body.engine_id} {body.product_type} operator blueprint",
            "display_order": 0,
            "orientation": body.orientation,
        },
        scenes=scenes,
        notes=[
            "Blueprint compiled from the external BOSMAX content pack.",
            "Use Generate Images before Generate Ingredients or Frames. Generate Text to Video remains visible only as a not-wired label.",
            "Direct single-step T2V is not exposed as a native queue type in this repo. Verified operator paths are prompt -> image -> video or ingredients/refs -> video.",
        ],
    )


@router.get("/runtime-storage-status")
async def runtime_storage_status(include_authority_context_count: bool = False):
    """Operator-visible proof of WHICH storage the running backend is bound to.

    The audit found runtime DB binding unproven: the live API reported zero
    products while a repo-local ``flow_agent.db`` held hundreds. The cause is a
    process launched from a different worktree, so ``config.DB_PATH`` resolves to
    THAT worktree's database. This endpoint runs INSIDE the live process, so its
    counts come from the ACTUAL bound storage (``crud`` -> ``get_db`` -> DB_PATH).
    Compare ``effective_db_path`` against the checkout you expect the data in.

    Secrets are never exposed — only paths and row counts.
    """
    config_db_path = Path(str(agent_config.DB_PATH))
    try:
        effective_db_path = str(config_db_path.resolve())
    except Exception:
        effective_db_path = str(config_db_path)
    db_exists = False
    db_size_bytes: int | None = None
    try:
        db_exists = config_db_path.exists()
        if db_exists:
            db_size_bytes = config_db_path.stat().st_size
    except Exception:
        pass

    # Counts come from the LIVE bound connection — this is the binding proof.
    product_count: int | None = None
    manual_product_count: int | None = None
    queue_count: int | None = None
    population: dict[str, Any] | None = None
    read_error: str | None = None
    try:
        product_count = await crud.count_products()
        manual_product_count = await crud.count_products(source="MANUAL")
        queue_stats = await crud.get_bulk_queue_stats()
        queue_count = int(queue_stats.get("total", 0))
        population = await catalog_population_summary()
    except Exception as exc:  # pragma: no cover - defensive
        read_error = str(exc)

    # Raw product_count remains available for storage binding diagnostics, but it
    # is never labelled canonical. The population service excludes fixtures and
    # PI-13 aliases using the same reporting predicates as Command Centre KPIs.
    canonical_product_count = (population or {}).get("real_canonical_products")
    authority_context_count_ceiling = canonical_product_count
    authority_context_count: int | None = None
    authority_context_count_source = "NOT_COMPUTED"
    if include_authority_context_count:
        try:
            from agent.services.bosmax_authority_registry import (
                get_prompt_tool_context,
            )

            ctx = await get_prompt_tool_context()
            authority_context_count = len(ctx.product.contexts)
            authority_context_count_source = (
                "bosmax_authority_registry.get_prompt_tool_context"
            )
        except Exception as exc:  # pragma: no cover - defensive
            authority_context_count_source = f"UNAVAILABLE:{exc}"

    warnings: list[str] = []
    if read_error:
        warnings.append(f"STORAGE_READ_FAILED:{read_error}")
    if product_count == 0 and (queue_count or 0) > 0:
        warnings.append("ACTIVE_STORAGE_HAS_QUEUE_BUT_ZERO_PRODUCTS")
    if product_count == 0 and manual_product_count == 0:
        warnings.append("ACTIVE_STORAGE_HAS_ZERO_MANUAL_PRODUCTS")
    if os.environ.get("FLOW_AGENT_DIR"):
        warnings.append("FLOW_AGENT_DIR_OVERRIDE_ACTIVE")
    try:
        cwd = os.getcwd()
        if str(config_db_path.parent.resolve()) != str(Path(cwd).resolve()):
            warnings.append("DB_PATH_PARENT_DIFFERS_FROM_CWD")
    except Exception:
        cwd = ""
    canonical_state_root = runtime_release.canonical_state_root()
    state_root_canonical = config_db_path.parent.resolve() == canonical_state_root.resolve()
    if not state_root_canonical:
        warnings.append("CANONICAL_STATE_ROOT_MISMATCH")

    return {
        "status": "ok" if not read_error else "degraded",
        "cwd": cwd,
        "base_dir": str(agent_config.BASE_DIR),
        "flow_agent_dir_override": os.environ.get("FLOW_AGENT_DIR") or None,
        "config_db_path": str(agent_config.DB_PATH),
        "effective_db_path": effective_db_path,
        "canonical_state_root": str(canonical_state_root),
        "state_root_canonical": state_root_canonical,
        "db_exists": db_exists,
        "db_size_bytes": db_size_bytes,
        "product_count": product_count,
        "manual_product_count": manual_product_count,
        "queue_count": queue_count,
        "catalog_population": population,
        "raw_product_rows": (population or {}).get("raw_database_rows", product_count),
        "active_product_count": (population or {}).get("active_products"),
        "archived_product_count": (population or {}).get("archived_products"),
        "merged_alias_count": (population or {}).get("merged_historical_aliases"),
        "real_canonical_product_count": (population or {}).get("real_canonical_products"),
        "production_eligible_product_count": (population or {}).get("production_eligible_products"),
        "canonical_product_count": canonical_product_count,
        "authority_context_count_ceiling": authority_context_count_ceiling,
        "authority_context_count": authority_context_count,
        "authority_context_count_source": authority_context_count_source,
        "warnings": warnings,
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }
