"""One-time backfill: recover Kalodata `sold_count` on the FastMoss bulk queue.

Root cause (see kalodata_import_service.build_staged_record): the Kalodata
`Item Sold` value was parsed at import but stored ONLY under
`record["kalodata_meta"]["sold_count"]`. The queue writer
(fastmoss_bulk_promotion_service.sync_bulk_queue) reads the TOP-LEVEL
`ref.get("sold_count")`, so the value was dropped and 335/367 Kalodata rows
persisted `sold_count = NULL` -> the UI renders "—".

The forward fix lifts it to the top level for future syncs. Existing rows were
already synced (INSERT OR IGNORE skips them), so this script fills them in place
from two exact-match sources (never fuzzy, never guessed):

  A. STAGED CATALOG (precise): match `fastmoss_bulk_draft_status.reference_id`
     to the staged record `id`, read `kalodata_meta.sold_count`.
  B. SOURCE WORKBOOK (fallback for rows the current staged catalog no longer
     holds): the Kalodata sheets' `Item Sold` column, keyed by EXACT normalized
     product title. A title whose rows disagree on the value is AMBIGUOUS and is
     skipped — a wrong number is worse than "—".

Zero cost (no provider calls). Dry-run by default; `--apply` backs up the DB
first and only fills rows whose `sold_count` is currently NULL/empty
(idempotent, never overwrites a real value).

Usage:
    python scripts/backfill_kalodata_sold_count.py            # dry-run
    python scripts/backfill_kalodata_sold_count.py --apply    # write
"""
from __future__ import annotations

import argparse
import re
import shutil
import sqlite3
import sys
from datetime import datetime, timezone
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
DB_PATH = REPO / "flow_agent.db"
WORKBOOK = REPO / ".hermes" / "desktop-attachments" / "KALODATA & FASTMOSS-BONUS 600 DATA PRODUK-REV1.xlsx"
# Sheets that carry Kalodata units-sold (raw export + the merged view).
_KALODATA_SHEETS = ("300 PRODUK-KALODATA", "MERGED PRODUCTS")

if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))


def _norm(s: object) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip()).lower()


def _as_int(raw: object) -> int | None:
    try:
        return int(raw) if raw is not None and str(raw).strip() != "" else None
    except (TypeError, ValueError):
        return None


def _staged_sold_by_id() -> dict[str, int]:
    """id -> units sold, from the Kalodata staged catalog (kalodata_meta)."""
    from agent.services.kalodata_import_service import load_staged_catalog

    out: dict[str, int] = {}
    for record in load_staged_catalog():
        rid = str(record.get("id") or "").strip()
        if not rid:
            continue
        meta = record.get("kalodata_meta") or {}
        val = _as_int(meta.get("sold_count"))
        if val is None:
            val = _as_int(record.get("sold_count"))  # tolerate post-fix shape
        if val is not None:
            out[rid] = val
    return out


def _workbook_sold_by_title() -> dict[str, int]:
    """normalized product title -> units sold, from the Kalodata workbook sheets.
    Titles whose rows disagree on the value are dropped as AMBIGUOUS."""
    if not WORKBOOK.exists():
        print(f"(workbook fallback unavailable: {WORKBOOK.name} not found)")
        return {}
    try:
        import openpyxl
    except ImportError:
        print("(workbook fallback unavailable: openpyxl not installed)")
        return {}

    seen: dict[str, set[int]] = {}
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    for sheet in _KALODATA_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        it = wb[sheet].iter_rows(values_only=True)
        try:
            header = [str(c or "") for c in next(it)]
        except StopIteration:
            continue
        name_i = next((i for i, h in enumerate(header) if _norm(h) == "product name"), None)
        sold_i = next((i for i, h in enumerate(header) if h.strip().lower().startswith("item sold")), None)
        if name_i is None or sold_i is None:
            continue
        for row in it:
            if name_i >= len(row) or sold_i >= len(row):
                continue
            title = _norm(row[name_i])
            val = _as_int(row[sold_i])
            if not title or val is None:
                continue
            seen.setdefault(title, set()).add(val)
    ambiguous = sum(1 for v in seen.values() if len(v) > 1)
    if ambiguous:
        print(f"(workbook: {ambiguous} ambiguous titles skipped)")
    return {t: next(iter(v)) for t, v in seen.items() if len(v) == 1}


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true", help="write changes (default: dry-run)")
    args = ap.parse_args()

    try:  # Windows consoles default to cp1252; product titles carry CJK/emoji.
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    if not DB_PATH.exists():
        print(f"DB not found: {DB_PATH}")
        return 2

    staged = _staged_sold_by_id()
    wb_title = _workbook_sold_by_title()
    print(f"sources: staged-by-id={len(staged)}  workbook-by-title={len(wb_title)}")

    con = sqlite3.connect(str(DB_PATH))
    con.row_factory = sqlite3.Row
    rows = con.execute(
        "SELECT reference_id, raw_product_title FROM fastmoss_bulk_draft_status "
        "WHERE (sold_count IS NULL OR TRIM(CAST(sold_count AS TEXT)) = '') "
        "AND batch_provenance LIKE 'Kalodata%'"
    ).fetchall()
    print(f"Kalodata rows with empty sold_count: {len(rows)}")

    fill: list[tuple[int, str, str]] = []  # (sold_count, reference_id, source)
    still_missing = 0
    n_id, n_title = 0, 0
    for r in rows:
        rid = str(r["reference_id"] or "").strip()
        if rid in staged:
            fill.append((staged[rid], rid, "id"))
            n_id += 1
        else:
            val = wb_title.get(_norm(r["raw_product_title"]))
            if val is not None:
                fill.append((val, rid, "title"))
                n_title += 1
            else:
                still_missing += 1

    print(f"\nWILL FILL: {len(fill)}   (by reference_id={n_id}, by workbook title={n_title})")
    print(f"STILL MISSING (no exact source): {still_missing}")

    if not args.apply:
        print("\nDRY-RUN — no changes written. Re-run with --apply to persist.")
        con.close()
        return 0

    if fill:
        stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        backup = DB_PATH.with_suffix(f".db.presold-{stamp}")
        shutil.copy2(DB_PATH, backup)
        print(f"\nbackup: {backup.name}")
        con.executemany(
            "UPDATE fastmoss_bulk_draft_status SET sold_count = ? "
            "WHERE reference_id = ? AND (sold_count IS NULL OR TRIM(CAST(sold_count AS TEXT)) = '')",
            [(sold, rid) for sold, rid, _ in fill],
        )
        con.commit()
        print(f"APPLIED: {con.total_changes} rows updated.")
    else:
        print("\nnothing to fill.")
    con.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
