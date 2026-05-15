import os
from openpyxl import load_workbook
from agent.config import OPERATOR_PACK_DIR

def inspect_workbook():
    path = OPERATOR_PACK_DIR / "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx"
    if not path.exists():
        print(f"File not found: {path}")
        return
    
    wb = load_workbook(path, read_only=True, data_only=True)
    target_sheets = ["Product Sales Rank", "Most Promoted Products", "Video Product List", "Product Search Data", "New Products Ranking", "Copywriting_Product_Map"]
    
    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            print(f"Sheet {sheet_name} not found.")
            continue
            
        ws = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")
        found_header = False
        for row in ws.iter_rows(values_only=True, max_row=20):
            values = [str(v).strip() if v is not None else "" for v in row]
            normalized = {v.lower() for v in values if v}
            if "product name" in normalized or "product title" in normalized or "rank" in normalized:
                print(f"Headers: {values[:15]}")
                found_header = True
                break
        if not found_header:
            print("Could not find header row.")

if __name__ == "__main__":
    inspect_workbook()
