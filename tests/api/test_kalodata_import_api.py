"""Kalodata import API contract tests."""
import io

import openpyxl
from fastapi import FastAPI
from fastapi.testclient import TestClient

from agent.api.kalodata_import import router
from agent.models.kalodata_import import (
    CopyIntelligenceDryRunReport,
    KalodataImportReport,
)


def _build_app() -> FastAPI:
    app = FastAPI()
    app.include_router(router, prefix="/api")
    return app


def test_import_endpoint_returns_report(monkeypatch):
    captured = {}

    def fake_import(source_path, existing_tids=None):
        captured["source_path"] = source_path
        captured["existing_tids"] = existing_tids
        return KalodataImportReport(
            source_path=str(source_path), parsed_merged=4, parsed_hub=3,
            staged=3, skipped_duplicate_in_file=1, product_id_from_url=2,
            price_ranges_parsed=1, hub_matched=1, hub_unmatched_rows=[99],
            staged_catalog_path="x.json", staged_hub_path="y.json",
        )

    monkeypatch.setattr(
        "agent.services.kalodata_import_service.import_workbook", fake_import
    )
    async def fake_tids():
        return {"tid-existing"}
    monkeypatch.setattr(
        "agent.services.kalodata_import_service.collect_system_tids", fake_tids
    )
    client = TestClient(_build_app())
    response = client.post("/api/kalodata/import", json={"source_path": "C:/tmp/kalo.xlsx"})
    assert response.status_code == 200
    payload = response.json()
    assert payload["staged"] == 3
    assert payload["hub_matched"] == 1
    assert captured["source_path"] == "C:/tmp/kalo.xlsx"


def test_import_endpoint_uses_default_path(monkeypatch):
    captured = {}

    def fake_import(source_path, existing_tids=None):
        captured["source_path"] = source_path
        captured["existing_tids"] = existing_tids
        return KalodataImportReport(source_path=str(source_path))

    monkeypatch.setattr(
        "agent.services.kalodata_import_service.import_workbook", fake_import
    )
    async def fake_tids():
        return {"tid-existing"}
    monkeypatch.setattr(
        "agent.services.kalodata_import_service.collect_system_tids", fake_tids
    )
    client = TestClient(_build_app())
    response = client.post("/api/kalodata/import", json={})
    assert response.status_code == 200
    assert "Kalodata-BONUS 300 DATA PRODUK.xlsx" in captured["source_path"]


def test_import_endpoint_404_on_missing_workbook(monkeypatch):
    def fake_import(source_path, existing_tids=None):
        raise FileNotFoundError(source_path)

    monkeypatch.setattr(
        "agent.services.kalodata_import_service.import_workbook", fake_import
    )
    async def fake_tids():
        return {"tid-existing"}
    monkeypatch.setattr(
        "agent.services.kalodata_import_service.collect_system_tids", fake_tids
    )
    client = TestClient(_build_app())
    response = client.post("/api/kalodata/import", json={"source_path": "C:/no.xlsx"})
    assert response.status_code == 404
    assert "WORKBOOK_NOT_FOUND" in response.text


def test_copy_intelligence_dry_run_requires_explicit_source_path():
    client = TestClient(_build_app())
    response = client.post("/api/kalodata/copy-intelligence/dry-run", json={})
    assert response.status_code == 422
    assert response.json()["detail"] == "SOURCE_PATH_REQUIRED"


def test_copy_intelligence_workbook_upload_preserves_full_xlsx_and_required_sheets(
    tmp_path, monkeypatch
):
    monkeypatch.setattr(
        "agent.services.kalodata_import_service.COPY_INTELLIGENCE_WORKBOOKS_DIR",
        tmp_path,
    )
    workbook = openpyxl.Workbook()
    workbook.active.title = "MERGED PRODUCTS"
    workbook.create_sheet("COPYWRITING HUB")
    payload = io.BytesIO()
    workbook.save(payload)
    original = payload.getvalue()

    client = TestClient(_build_app())
    response = client.post(
        "/api/kalodata/copy-intelligence/workbooks",
        files={
            "workbook": (
                "Kalodata & Fastmoss 600.xlsx",
                original,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    report = response.json()
    assert report["original_filename"] == "Kalodata & Fastmoss 600.xlsx"
    assert report["source_id"] == report["fingerprint"]
    assert report["required_sheets"] == ["COPYWRITING HUB", "MERGED PRODUCTS"]
    assert report["sheet_names"] == ["MERGED PRODUCTS", "COPYWRITING HUB"]
    assert (tmp_path / f"{report['source_id']}.xlsx").read_bytes() == original


def test_copy_intelligence_workbook_upload_rejects_non_xlsx(tmp_path, monkeypatch):
    monkeypatch.setattr(
        "agent.services.kalodata_import_service.COPY_INTELLIGENCE_WORKBOOKS_DIR",
        tmp_path,
    )
    client = TestClient(_build_app())
    response = client.post(
        "/api/kalodata/copy-intelligence/workbooks",
        files={"workbook": ("copy.csv", b"not a workbook", "text/csv")},
    )

    assert response.status_code == 422
    assert response.json()["detail"].endswith("XLSX_REQUIRED")
    assert list(tmp_path.iterdir()) == []


def test_copy_intelligence_workbook_upload_requires_both_full_workbook_sheets(
    tmp_path, monkeypatch
):
    monkeypatch.setattr(
        "agent.services.kalodata_import_service.COPY_INTELLIGENCE_WORKBOOKS_DIR",
        tmp_path,
    )
    workbook = openpyxl.Workbook()
    workbook.active.title = "COPYWRITING HUB"
    payload = io.BytesIO()
    workbook.save(payload)

    client = TestClient(_build_app())
    response = client.post(
        "/api/kalodata/copy-intelligence/workbooks",
        files={"workbook": ("hub-only.xlsx", payload.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 422
    assert response.json()["detail"].endswith("MISSING_REQUIRED_SHEETS:MERGED PRODUCTS")
    assert list(tmp_path.iterdir()) == []


def test_uploaded_copy_intelligence_dry_run_uses_stored_source_only(tmp_path, monkeypatch):
    source_id = "a" * 64
    source_path = tmp_path / f"{source_id}.xlsx"
    source_path.write_bytes(b"full workbook bytes")
    monkeypatch.setattr(
        "agent.services.kalodata_import_service.COPY_INTELLIGENCE_WORKBOOKS_DIR",
        tmp_path,
    )
    captured = {}

    async def fake_dry_run(path):
        captured["path"] = path
        return CopyIntelligenceDryRunReport(source_workbook=path.name)

    monkeypatch.setattr(
        "agent.services.kalodata_import_service.build_copy_intelligence_dry_run_for_system",
        fake_dry_run,
    )
    client = TestClient(_build_app())
    response = client.post(
        "/api/kalodata/copy-intelligence/dry-run-upload",
        json={"source_id": source_id},
    )

    assert response.status_code == 200
    assert captured["path"] == source_path


def test_uploaded_copy_intelligence_dry_run_never_calls_seed_primitive(
    tmp_path, monkeypatch
):
    source_id = "b" * 64
    (tmp_path / f"{source_id}.xlsx").write_bytes(b"full workbook bytes")
    monkeypatch.setattr(
        "agent.services.kalodata_import_service.COPY_INTELLIGENCE_WORKBOOKS_DIR",
        tmp_path,
    )

    async def fake_dry_run(_path):
        return CopyIntelligenceDryRunReport(source_workbook="uploaded.xlsx")

    async def seed_must_not_run(_records):
        raise AssertionError("seed primitive must not run during dry-run")

    monkeypatch.setattr(
        "agent.services.kalodata_import_service.build_copy_intelligence_dry_run_for_system",
        fake_dry_run,
    )
    monkeypatch.setattr(
        "agent.services.kalodata_import_service.persist_copy_intelligence_seed_records",
        seed_must_not_run,
    )
    client = TestClient(_build_app())
    response = client.post(
        "/api/kalodata/copy-intelligence/dry-run-upload",
        json={"source_id": source_id},
    )

    assert response.status_code == 200


def test_copy_intelligence_seed_ledger_lists_review_rows_without_seed_write(monkeypatch):
    captured = {}

    async def fake_list(*, confidence=None, status=None, search=None, limit=100):
        captured.update(
            confidence=confidence, status=status, search=search, limit=limit
        )
        return {
            "total": 1,
            "items": [{
                "seed_id": "seed-1", "source_row": 12,
                "source_product_name": "Ledger product", "target_avatar": "Parents",
                "pain_point": "Time", "emotion_trigger": "Relief",
                "dream_outcome": "Easier routine", "key_ingredients_features": "Feature A",
                "hook_script": "Hook", "cta_script": "CTA", "confidence": "HIGH",
                "match_method": "TIKTOK_PRODUCT_ID_MATCH", "status": "NEEDS_REVIEW",
                "source_workbook": "seed.xlsx", "source_sheet": "COPYWRITING HUB",
                "provenance": {"source_row": "12"},
            }],
        }

    async def seed_must_not_run(_records):
        raise AssertionError("ledger read must not invoke the seed primitive")

    monkeypatch.setattr(
        "agent.services.kalodata_import_service.list_copy_intelligence_seed_records",
        fake_list,
    )
    monkeypatch.setattr(
        "agent.services.kalodata_import_service.persist_copy_intelligence_seed_records",
        seed_must_not_run,
    )
    client = TestClient(_build_app())
    response = client.get(
        "/api/kalodata/copy-intelligence/seeds",
        params={"confidence": "HIGH", "status": "NEEDS_REVIEW", "search": "Ledger", "limit": 25},
    )

    assert response.status_code == 200
    assert response.json()["items"][0]["source_product_name"] == "Ledger product"
    assert captured == {
        "confidence": "HIGH", "status": "NEEDS_REVIEW", "search": "Ledger", "limit": 25,
    }


def test_apply_hub_enrichment_delegates(monkeypatch):
    captured = {}

    async def fake_apply(reference_ids=None):
        captured["reference_ids"] = reference_ids
        return {"total": 2, "recomputed": 2, "skipped": 0, "failed": 0, "results": []}

    monkeypatch.setattr(
        "agent.services.kalodata_import_service.apply_hub_enrichment", fake_apply
    )
    client = TestClient(_build_app())
    response = client.post(
        "/api/kalodata/apply-hub-enrichment",
        json={"reference_ids": ["fastmoss-ref:abc123"]},
    )
    assert response.status_code == 200
    assert response.json()["recomputed"] == 2
    assert captured["reference_ids"] == ["fastmoss-ref:abc123"]


def test_hub_gaps_shape(monkeypatch):
    async def fake_gaps():
        return {"items": [], "totals": {"staged": 0, "fully_enriched": 0, "with_any_gap": 0}}

    monkeypatch.setattr("agent.services.kalodata_import_service.hub_gaps", fake_gaps)
    client = TestClient(_build_app())
    response = client.get("/api/kalodata/hub-gaps")
    assert response.status_code == 200
    assert response.json()["totals"]["staged"] == 0


def test_cache_images_bounds():
    client = TestClient(_build_app())
    assert client.post("/api/kalodata/cache-images", json={"product_ids": []}).status_code == 422
    assert (
        client.post(
            "/api/kalodata/cache-images",
            json={"product_ids": [f"p{i}" for i in range(26)]},
        ).status_code
        == 422
    )


def test_cache_images_sequential_with_results(monkeypatch):
    calls: list[str] = []

    async def fake_cache(product_id):
        calls.append(product_id)
        if product_id == "bad":
            return {"status": "failed", "detail": "404", "image_asset_status": "FAILED"}
        return {"status": "success", "local_image_path": f"/img/{product_id}.jpg",
                "image_asset_status": "READY"}

    monkeypatch.setattr("agent.api.products.cache_product_image", fake_cache)
    # retries should not sleep the suite
    import agent.api.kalodata_import as api_mod

    async def no_sleep(_seconds):
        return None

    monkeypatch.setattr(api_mod.asyncio, "sleep", no_sleep)

    client = TestClient(_build_app())
    response = client.post(
        "/api/kalodata/cache-images", json={"product_ids": ["p1", "bad", "p2"]}
    )
    assert response.status_code == 200
    payload = response.json()
    assert payload["cached"] == 2
    assert payload["failed"] == 1
    # 'bad' retried 3 times, successes once each
    assert calls.count("bad") == 3
    assert calls.count("p1") == calls.count("p2") == 1


# ── Copy Intelligence seed review/approval endpoints ─────────────────────────
def test_approve_seed_endpoint_routes_and_returns_result(monkeypatch):
    captured = {}

    async def fake_review(seed_id, *, action, reviewed_by, review_note, confirmation_phrase):
        captured.update(
            seed_id=seed_id, action=action, reviewed_by=reviewed_by,
            review_note=review_note, confirmation_phrase=confirmation_phrase,
        )
        return {
            "seed_id": seed_id, "previous_status": "NEEDS_REVIEW", "new_status": "APPROVED",
            "confidence": "HIGH", "reviewed_by": reviewed_by,
            "reviewed_at": "2026-07-15T00:00:00Z", "review_note": review_note,
        }

    monkeypatch.setattr(
        "agent.services.kalodata_import_service.review_copy_intelligence_seed", fake_review
    )
    client = TestClient(_build_app())
    response = client.post(
        "/api/kalodata/copy-intelligence/seeds/seed-1/approve",
        json={"reviewed_by": "owner", "review_note": "verified", "confirmation_phrase": "APPROVE COPY INTELLIGENCE"},
    )
    assert response.status_code == 200
    assert response.json()["new_status"] == "APPROVED"
    assert captured == {
        "seed_id": "seed-1", "action": "APPROVE", "reviewed_by": "owner",
        "review_note": "verified", "confirmation_phrase": "APPROVE COPY INTELLIGENCE",
    }


def test_reject_seed_endpoint_routes_with_reject_action(monkeypatch):
    captured = {}

    async def fake_review(seed_id, *, action, reviewed_by, review_note, confirmation_phrase):
        captured["action"] = action
        return {
            "seed_id": seed_id, "previous_status": "NEEDS_REVIEW", "new_status": "REJECTED",
            "confidence": "HIGH", "reviewed_by": reviewed_by,
            "reviewed_at": "2026-07-15T00:00:00Z", "review_note": review_note,
        }

    monkeypatch.setattr(
        "agent.services.kalodata_import_service.review_copy_intelligence_seed", fake_review
    )
    client = TestClient(_build_app())
    response = client.post(
        "/api/kalodata/copy-intelligence/seeds/seed-1/reject",
        json={"reviewed_by": "owner", "review_note": "wrong", "confirmation_phrase": "REJECT COPY INTELLIGENCE"},
    )
    assert response.status_code == 200
    assert response.json()["new_status"] == "REJECTED"
    assert captured["action"] == "REJECT"


def test_review_endpoint_maps_fail_closed_errors_to_status_codes(monkeypatch):
    from agent.services import kalodata_import_service as _svc

    scenarios = [
        ("COPY_INTELLIGENCE_SEED_NOT_FOUND", 404),
        ("INVALID_SOURCE_STATUS", 409),
        ("CONFIRMATION_PHRASE_MISMATCH", 422),
        ("MEDIUM_CONFIDENCE_PHRASE_REQUIRED", 422),
    ]
    for code, status_code in scenarios:
        async def fake_review(*args, _code=code, _sc=status_code, **kwargs):
            raise _svc.CopyIntelligenceReviewError(_code, _sc, {"seed_id": "seed-1"})

        monkeypatch.setattr(
            "agent.services.kalodata_import_service.review_copy_intelligence_seed", fake_review
        )
        client = TestClient(_build_app())
        response = client.post(
            "/api/kalodata/copy-intelligence/seeds/seed-1/approve",
            json={"reviewed_by": "owner", "review_note": "note", "confirmation_phrase": "x"},
        )
        assert response.status_code == status_code
        assert response.json()["detail"]["error"] == code


def test_review_endpoint_requires_body_fields():
    client = TestClient(_build_app())
    response = client.post(
        "/api/kalodata/copy-intelligence/seeds/seed-1/approve",
        json={"reviewed_by": "owner"},
    )
    assert response.status_code == 422


# ── Approved Copy Intelligence context endpoint ──────────────────────────────
def test_approved_context_endpoint_routes_read_only(monkeypatch):
    captured = {}

    async def fake_context(*, target_product_id=None, reference_id=None, seed_id=None, limit=100):
        captured.update(
            target_product_id=target_product_id, reference_id=reference_id,
            seed_id=seed_id, limit=limit,
        )
        return {
            "total": 1,
            "items": [{
                "seed_id": "seed-1", "source_product_name": "Approved product",
                "target_product_id": "prod-1", "reference_id": None,
                "target_avatar": "Parents", "pain_point": "Time",
                "emotion_trigger": "Relief", "dream_outcome": "Easier",
                "key_ingredients_features": "Feature", "hook_script": "Hook",
                "cta_script": "CTA", "confidence": "HIGH",
                "match_method": "TIKTOK_PRODUCT_ID_MATCH", "status": "APPROVED",
                "source_workbook": "seed.xlsx", "source_sheet": "COPYWRITING HUB",
                "provenance": {"source_row": "2"}, "reviewed_by": "owner",
                "reviewed_at": "2026-07-15T00:00:00Z", "review_note": "ok",
            }],
        }

    async def review_must_not_run(*args, **kwargs):
        raise AssertionError("approved-context read must not invoke review/transition")

    async def seed_must_not_run(_records):
        raise AssertionError("approved-context read must not invoke the seed primitive")

    monkeypatch.setattr(
        "agent.services.kalodata_import_service.get_approved_copy_intelligence_context", fake_context
    )
    monkeypatch.setattr(
        "agent.services.kalodata_import_service.review_copy_intelligence_seed", review_must_not_run
    )
    monkeypatch.setattr(
        "agent.services.kalodata_import_service.persist_copy_intelligence_seed_records", seed_must_not_run
    )
    client = TestClient(_build_app())
    response = client.get(
        "/api/kalodata/copy-intelligence/approved-context",
        params={"target_product_id": "prod-1", "reference_id": "ref-9", "seed_id": "seed-1", "limit": 50},
    )
    assert response.status_code == 200
    body = response.json()
    assert body["total"] == 1
    assert body["items"][0]["status"] == "APPROVED"
    assert captured == {
        "target_product_id": "prod-1", "reference_id": "ref-9", "seed_id": "seed-1", "limit": 50,
    }


def test_approved_context_endpoint_empty_ok(monkeypatch):
    async def fake_context(*, target_product_id=None, reference_id=None, seed_id=None, limit=100):
        return {"total": 0, "items": []}

    monkeypatch.setattr(
        "agent.services.kalodata_import_service.get_approved_copy_intelligence_context", fake_context
    )
    client = TestClient(_build_app())
    response = client.get("/api/kalodata/copy-intelligence/approved-context")
    assert response.status_code == 200
    assert response.json() == {"total": 0, "items": []}
