"""One-time backfill: populate `sell_price` + `commission_amount` on the FastMoss
bulk queue so the new Sell-price / Comm-amt columns show data for existing rows.

Both values were available upstream (the enriched reference carries `price`; the
queue row already stores `commission_rate`) but were dropped at the queue write
boundary. The forward fix reads them in `sync_bulk_queue`; existing rows were
already synced (INSERT OR IGNORE), so this fills them in place.

- `sell_price` (REAL, MYR): sourced by exact match only — reference row by
  `reference_id`, else the source workbook `Price` column by exact normalized
  title (ambiguous titles skipped).
- `commission_amount` (REAL, MYR): computed `sell_price * normalized_rate`, using
  the row's OWN stored `commission_rate` (so it is self-consistent with the
  displayed Sell price + Comm%). Uses the SAME helper as the writer, which
  normalizes both FastMoss percent strings ("5%") and Kalodata decimals ("0.05").

Zero cost. Dry-run by default; `--apply` backs up the DB first and only fills
rows whose `sell_price` is currently NULL (idempotent).

Usage:
    python scripts/backfill_fastmoss_price_comm.py            # dry-run
    python scripts/backfill_fastmoss_price_comm.py --apply    # write
"""
from __future__ import annotations

import argparse
import asyncio
import re
import shutil
import sqlite3
import sys
from datetime import datetime, timezone
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
DB_PATH = REPO / "flow_agent.db"
WORKBOOK = REPO / ".hermes" / "desktop-attachments" / "KALODATA & FASTMOSS-BONUS 600 DATA PRODUK-REV1.xlsx"
_PRICE_SHEETS = ("300 PRODUK-KALODATA", "300 PRODUK-FASTMOSS", "MERGED PRODUCTS")

if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))


def _norm(s: object) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip()).lower()


def _to_price(raw: object) -> float | None:
    """First numeric value in a price cell (handles 'RM0.89 - 1.99', '26.5')."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return round(float(raw), 2)
    m = re.search(r"\d+(?:\.\d+)?", str(raw).replace(",", ""))
    return round(float(m.group(0)), 2) if m else None


async def _ref_price_by_id() -> dict[str, float]:
    from agent.services.fastmoss_product_reference_service import list_fastmoss_reference_products

    out: dict[str, float] = {}
    for r in await list_fastmoss_reference_products(limit=2000):
        rid = str(r.get("id") or "").strip()
        p = _to_price(r.get("price"))
        if rid and p is not None:
            out[rid] = p
    return out


def _workbook_price_by_title() -> dict[str, float]:
    if not WORKBOOK.exists():
        print(f"(workbook fallback unavailable: {WORKBOOK.name} not found)")
        return {}
    try:
        import openpyxl
    except ImportError:
        print("(workbook fallback unavailable: openpyxl not installed)")
        return {}

    seen: dict[str, set[float]] = {}
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    for sheet in _PRICE_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        it = wb[sheet].iter_rows(values_only=True)
        try:
            header = [str(c or "") for c in next(it)]
        except StopIteration:
            continue
        name_i = next((i for i, h in enumerate(header) if _norm(h) == "product name"), None)
        price_i = next((i for i, h in enumerate(header)
                        if _norm(h).startswith("price") or "selling price" in _norm(h)), None)
        if name_i is None or price_i is None:
            continue
        for row in it:
            if name_i >= len(row) or price_i >= len(row):
                continue
            t = _norm(row[name_i])
            p = _to_price(row[price_i])
            if t and p is not None:
                seen.setdefault(t, set()).add(p)
    return {t: next(iter(v)) for t, v in seen.items() if len(v) == 1}


async def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true", help="write changes (default: dry-run)")
    args = ap.parse_args()

    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    if not DB_PATH.exists():
        print(f"DB not found: {DB_PATH}")
        return 2

    from agent.services.fastmoss_bulk_promotion_service import _queue_commission_amount

    ref_price = await _ref_price_by_id()
    wb_price = _workbook_price_by_title()
    print(f"price sources: reference-by-id={len(ref_price)}  workbook-by-title={len(wb_price)}")

    con = sqlite3.connect(str(DB_PATH))
    con.row_factory = sqlite3.Row
    rows = con.execute(
        "SELECT reference_id, raw_product_title, commission_rate FROM fastmoss_bulk_draft_status "
        "WHERE sell_price IS NULL"
    ).fetchall()
    print(f"rows missing sell_price: {len(rows)}")

    updates: list[tuple[float, float | None, str]] = []  # (sell_price, commission_amount, reference_id)
    n_id, n_title, no_price, comm_filled = 0, 0, 0, 0
    for r in rows:
        rid = str(r["reference_id"] or "").strip()
        price = ref_price.get(rid)
        if price is not None:
            n_id += 1
        else:
            price = wb_price.get(_norm(r["raw_product_title"]))
            if price is not None:
                n_title += 1
        if price is None:
            no_price += 1
            continue
        comm = _queue_commission_amount(price, r["commission_rate"])
        if comm is not None:
            comm_filled += 1
        updates.append((price, comm, rid))

    print(f"\nWILL FILL sell_price: {len(updates)} (by reference_id={n_id}, by workbook title={n_title})")
    print(f"  of those, commission_amount computed: {comm_filled}  (rest: rate missing/'-')")
    print(f"NO PRICE SOURCE (stay NULL): {no_price}")

    if not args.apply:
        print("\nDRY-RUN — no changes written. Re-run with --apply to persist.")
        con.close()
        return 0

    if updates:
        stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        backup = DB_PATH.with_suffix(f".db.preprice-{stamp}")
        shutil.copy2(DB_PATH, backup)
        print(f"\nbackup: {backup.name}")
        con.executemany(
            "UPDATE fastmoss_bulk_draft_status SET sell_price = ?, commission_amount = ? "
            "WHERE reference_id = ? AND sell_price IS NULL",
            updates,
        )
        con.commit()
        print(f"APPLIED: {con.total_changes} rows updated.")
    else:
        print("\nnothing to fill.")
    con.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(asyncio.run(main()))
