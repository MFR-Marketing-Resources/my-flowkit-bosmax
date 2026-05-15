import os
import json
import logging
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, List, Optional
from openpyxl import load_workbook
from agent.config import OPERATOR_PACK_DIR
from agent.services.product_mapping import normalize_mapping_text, resolve_product_mapping

logger = logging.getLogger(__name__)

# Status constants from contract
STATUS_SOURCE_ANCHOR_PRESENT = "SOURCE_ANCHOR_PRESENT"
STATUS_SOURCE_ANCHOR_PARTIAL = "SOURCE_ANCHOR_PARTIAL"
STATUS_SOURCE_ANCHOR_MISSING = "STATUS_SOURCE_ANCHOR_MISSING"
STATUS_SOURCE_ANCHOR_WEAK_FILE_HINT_ONLY = "SOURCE_ANCHOR_WEAK_FILE_HINT_ONLY"
STATUS_SOURCE_ANCHOR_COLUMN_NOT_FOUND = "SOURCE_ANCHOR_COLUMN_NOT_FOUND"
STATUS_SOURCE_ANCHOR_POTENTIALLY_CONTAMINATED = "SOURCE_ANCHOR_POTENTIALLY_CONTAMINATED"
STATUS_SOURCE_ANCHOR_KEYWORD_DERIVED = "SOURCE_ANCHOR_KEYWORD_DERIVED"
STATUS_SOURCE_ANCHOR_VERIFIED_FROM_RAW_SOURCE = "SOURCE_ANCHOR_VERIFIED_FROM_RAW_SOURCE"

# Reconciliation results
RECON_MATCHES_RAW_SOURCE = "MATCHES_RAW_SOURCE"
RECON_DIFFERS_FROM_RAW_SOURCE = "DIFFERS_FROM_RAW_SOURCE"
RECON_RAW_SOURCE_NOT_AVAILABLE = "RAW_SOURCE_NOT_AVAILABLE"
RECON_RAW_SOURCE_COLUMNS_MISSING = "RAW_SOURCE_COLUMNS_MISSING"
RECON_TITLE_KEYWORD_OVERRIDE_SUSPECTED = "TITLE_KEYWORD_OVERRIDE_SUSPECTED"
RECON_FILE_HINT_ONLY = "FILE_HINT_ONLY"

FASTMOSS_WORKBOOK_NAME = "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx"
FASTMOSS_TARGET_SHEET = "Copywriting_Product_Map"

class FastMossTaxonomyReconciliationService:
    @staticmethod
    @lru_cache(maxsize=1)
    def load_fastmoss_source_data() -> Dict[str, Dict[str, Any]]:
        """
        Loads the raw taxonomy from the FastMoss workbook.
        Maps normalized titles to raw source rows.
        """
        path = OPERATOR_PACK_DIR / FASTMOSS_WORKBOOK_NAME
        if not path.exists():
            logger.warning(f"FastMoss workbook not found at {path}")
            return {}

        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            if FASTMOSS_TARGET_SHEET not in wb.sheetnames:
                logger.warning(f"Sheet {FASTMOSS_TARGET_SHEET} not found in {FASTMOSS_WORKBOOK_NAME}")
                return {}

            ws = wb[FASTMOSS_TARGET_SHEET]
            headers: List[str] = []
            lookup: Dict[str, Dict[str, Any]] = {}

            for row in ws.iter_rows(values_only=True):
                values = list(row)
                if not any(v is not None and str(v).strip() for v in values):
                    continue
                
                # Identify header row
                if not headers and any(str(v).strip().lower() == "product name" for v in values if v):
                    headers = [str(v).strip() if v is not None else f"Col{i}" for i, v in enumerate(values)]
                    continue
                
                if not headers:
                    continue

                data = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
                name = str(data.get("Product Name") or "").strip()
                if not name:
                    continue

                # Use normalized title as key for cross-matching
                key = normalize_mapping_text(name)
                # Keep the first occurrence (usually the one with most info)
                if key not in lookup:
                    lookup[key] = data

            return lookup
        except Exception as e:
            logger.exception(f"Error loading FastMoss workbook: {e}")
            return {}

    @staticmethod
    def audit_fastmoss_product(product: Dict[str, Any]) -> Dict[str, Any]:
        """
        Audits a FastMoss product against raw source taxonomy.
        """
        source_data = FastMossTaxonomyReconciliationService.load_fastmoss_source_data()
        
        raw_title = product.get("raw_product_title") or product.get("product_display_name") or ""
        normalized_title = normalize_mapping_text(raw_title)
        
        source_row = source_data.get(normalized_title)
        
        db_category = (product.get("category") or "").strip()
        db_subcategory = (product.get("subcategory") or "").strip()
        db_type = (product.get("type") or "").strip()
        
        # Determine anchor status
        status = STATUS_SOURCE_ANCHOR_MISSING
        reconciliation = RECON_RAW_SOURCE_NOT_AVAILABLE
        origin = "DB_FALLBACK"
        notes = []
        discovered_columns = []

        if source_row:
            origin = f"WORKBOOK:{FASTMOSS_WORKBOOK_NAME}"
            src_category = str(source_row.get("Category") or "").strip()
            src_subcategory = str(source_row.get("Sub Category") or "").strip()
            src_type = str(source_row.get("Type / Product Angle") or "").strip()
            
            discovered_columns = [k for k, v in source_row.items() if v is not None]
            
            if src_category and src_subcategory:
                status = STATUS_SOURCE_ANCHOR_PRESENT
            elif src_category or src_subcategory:
                status = STATUS_SOURCE_ANCHOR_PARTIAL
            else:
                status = STATUS_SOURCE_ANCHOR_COLUMN_NOT_FOUND
            
            # Reconciliation
            if db_category == src_category and db_subcategory == src_subcategory:
                reconciliation = RECON_MATCHES_RAW_SOURCE
                status = STATUS_SOURCE_ANCHOR_VERIFIED_FROM_RAW_SOURCE
            else:
                reconciliation = RECON_DIFFERS_FROM_RAW_SOURCE
                notes.append(f"DB says '{db_category}/{db_subcategory}', Source says '{src_category}/{src_subcategory}'")

            # Check if DB values match what a KEYWORD resolve would produce
            # If DB differs from Source AND DB matches Keywords -> Suspicion of contamination
            keyword_mapping = resolve_product_mapping(product_name=raw_title, source_hint="FASTMOSS")
            kw_category = keyword_mapping.get("category")
            
            if db_category == kw_category and db_category != src_category:
                reconciliation = RECON_TITLE_KEYWORD_OVERRIDE_SUSPECTED
                status = STATUS_SOURCE_ANCHOR_KEYWORD_DERIVED
                notes.append("DB value matches keyword mapping but differs from raw source.")
        else:
            if product.get("fastmoss_source_file"):
                status = STATUS_SOURCE_ANCHOR_WEAK_FILE_HINT_ONLY
                reconciliation = RECON_FILE_HINT_ONLY
                origin = "FILE_HINT"

        return {
            "source_anchor_status": status,
            "source_anchor_origin": origin,
            "reconciliation_status": reconciliation,
            "discovered_columns": discovered_columns,
            "notes": notes,
            "raw_source_available": source_row is not None,
            "raw_values": {
                "category": source_row.get("Category") if source_row else None,
                "subcategory": source_row.get("Sub Category") if source_row else None,
                "type": source_row.get("Type / Product Angle") if source_row else None,
            } if source_row else None
        }

    @staticmethod
    async def perform_full_fastmoss_audit(limit: int = 20) -> Dict[str, Any]:
        """
        Runs the audit over all FastMoss products in the database.
        """
        from agent.db import crud
        fastmoss_products = await crud.list_products(source="FASTMOSS", limit=1000)
        
        results = []
        status_dist = {}
        recon_dist = {}
        contaminated_count = 0
        
        for p in fastmoss_products:
            audit = FastMossTaxonomyReconciliationService.audit_fastmoss_product(dict(p))
            results.append({
                "product_id": p.id,
                "title": p.raw_product_title,
                **audit
            })
            
            status = audit["source_anchor_status"]
            status_dist[status] = status_dist.get(status, 0) + 1
            
            recon = audit["reconciliation_status"]
            recon_dist[recon] = recon_dist.get(recon, 0) + 1
            
            if recon == RECON_TITLE_KEYWORD_OVERRIDE_SUSPECTED or status == STATUS_SOURCE_ANCHOR_POTENTIALLY_CONTAMINATED:
                contaminated_count += 1

        suspicious = [r for r in results if r["reconciliation_status"] == RECON_TITLE_KEYWORD_OVERRIDE_SUSPECTED][:10]

        return {
            "total_fastmoss_products": len(fastmoss_products),
            "source_anchor_status_distribution": status_dist,
            "reconciliation_status_distribution": recon_dist,
            "contaminated_or_keyword_derived_count": contaminated_count,
            "samples": results[:limit],
            "suspicious_examples": suspicious,
            "no_write_back": True
        }
