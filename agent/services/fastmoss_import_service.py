from __future__ import annotations

import csv
import io
import json
import re
from collections import Counter
from functools import lru_cache
from pathlib import Path
from typing import Any
from uuid import uuid4

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook

from agent.config import BASE_DIR
from agent.models.fastmoss_import import (
    FastMossImportBatchReport,
    FastMossImportFileReport,
    FastMossSalesMetricScopeEntry,
)


FASTMOSS_IMPORTS_DIR = BASE_DIR / "data" / "fastmoss" / "imports"
LATEST_BATCH_POINTER = FASTMOSS_IMPORTS_DIR / "latest.json"
EXPECTED_FILE_FIELD_ORDER = [
    "creator_search",
    "export_ad_list",
    "export_advertiser_list",
    "shop_list",
    "sales_rank",
    "new_products_ranking",
    "product_search_data",
    "product_search_sales_rank",
    "most_promoted_products_rank",
    "video_product_list",
]

FILE_TYPE_CONFIGS: dict[str, dict[str, Any]] = {
    "creator_search": {
        "file_type_id": "CREATOR_SEARCH",
        "label": "Creator Search",
        "filename_patterns": ["creator_search", "creator search"],
        "required_columns": ["Creator Name"],
        "optional_columns": ["Follower Count", "Shop Name", "Category"],
        "sample_columns": ["Creator Name", "Shop Name", "Category"],
    },
    "export_ad_list": {
        "file_type_id": "EXPORT_AD_LIST",
        "label": "Export Ad List",
        "filename_patterns": ["export ad list"],
        "required_columns": ["Product Name"],
        "optional_columns": ["Shop Name", "Orders", "Total Units Sold"],
        "sample_columns": ["Product Name", "Shop Name", "Orders"],
    },
    "export_advertiser_list": {
        "file_type_id": "EXPORT_ADVERTISER_LIST",
        "label": "Export Advertiser List",
        "filename_patterns": ["export advertiser list"],
        "required_columns": ["Shop Name"],
        "optional_columns": ["Total Sales Volume", "Shop Units Sold", "Orders"],
        "sample_columns": ["Shop Name", "Total Sales Volume"],
    },
    "shop_list": {
        "file_type_id": "SHOP_LIST",
        "label": "Shop List",
        "filename_patterns": ["shop list"],
        "required_columns": ["Shop Name"],
        "optional_columns": ["Shop Total Units Sold", "Total Sales Volume", "FastMoss Shop Detail"],
        "sample_columns": ["Shop Name", "Shop Total Units Sold"],
    },
    "sales_rank": {
        "file_type_id": "SALES_RANK",
        "label": "Sales Rank",
        "filename_patterns": ["sales rank"],
        "required_columns": ["Product Name"],
        "optional_columns": ["Shop Name", "Total Units Sold", "Shop Total Units Sold", "Orders", "FastMoss Product Detail", "TikTok Product Detail"],
        "sample_columns": ["Product Name", "Shop Name", "Total Units Sold", "Shop Total Units Sold"],
    },
    "new_products_ranking": {
        "file_type_id": "NEW_PRODUCTS_RANKING",
        "label": "New Products Ranking",
        "filename_patterns": ["new_products_ranking", "new products ranking"],
        "required_columns": ["Product Name"],
        "optional_columns": ["Shop", "Units Sold", "Shop Units Sold"],
        "sample_columns": ["Product Name", "Shop", "Units Sold"],
    },
    "product_search_data": {
        "file_type_id": "PRODUCT_SEARCH_DATA",
        "label": "Product Search Data",
        "filename_patterns": ["product search data"],
        "required_columns": ["Product Name"],
        "optional_columns": ["Store Name", "Total Sales Volume", "7-Day Sales Volume"],
        "sample_columns": ["Product Name", "Store Name", "Total Sales Volume"],
    },
    "product_search_sales_rank": {
        "file_type_id": "PRODUCT_SEARCH_SALES_RANK",
        "label": "Product Search Sales Rank",
        "filename_patterns": ["product_search_sales_rank", "product search sales rank"],
        "required_columns": ["Product Name"],
        "optional_columns": ["Shop", "Units Sold", "Shop Units Sold"],
        "sample_columns": ["Product Name", "Shop", "Units Sold", "Shop Units Sold"],
    },
    "most_promoted_products_rank": {
        "file_type_id": "MOST_PROMOTED_PRODUCTS_RANK",
        "label": "Most Promoted Products Rank",
        "filename_patterns": ["most promoted products", "tt_most_promoted_products_rank", "most_promoted_products_rank"],
        "required_columns": ["Product Name"],
        "optional_columns": ["Shop Name", "Total Units Sold", "Shop Units Sold", "FastMoss Product Detail"],
        "sample_columns": ["Product Name", "Shop Name", "Total Units Sold", "Shop Units Sold"],
    },
    "video_product_list": {
        "file_type_id": "VIDEO_PRODUCT_LIST",
        "label": "Video Product List",
        "filename_patterns": ["video_product_list", "video product list"],
        "required_columns": ["Product Title"],
        "optional_columns": ["Video Total Units Sold", "Video Units Sold", "FastMoss Product Detail Page Link"],
        "sample_columns": ["Product Title", "Video Total Units Sold", "Video Units Sold"],
    },
}
FILE_TYPE_ID_TO_KEY = {config["file_type_id"]: key for key, config in FILE_TYPE_CONFIGS.items()}


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _normalize_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", _normalize_header(value).lower()).strip("_")


def _guess_extension(filename: str) -> str:
    suffix = Path(filename).suffix.lower()
    return suffix if suffix else ""


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=True, indent=2), encoding="utf-8")


def _read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def detect_fastmoss_file_type(
    *,
    upload_field_key: str | None,
    filename: str | None,
    headers: list[str] | None = None,
    sheet_names: list[str] | None = None,
) -> tuple[str | None, str]:
    if upload_field_key and upload_field_key in FILE_TYPE_CONFIGS:
        return FILE_TYPE_CONFIGS[upload_field_key]["file_type_id"], "field_key"

    filename_key = _normalize_key(filename or "")
    for key, config in FILE_TYPE_CONFIGS.items():
        if any(pattern in filename_key for pattern in (_normalize_key(item) for item in config["filename_patterns"])):
            return config["file_type_id"], "filename_pattern"

    normalized_headers = {_normalize_key(item) for item in (headers or [])}
    sheet_key = " ".join(_normalize_key(name) for name in (sheet_names or []))
    best_type_id: str | None = None
    best_score = 0
    for key, config in FILE_TYPE_CONFIGS.items():
        required = {_normalize_key(item) for item in config["required_columns"]}
        optional = {_normalize_key(item) for item in config["optional_columns"]}
        score = len(required & normalized_headers) * 3 + len(optional & normalized_headers)
        if any(pattern in sheet_key for pattern in (_normalize_key(item) for item in config["filename_patterns"])):
            score += 2
        if score > best_score:
            best_score = score
            best_type_id = config["file_type_id"]
    if best_type_id:
        return best_type_id, "header_signature"
    return None, "unrecognized"


def classify_sales_metric_column(file_type_id: str, column_name: str) -> FastMossSalesMetricScopeEntry | None:
    normalized = _normalize_key(column_name)
    metric_name = "unknown_sold_metric"
    metric_scope = "UNKNOWN"
    truth_status = "NOT_VERIFIED"
    warning: str | None = "SALES_METRIC_SCOPE_NOT_VERIFIED"

    if "shop_total_units_sold" in normalized:
        metric_name = "shop_total_sold_count"
        metric_scope = "SHOP"
        truth_status = "SHOP_LEVEL_AGGREGATE"
        warning = "SHOP_LEVEL_METRIC_NOT_PRODUCT_SALES"
    elif "shop_units_sold" in normalized:
        metric_name = "shop_units_sold"
        metric_scope = "SHOP"
        truth_status = "SHOP_LEVEL_AGGREGATE"
        warning = "SHOP_LEVEL_METRIC_NOT_PRODUCT_SALES"
    elif normalized in {"total_units_sold", "units_sold", "video_total_units_sold", "video_units_sold"}:
        metric_name = "product_units_sold" if "video" not in normalized else "product_sold_count"
        metric_scope = "PRODUCT"
        truth_status = "VERIFIED_PRODUCT_LEVEL"
        warning = None
    elif normalized in {"orders", "order_count"}:
        metric_name = "order_count"
        metric_scope = "UNKNOWN"
        truth_status = "NOT_VERIFIED"
    elif "total_sales_volume" in normalized or "sales_volume" in normalized:
        metric_name = "total_sales_volume"
        metric_scope = "SHOP" if file_type_id in {"SHOP_LIST", "EXPORT_ADVERTISER_LIST"} else "UNKNOWN"
        truth_status = "SHOP_LEVEL_AGGREGATE" if metric_scope == "SHOP" else "NOT_VERIFIED"
        warning = "SHOP_LEVEL_METRIC_NOT_PRODUCT_SALES" if metric_scope == "SHOP" else "SALES_METRIC_SCOPE_NOT_VERIFIED"
    else:
        return None

    return FastMossSalesMetricScopeEntry(
        file_type_id=file_type_id,
        metric_name=metric_name,
        source_column=column_name,
        metric_scope=metric_scope,
        truth_status=truth_status,
        warning=warning,
    )


def _find_header_in_rows(rows: list[list[Any]], expected_columns: list[str]) -> tuple[int, list[str]]:
    expected = {_normalize_key(column) for column in expected_columns}
    best_index = -1
    best_headers: list[str] = []
    best_score = -1
    for index, row in enumerate(rows[:12]):
        headers = [_normalize_header(value) for value in row]
        normalized = {_normalize_key(header) for header in headers if header}
        score = len(expected & normalized)
        if score > best_score and normalized:
            best_score = score
            best_index = index
            best_headers = headers
    return best_index, best_headers


def _parse_csv_bytes(payload: bytes, config: dict[str, Any]) -> tuple[list[str], str, list[str], int, list[dict[str, Any]]]:
    text = payload.decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(text)))
    header_index, headers = _find_header_in_rows(rows, config["required_columns"] + config["optional_columns"])
    if header_index < 0 or not headers:
        raise ValueError("CSV header row not found")
    data_rows = rows[header_index + 1 :]
    row_count = sum(1 for row in data_rows if any(_normalize_header(cell) for cell in row))
    samples: list[dict[str, Any]] = []
    for row in data_rows:
        if not any(_normalize_header(cell) for cell in row):
            continue
        mapping = {
            headers[idx]: row[idx]
            for idx in range(min(len(headers), len(row)))
            if headers[idx]
        }
        sample = {column: mapping.get(column) for column in config["sample_columns"] if mapping.get(column) not in (None, "")}
        if sample:
            samples.append(sample)
        if len(samples) >= 5:
            break
    return ["CSV"], "CSV", headers, row_count, samples


def _parse_workbook_bytes(payload: bytes, config: dict[str, Any]) -> tuple[list[str], str, list[str], int, list[dict[str, Any]]]:
    workbook = load_workbook(filename=io.BytesIO(payload), read_only=True, data_only=True)
    sheet_names = list(workbook.sheetnames)
    best_sheet = sheet_names[0] if sheet_names else "UNKNOWN"
    best_headers: list[str] = []
    best_header_index = -1
    best_score = -1

    for sheet_name in sheet_names:
        ws = workbook[sheet_name]
        preview_rows = [list(row) for row in ws.iter_rows(values_only=True, max_row=12)]
        header_index, headers = _find_header_in_rows(preview_rows, config["required_columns"] + config["optional_columns"])
        normalized = {_normalize_key(item) for item in headers if item}
        score = len(normalized & {_normalize_key(item) for item in config["required_columns"]}) * 3
        score += len(normalized & {_normalize_key(item) for item in config["optional_columns"]})
        if score > best_score:
            best_score = score
            best_sheet = sheet_name
            best_headers = headers
            best_header_index = header_index

    if best_header_index < 0 or not best_headers:
        raise ValueError("Workbook header row not found")

    ws = workbook[best_sheet]
    row_count = 0
    samples: list[dict[str, Any]] = []
    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=0):
        values = list(row)
        if row_index <= best_header_index:
            continue
        if not any(_normalize_header(cell) for cell in values):
            continue
        row_count += 1
        mapping = {
            best_headers[idx]: values[idx]
            for idx in range(min(len(best_headers), len(values)))
            if best_headers[idx]
        }
        sample = {column: mapping.get(column) for column in config["sample_columns"] if mapping.get(column) not in (None, "")}
        if sample and len(samples) < 5:
            samples.append(sample)
    return sheet_names, best_sheet, best_headers, row_count, samples


def _build_column_validation(config: dict[str, Any], headers: list[str]) -> tuple[list[str], list[str], list[str], list[str]]:
    normalized_headers = {_normalize_key(header): header for header in headers if header}
    required_present = [column for column in config["required_columns"] if _normalize_key(column) in normalized_headers]
    optional_present = [column for column in config["optional_columns"] if _normalize_key(column) in normalized_headers]
    missing_required = [column for column in config["required_columns"] if _normalize_key(column) not in normalized_headers]
    known = {_normalize_key(column) for column in config["required_columns"] + config["optional_columns"]}
    unknown_columns = [header for header in headers if _normalize_key(header) and _normalize_key(header) not in known]
    return required_present, optional_present, missing_required, unknown_columns


def _parse_saved_file(*, storage_path: Path, config: dict[str, Any], upload_field_key: str, original_filename: str) -> FastMossImportFileReport:
    extension = _guess_extension(original_filename)
    report = FastMossImportFileReport(
        upload_field_key=upload_field_key,
        file_type_id=config["file_type_id"],
        label=config["label"],
        original_filename=original_filename,
        detected_by="field_key",
        storage_path=str(storage_path),
        extension=extension,
    )
    payload = storage_path.read_bytes()
    try:
        if extension in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            sheet_names, selected_sheet, headers, row_count, samples = _parse_workbook_bytes(payload, config)
        elif extension == ".csv":
            sheet_names, selected_sheet, headers, row_count, samples = _parse_csv_bytes(payload, config)
        elif extension == ".xls":
            report.parse_status = "UNSUPPORTED_FORMAT"
            report.parse_errors.append("Legacy .xls parsing is not supported in this checkout.")
            return report
        else:
            report.parse_status = "UNSUPPORTED_FORMAT"
            report.parse_errors.append(f"Unsupported file extension: {extension or 'NONE'}")
            return report

        report.sheet_names = sheet_names
        report.selected_sheet = selected_sheet
        report.headers = headers
        report.row_count = row_count
        report.sample_records = samples
        required_present, optional_present, missing_required, unknown_columns = _build_column_validation(config, headers)
        report.required_columns_present = required_present
        report.optional_columns_present = optional_present
        report.missing_required_columns = missing_required
        report.unknown_columns = unknown_columns
        report.sales_metric_scope_report = [
            metric
            for header in headers
            if (metric := classify_sales_metric_column(config["file_type_id"], header)) is not None
        ]
        if missing_required:
            report.parse_status = "COLUMN_VALIDATION_FAILED"
            report.parse_warnings.append("Missing required columns: " + ", ".join(missing_required))
        else:
            report.parse_status = "PARSED"
    except Exception as exc:  # pragma: no cover - safety wrapper
        report.parse_status = "PARSE_ERROR"
        report.parse_errors.append(str(exc))
    return report


def _materialize_report(*, batch_id: str, file_reports: list[FastMossImportFileReport], raw_dir: Path) -> FastMossImportBatchReport:
    recognized = [report.file_type_id for report in file_reports if report.file_type_id]
    duplicates = [file_type for file_type, count in Counter(recognized).items() if count > 1]
    missing = [
        FILE_TYPE_CONFIGS[field_key]["file_type_id"]
        for field_key in EXPECTED_FILE_FIELD_ORDER
        if FILE_TYPE_CONFIGS[field_key]["file_type_id"] not in recognized
    ]
    parse_warnings = [
        f"{report.file_type_id or report.upload_field_key}: {warning}"
        for report in file_reports
        for warning in report.parse_warnings
    ]
    parse_errors = [
        f"{report.file_type_id or report.upload_field_key}: {error}"
        for report in file_reports
        for error in report.parse_errors
    ]
    row_counts = {
        report.file_type_id: report.row_count
        for report in file_reports
        if report.file_type_id
    }
    column_validation = {
        report.file_type_id: {
            "required_columns_present": report.required_columns_present,
            "missing_required_columns": report.missing_required_columns,
            "optional_columns_present": report.optional_columns_present,
            "unknown_columns": report.unknown_columns,
            "parse_status": report.parse_status,
        }
        for report in file_reports
        if report.file_type_id
    }
    sales_metric_scope_report = [
        metric
        for report in file_reports
        for metric in report.sales_metric_scope_report
    ]
    product_reference_sample: list[dict[str, Any]] = []
    for report in file_reports:
        for sample in report.sample_records:
            enriched = {"file_type_id": report.file_type_id, **sample}
            product_reference_sample.append(enriched)
            if len(product_reference_sample) >= 12:
                break
        if len(product_reference_sample) >= 12:
            break

    has_errors = bool(parse_errors)
    ready_for_processing = bool(file_reports) and not has_errors
    import_status = "IMPORT_PREVIEW_READY" if ready_for_processing else "IMPORT_PREVIEW_WITH_ERRORS"
    return FastMossImportBatchReport(
        batch_id=batch_id,
        import_status=import_status,
        uploaded_files=len(file_reports),
        recognized_file_types=recognized,
        missing_expected_file_types=missing,
        duplicate_file_types=duplicates,
        row_counts_by_file_type=row_counts,
        column_validation_by_file_type=column_validation,
        sales_metric_scope_report=sales_metric_scope_report,
        product_reference_sample=product_reference_sample,
        parse_warnings=parse_warnings,
        parse_errors=parse_errors,
        ready_for_processing=ready_for_processing,
        raw_file_storage_path=str(raw_dir),
        provenance=[
            "fastmoss_import:read_only_preview",
            "fastmoss_import:latest_reference_only",
            "fastmoss_import:growth_analytics_disabled",
        ],
        files=file_reports,
    )


async def import_fastmoss_batch(files_by_field: dict[str, UploadFile | None]) -> dict[str, Any]:
    submitted = {
        field_key: upload
        for field_key, upload in files_by_field.items()
        if upload is not None and upload.filename
    }
    if not submitted:
        raise HTTPException(status_code=400, detail="No FastMoss files were uploaded.")

    batch_id = uuid4().hex
    raw_dir = FASTMOSS_IMPORTS_DIR / batch_id / "raw"
    raw_dir.mkdir(parents=True, exist_ok=True)

    file_reports: list[FastMossImportFileReport] = []
    for field_key in EXPECTED_FILE_FIELD_ORDER:
        upload = submitted.get(field_key)
        if not upload:
            continue
        config = FILE_TYPE_CONFIGS[field_key]
        original_filename = Path(upload.filename or f"{field_key}.bin").name
        extension = _guess_extension(original_filename)
        storage_path = raw_dir / f"{config['file_type_id'].lower()}{extension or '.bin'}"
        payload = await upload.read()
        storage_path.write_bytes(payload)
        file_reports.append(
            _parse_saved_file(
                storage_path=storage_path,
                config=config,
                upload_field_key=field_key,
                original_filename=original_filename,
            )
        )

    batch_report = _materialize_report(batch_id=batch_id, file_reports=file_reports, raw_dir=raw_dir)
    report_path = FASTMOSS_IMPORTS_DIR / batch_id / "report.json"
    _write_json(report_path, batch_report.model_dump(mode="json"))
    _write_json(LATEST_BATCH_POINTER, {"batch_id": batch_id, "report_path": str(report_path)})
    return batch_report.model_dump(mode="json")


def get_fastmoss_import_batch(batch_id: str) -> dict[str, Any]:
    report_path = FASTMOSS_IMPORTS_DIR / batch_id / "report.json"
    if not report_path.exists():
        raise HTTPException(status_code=404, detail="FastMoss import batch not found")
    return _read_json(report_path)


def get_latest_fastmoss_import_batch() -> dict[str, Any]:
    if not LATEST_BATCH_POINTER.exists():
        raise HTTPException(status_code=404, detail="No latest FastMoss import batch is available")
    pointer = _read_json(LATEST_BATCH_POINTER)
    report_path = Path(pointer["report_path"])
    if not report_path.exists():
        raise HTTPException(status_code=404, detail="Latest FastMoss import report is missing")
    return _read_json(report_path)


FASTMOSS_REFERENCE_COLUMN_HINTS: dict[str, dict[str, list[str]]] = {
    "EXPORT_AD_LIST": {
        "name_columns": ["Product Name", "Description"],
        "shop_columns": ["Shop Name", "Handle"],
        "source_url_columns": ["FastMoss Product Detail", "Homepage", "Video Url"],
        "tiktok_url_columns": ["TikTok Product Detail"],
    },
    "EXPORT_ADVERTISER_LIST": {
        "name_columns": ["Name"],
        "shop_columns": ["Shop Name", "Name"],
        "source_url_columns": [],
        "tiktok_url_columns": [],
    },
    "SHOP_LIST": {
        "name_columns": [],
        "shop_columns": ["Shop Name", "Company Name"],
        "source_url_columns": ["FastMoss Shop Detail", "FastMoss Shop Detail Page Link"],
        "tiktok_url_columns": [],
    },
    "SALES_RANK": {
        "name_columns": ["Product Name"],
        "shop_columns": ["Shop Name", "Company Name"],
        "source_url_columns": ["FastMoss Product Detail", "FastMoss Shop Detail", "FastMoss Shop Detail Page Link"],
        "tiktok_url_columns": ["TikTok Product Detail"],
    },
    "NEW_PRODUCTS_RANKING": {
        "name_columns": ["Product Name"],
        "shop_columns": ["Shop"],
        "source_url_columns": ["FastMoss Product Detail", "FastMoss Product Detail Page Link"],
        "tiktok_url_columns": ["TikTok Product Detail"],
    },
    "PRODUCT_SEARCH_DATA": {
        "name_columns": ["Product Name"],
        "shop_columns": ["Store Name"],
        "source_url_columns": ["FastMoss", "FastMoss Product Detail", "FastMoss Shop"],
        "tiktok_url_columns": ["TikTok", "TikTok Product Detail"],
    },
    "PRODUCT_SEARCH_SALES_RANK": {
        "name_columns": ["Product Name"],
        "shop_columns": ["Shop Name", "Shop"],
        "source_url_columns": ["FastMoss Product Detail", "FastMoss Shop Detail"],
        "tiktok_url_columns": ["TikTok Product Detail"],
    },
    "MOST_PROMOTED_PRODUCTS_RANK": {
        "name_columns": ["Product Name"],
        "shop_columns": ["Shop Name"],
        "source_url_columns": ["FastMoss Product Detail", "FastMoss Shop Detail"],
        "tiktok_url_columns": ["TikTok Product Detail"],
    },
    "VIDEO_PRODUCT_LIST": {
        "name_columns": ["Product Title"],
        "shop_columns": [],
        "source_url_columns": ["FastMoss Product Detail Page Link", "Video Link"],
        "tiktok_url_columns": ["TikTok Product Link"],
    },
}


def _to_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    text = _normalize_header(value).replace(",", "")
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _iter_saved_report_rows(*, storage_path: Path, extension: str, selected_sheet: str, headers: list[str]) -> list[dict[str, Any]]:
    normalized_headers = [_normalize_header(header) for header in headers if _normalize_header(header)]
    if not normalized_headers:
        return []

    def _map_rows(rows: list[list[Any]]) -> list[dict[str, Any]]:
        header_index = -1
        expected = [_normalize_key(header) for header in normalized_headers]
        for index, row in enumerate(rows[:20]):
            candidate = [_normalize_key(value) for value in row[: len(normalized_headers)]]
            if candidate == expected:
                header_index = index
                break
        if header_index < 0:
            return []
        mapped: list[dict[str, Any]] = []
        for row in rows[header_index + 1 :]:
            if not any(_normalize_header(cell) for cell in row):
                continue
            mapped.append(
                {
                    normalized_headers[idx]: row[idx]
                    for idx in range(min(len(normalized_headers), len(row)))
                    if normalized_headers[idx]
                }
            )
        return mapped

    if extension in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        workbook = load_workbook(filename=storage_path, read_only=True, data_only=True)
        if selected_sheet not in workbook.sheetnames:
            return []
        rows = [list(row) for row in workbook[selected_sheet].iter_rows(values_only=True)]
        return _map_rows(rows)

    if extension == ".csv":
        text = storage_path.read_text(encoding="utf-8-sig")
        rows = list(csv.reader(io.StringIO(text)))
        return _map_rows(rows)

    return []


def _non_empty_values(row: dict[str, Any], columns: list[str]) -> list[str]:
    values: list[str] = []
    seen: set[str] = set()
    normalized_row = {_normalize_key(key): value for key, value in row.items()}
    for column in columns:
        value = normalized_row.get(_normalize_key(column))
        text = _normalize_header(value)
        if not text or text in seen:
            continue
        seen.add(text)
        values.append(text)
    return values


@lru_cache(maxsize=4)
def _latest_reference_index(batch_id: str, report_mtime: float) -> dict[str, Any]:
    report_path = FASTMOSS_IMPORTS_DIR / batch_id / "report.json"
    report = _read_json(report_path)
    records: list[dict[str, Any]] = []

    for file_report in report.get("files", []):
        file_type_id = str(file_report.get("file_type_id") or "").strip()
        storage_path = Path(str(file_report.get("storage_path") or ""))
        if not file_type_id or not storage_path.exists():
            continue
        if file_report.get("parse_errors"):
            continue
        column_hints = FASTMOSS_REFERENCE_COLUMN_HINTS.get(file_type_id, {})
        metric_columns = file_report.get("sales_metric_scope_report") or []
        if not metric_columns:
            continue

        rows = _iter_saved_report_rows(
            storage_path=storage_path,
            extension=str(file_report.get("extension") or "").lower(),
            selected_sheet=str(file_report.get("selected_sheet") or ""),
            headers=list(file_report.get("headers") or []),
        )
        for row in rows:
            metric_values = []
            for metric in metric_columns:
                source_column = str(metric.get("source_column") or "")
                value = _to_int(row.get(source_column))
                if value is None:
                    continue
                metric_values.append(
                    {
                        "metric_name": metric.get("metric_name"),
                        "source_column": source_column,
                        "metric_scope": metric.get("metric_scope"),
                        "truth_status": metric.get("truth_status"),
                        "warning": metric.get("warning"),
                        "value": value,
                    }
                )
            if not metric_values:
                continue

            names = _non_empty_values(row, column_hints.get("name_columns", []))
            source_urls = _non_empty_values(row, column_hints.get("source_url_columns", []))
            tiktok_urls = _non_empty_values(row, column_hints.get("tiktok_url_columns", []))
            shop_names = _non_empty_values(row, column_hints.get("shop_columns", []))
            if not names and not source_urls and not tiktok_urls:
                continue

            records.append(
                {
                    "batch_id": batch_id,
                    "file_type_id": file_type_id,
                    "detected_by": file_report.get("detected_by") or "field_key",
                    "names": names,
                    "shop_names": shop_names,
                    "source_urls": source_urls,
                    "tiktok_urls": tiktok_urls,
                    "metric_values": metric_values,
                }
            )

    by_name: dict[str, list[dict[str, Any]]] = {}
    by_source_url: dict[str, list[dict[str, Any]]] = {}
    by_tiktok_url: dict[str, list[dict[str, Any]]] = {}
    for record in records:
        for name in record["names"]:
            by_name.setdefault(_normalize_key(name), []).append(record)
        for url in record["source_urls"]:
            by_source_url.setdefault(_normalize_header(url), []).append(record)
        for url in record["tiktok_urls"]:
            by_tiktok_url.setdefault(_normalize_header(url), []).append(record)
    return {
        "batch_id": batch_id,
        "report": report,
        "records": records,
        "by_name": by_name,
        "by_source_url": by_source_url,
        "by_tiktok_url": by_tiktok_url,
    }


def get_latest_fastmoss_reference_index() -> dict[str, Any] | None:
    if not LATEST_BATCH_POINTER.exists():
        return None
    pointer = _read_json(LATEST_BATCH_POINTER)
    batch_id = str(pointer.get("batch_id") or "").strip()
    report_path = FASTMOSS_IMPORTS_DIR / batch_id / "report.json"
    if not batch_id or not report_path.exists():
        return None
    return _latest_reference_index(batch_id, report_path.stat().st_mtime)
