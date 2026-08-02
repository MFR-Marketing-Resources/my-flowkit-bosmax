#!/usr/bin/env python
"""PI-13 Master Product Recovery workbook builder — identity-locked, evidence-graded, 6 sheets."""
import sqlite3, json, hashlib, collections, datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[2]
TS = datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
c = sqlite3.connect(f"file:{(REPO/'flow_agent.db').as_posix()}?mode=ro", uri=True); c.row_factory = sqlite3.Row
R = json.load(open(REPO / "outputs/mission-pi12/residual_reasons.json", encoding="utf-8"))


def prod(pid):
    return dict(c.execute("SELECT product_display_name,raw_product_title,brand,shop_name,category,source_url,tiktok_product_url FROM product WHERE id=?", (pid,)).fetchone())


wb = openpyxl.Workbook()
HFILL = PatternFill("solid", fgColor="1F4E78"); HFONT = Font(color="FFFFFF", bold=True, size=10)


def sheet(name, head, widths):
    ws = wb.create_sheet(name); ws.append(head)
    for cell in ws[1]:
        cell.fill = HFILL; cell.font = HFONT; cell.alignment = Alignment(wrap_text=True, vertical="top")
    for i, h in enumerate(head, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 14)
    ws.freeze_panes = "B2"; return ws


wb.remove(wb.active)

# --- Sheet 1: Recovery Master (identity lock, full DB title) ---
M = sheet("Recovery Master",
          ["product_id", "full_db_title", "brand", "shop_name", "category", "debt_type", "missing_fields",
           "stored_source_url", "stored_tiktok_url", "identity_confidence", "recovery_status", "suggested_next_action"],
          {"product_id": 38, "full_db_title": 60, "brand": 14, "shop_name": 16, "category": 18, "debt_type": 20,
           "missing_fields": 26, "stored_source_url": 44, "stored_tiktok_url": 42, "identity_confidence": 16,
           "recovery_status": 30, "suggested_next_action": 40})
cohort = []
for x in R["incomplete"]:
    pid = x["product_id"]; p = prod(pid); cohort.append(pid)
    M.append([pid, p["raw_product_title"] or p["product_display_name"], p["brand"] or "", p["shop_name"] or "",
              p["category"], "INCOMPLETE", (x["reason"] or "").replace("MISSING_REQUIRED_FIELDS:", ""),
              p["source_url"] or "", p["tiktok_product_url"] or "", "LOCKED(DB)", "PENDING_RECOVERY",
              "Exhaustive 21-source sweep; capture excerpt+risk per field before any write"])
for x in R["review"]:
    pid = x["product_id"]; p = prod(pid); cohort.append(pid); tok = (x.get("claim_tokens") or [None])[0]
    M.append([pid, p["raw_product_title"] or p["product_display_name"], p["brand"] or "", p["shop_name"] or "",
              p["category"], x["reason"], f"claim_token: {tok}", p["source_url"] or "", p["tiktok_product_url"] or "",
              "LOCKED(DB)", "CLAIM_ADJUDICATION", "See Claim Review sheet"])

# --- Sheet 2: Source Evidence (per-field capture; corrected proof-of-concept) ---
E = sheet("Source Evidence",
          ["product_id", "field", "recovered_value", "source_url", "source_type", "source_excerpt",
           "retrieval_ts", "identity_match", "claim_risk", "confidence", "review_status"],
          {"product_id": 38, "field": 18, "recovered_value": 50, "source_url": 48, "source_type": 20,
           "source_excerpt": 54, "retrieval_ts": 20, "identity_match": 18, "claim_risk": 16, "confidence": 12, "review_status": 20})
seed = [
    ["01f3274b-2fc0-4977-b6d5-efaaa2cb2bf7", "usage_text", "Take 1 chewable tablet daily",
     "https://pentavite.com.my/collections/vitamins-minerals/products/mens-multi", "official_brand_store",
     "Men's Multi chewable tablet; 20 tailored nutrients (10 vit/min + 9 fruit&veg + Taurine)", TS,
     "HIGH (brand + 20-nutrient men-multi match)", "LOW", "MEDIUM", "UNVERIFIED_PENDING_EXACT_PAGE"],
    ["01f3274b-2fc0-4977-b6d5-efaaa2cb2bf7", "benefits_json", "Energy support; general health maintenance",
     "https://pentavite.com.my/collections/vitamins-minerals/products/mens-multi", "official_brand_store",
     "B-complex + Taurine for energy; Zinc/Vit C/D to maintain health (official)", TS, "HIGH",
     "MED (health-benefit wording)", "MEDIUM", "UNVERIFIED_PENDING_EXACT_PAGE"],
    ["0eed93fb-504e-44f4-bfe6-fcd620ae2d81", "identity", "MAVERIX MaxOil LIMITED PREMIUM Set (5 Bottles)",
     "https://www.lazada.com.my/products/maverix-maxoil-limited-premium-set-5-bottles-i14857278616.html",
     "authorised_marketplace", "Lazada title matches; variants COMBO5/SINGLE/DOUBLE/TRIPPLE STAR", TS,
     "HIGH (name + set match)", "N/A", "HIGH", "IDENTITY_VERIFIED"],
    ["0eed93fb-504e-44f4-bfe6-fcd620ae2d81", "usp_json", "(seller claim) tanpa-bahan-kimia / organic — NOT approved",
     "https://www.lazada.com.my/products/maverix-maxoil-limited-premium-set-5-bottles-i14857278616.html",
     "seller_listing", "seller phrase only; 'organic' != 'non-chemical'; needs exact excerpt + risk review", TS,
     "HIGH", "HIGH (organic/chemical)", "LOW", "REJECTED_needs_stronger_evidence"],
    ["0f6bdfdd-474d-4a30-aa92-400aa0f1ed51", "identity", "ANAS Powder Blusher DUO (brand OK; shade #06 NOT matched)",
     "https://www.lazada.com.my/products/new-2026-anas-powder-blusher-duo-01-love-rosie--peach-please-i14842797624.html",
     "authorised_marketplace", "found #01 Love Rosie/Peach Please; residual is #06 -> variant mismatch", TS,
     "PARTIAL (brand yes, shade #06 no)", "N/A", "LOW", "IDENTITY_PARTIAL_need_shade06"],
    ["0f6bdfdd-474d-4a30-aa92-400aa0f1ed51", "benefits_json", "(snippet) silky/buildable/long-lasting — NOT approved",
     "(search snippet)", "search_snippet", "snippet-level; not exact product page for #06", TS, "PARTIAL", "MED",
     "LOW", "REJECTED_snippet_only"],
    ["0ab23ee4-6947-4966-96b1-906a3cdc98e6", "identity", "Yaya Empire Vanilla Crush 10ml (Ara Johari/Ervan collab)",
     "https://www.lazada.com.my/products/yaya-empire-vanilla-crush-ara-johari-mia-azahar-ervan-best-selling-perfume-10-ml-i14890549810.html",
     "authorised_marketplace", "Lazada title matches name + 10ml + collaborators", TS, "HIGH (name+size)", "N/A",
     "HIGH", "IDENTITY_VERIFIED"],
    ["0ab23ee4-6947-4966-96b1-906a3cdc98e6", "usp_json", "(snippet) sweet vanilla candy-fruity — NOT approved",
     "(search snippet)", "search_snippet", "scent from snippet, not authoritative page", TS, "HIGH", "LOW", "LOW",
     "REJECTED_snippet_only"],
]
for r in seed:
    E.append(r)

# --- Sheet 3: Claim Review (11) ---
CR = sheet("Claim Review",
           ["product_id", "product", "category", "claim_token", "where_found", "adjudication", "decision_required", "status"],
           {"product_id": 38, "product": 36, "category": 16, "claim_token": 16, "where_found": 40, "adjudication": 46,
            "decision_required": 40, "status": 22})
claim_map = {
    "012e9fd6-dea4-4ebc-a4ca-27669b2eaccb": ("eye treatment / Eye Treatments (category)", "cosmetic category term, not medical"),
    "04a42d8c-755e-409e-8b60-b8b91f599dc7": ("Anti-bacterial treatment (fabric)", "fabric finish, not medical"),
    "3b56ffdb-e703-4af2-bfc9-3197bc7b14c9": ("cereal-based treats / a treat", "snack, not medical"),
    "49677116-dfc6-42da-ac7d-a25a49167c16": ("lip treatment / Lip Treatments", "cosmetic category term"),
    "9ef257e3-00b1-4f50-b7d9-419b8b2403a7": ("lip treatment", "cosmetic category term"),
    "db2dbbeb-79dc-4b78-b1ce-2257257cb7f8": ("lip treatment", "cosmetic category term"),
    "a0be58cd-9bdb-4f1f-a7dc-91b845e2e591": ("popcorn treat / crunchy treat", "snack, not medical"),
    "ae47b55b-58d4-441e-97d3-0d6c785bb530": ("targeted treatment (pest)", "pest control, not medical"),
    "e06d8afd-44fa-47d7-8d79-d7dc55d52c41": ("leave in as a treatment / Hair Treatments", "cosmetic haircare"),
    "efcc1f31-c91b-46fb-acc6-d736a5271884": ("conditioner or treatment", "cosmetic haircare"),
    "9311100f-6271-4965-8b26-f6e04e3843a1": ("anti-inflammatory in benefits/usp/description", "REAL efficacy claim, AI-generated, no acquired evidence"),
}
for x in R["review"]:
    pid = x["product_id"]; p = prod(pid); tok = (x.get("claim_tokens") or [None])[0]
    where, adj = claim_map.get(pid, ("", ""))
    if tok == "treat":
        CR.append([pid, (p["product_display_name"] or "")[:36], p["category"], tok, where, "FALSE POSITIVE — " + adj,
                   "Owner records benign-context decision (per-product; no global lexicon change)", "AWAIT_OWNER_DECISION"])
    else:
        CR.append([pid, (p["product_display_name"] or "")[:36], p["category"], tok, where, "REAL unsupported claim — " + adj,
                   "Owner: acquire evidence+approve OR remove claim + keep blocked", "AWAIT_OWNER_DECISION"])

# --- Sheets 4-6 ---
X = sheet("Exceptions", ["product_id", "product", "reason", "sources_tried", "exhaustion_proof"],
          {"product_id": 38, "product": 40, "reason": 30, "sources_tried": 50, "exhaustion_proof": 40})
S = wb.create_sheet("Summary")
byf = collections.Counter((x["reason"] or "").replace("MISSING_REQUIRED_FIELDS:", "") for x in R["incomplete"])
S.append(["PI-13 Master Product Recovery — identity-locked, evidence-graded"]); S["A1"].font = Font(bold=True, size=13)
S.append(["built_utc", TS]); S.append(["cohort_total", len(cohort)]); S.append(["unique_ids", len(set(cohort))])
S.append(["incomplete", len(R["incomplete"])]); S.append(["claim", len(R["review"])])
S.append([]); S.append(["Missing-field breakdown (incomplete)"])
for k, v in byf.most_common():
    S.append(["  " + k, v])
S.append([]); S.append(["Recovery status", "all PENDING — NO DB writes until identity-lock + workbook QA + 10-pilot pass"])
S.column_dimensions["A"].width = 64; S.column_dimensions["B"].width = 30
D = sheet("Deletion Candidates", ["product_id", "product", "sources_exhausted", "exhaustion_proof", "recommendation"],
          {"product_id": 38, "product": 40, "sources_exhausted": 50, "exhaustion_proof": 40, "recommendation": 30})
D.append(["-", "EMPTY BY DESIGN", "-", "populated only after ALL 21 sources tried with proof", "never delete to lower a count"])

path = REPO / "outputs/mission-pi12/Master_Product_Recovery.xlsx"
wb.save(path)
sha = hashlib.sha256(open(path, "rb").read()).hexdigest()
(REPO / "outputs/mission-pi12/Master_Product_Recovery.sha256").write_text(sha + "  Master_Product_Recovery.xlsx\n")
print("saved", path)
print("SHA-256:", sha)
print("cohort:", len(cohort), "unique:", len(set(cohort)), "sheets:", wb.sheetnames)
